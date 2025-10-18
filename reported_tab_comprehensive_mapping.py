#!/usr/bin/env python3
"""
Comprehensive Field Mapping for Reported Tab

This script creates a detailed mapping of all fields in the Reported tab
of the target file with full hierarchical scoping and enhanced context information.

Author: AI Assistant
Date: October 2025
"""

import openpyxl
import csv
from pathlib import Path
import re
from typing import Dict, List, Tuple, Optional


def analyze_reported_tab_structure(target_file: str) -> Tuple[List[Dict], List[Dict]]:
    """
    Analyze the Reported tab and create comprehensive field mappings.
    """
    print("Loading target workbook and analyzing Reported tab...")
    wb = openpyxl.load_workbook(target_file, data_only=True)
    
    if 'Reported' not in wb.sheetnames:
        print(f"ERROR: Reported tab not found. Available sheets: {wb.sheetnames}")
        wb.close()
        return [], []
    
    sheet = wb['Reported']
    print(f"Reported tab has {sheet.max_row} rows and {sheet.max_column} columns")
    
    # Get all column headers (quarters/periods) from row 5
    header_row = 5
    column_headers = []
    for col_idx in range(2, min(sheet.max_column + 1, 50)):  # Start from col 2, check up to 50
        header_val = sheet.cell(header_row, col_idx).value
        if header_val and str(header_val).strip():
            column_headers.append({
                'col_idx': col_idx,
                'header': str(header_val).strip(),
                'clean_header': clean_quarter_header(str(header_val))
            })
    
    print(f"Found {len(column_headers)} data columns: {[h['clean_header'] for h in column_headers[:8]]}...")
    
    # Process all rows to build comprehensive mapping
    field_mappings = []
    
    # Context tracking
    current_major_section = None
    current_section = None
    current_subsection = None
    
    for row_idx in range(1, sheet.max_row + 1):
        first_col = sheet.cell(row_idx, 1).value
        if not first_col:
            continue
            
        first_col_str = str(first_col).strip()
        
        # Skip header/empty rows
        if (row_idx <= 6 or 
            not first_col_str or
            first_col_str.lower().startswith('ipg photonics') or
            first_col_str.lower() in ['reported', 'quarterly']):
            continue
        
        # Classify row type and update context
        row_type = classify_reported_row(first_col_str, row_idx)
        
        if row_type == 'major_section':
            current_major_section = clean_field_name(first_col_str)
            current_section = None
            current_subsection = None
            print(f"Major Section: {current_major_section}")
            
        elif row_type == 'section_header':
            current_section = clean_field_name(first_col_str)
            current_subsection = None
            print(f"  Section: {current_section}")
            
        elif row_type == 'subsection_header':
            current_subsection = clean_field_name(first_col_str.rstrip(':'))
            print(f"    Subsection: {current_subsection}")
            
        elif row_type == 'data_field':
            # Get all values for this row across columns
            row_values = {}
            for col_info in column_headers:
                col_idx = col_info['col_idx']
                value = sheet.cell(row_idx, col_idx).value
                row_values[col_info['clean_header']] = value
            
            # Create comprehensive field mapping
            field_name = clean_field_name(first_col_str)
            scoped_name = build_reported_scoped_name(
                current_major_section, current_section, current_subsection, field_name
            )
            enhanced_scoped_name = apply_reported_enhanced_scoping(
                scoped_name, field_name, current_major_section, current_section
            )
            
            field_mapping = {
                'row_number': row_idx,
                'original_field_name': first_col_str,
                'cleaned_field_name': field_name,
                'major_section_context': current_major_section or '',
                'section_context': current_section or '',
                'subsection_context': current_subsection or '',
                'basic_scoped_name': scoped_name,
                'enhanced_scoped_name': enhanced_scoped_name,
                'row_type': row_type,
                'values': row_values,
                'has_data': any(v is not None and str(v).strip() for v in row_values.values())
            }
            
            field_mappings.append(field_mapping)
            
            # Print key fields for monitoring
            if any(keyword in first_col_str.lower() for keyword in 
                   ['total', 'revenue', 'china', 'germany', 'north america', 'net sales', 'income']):
                print(f"      Field: {first_col_str} → {enhanced_scoped_name}")
    
    wb.close()
    return field_mappings, column_headers


def classify_reported_row(first_col: str, row_idx: int) -> str:
    """
    Classify Reported tab row types.
    """
    first_col_lower = first_col.lower()
    first_col_upper = first_col.upper()
    
    # Major sections (all caps or specific patterns)
    if (first_col == first_col_upper and len(first_col) > 10) or any(pattern in first_col_lower for pattern in [
        'financial statements', 'segment information', 'income statement', 'balance sheet', 'cash flow'
    ]):
        return 'major_section'
    
    # Section headers (specific financial statement sections)
    section_patterns = [
        r'segment breakdown', r'end market breakdown', r'revenue by', r'by region',
        r'product breakdown', r'application breakdown', r'geographic breakdown',
        r'current assets', r'non.?current assets', r'current liabilities', r'equity',
        r'operating activities', r'investing activities', r'financing activities'
    ]
    
    for pattern in section_patterns:
        if re.search(pattern, first_col_lower):
            return 'section_header'
    
    # Subsection headers (lines ending with colon or specific indentation patterns)
    if first_col.endswith(':') or first_col.endswith(', of which'):
        return 'subsection_header'
    
    # Data fields - any meaningful content that's not clearly a header
    if (len(first_col) > 1 and 
        not first_col_lower.startswith('ipg photonics') and
        first_col_lower not in ['reported', 'quarterly'] and
        not first_col.startswith('(') and
        not re.match(r'^\d{4}', first_col)):  # Not a year
        return 'data_field'
    
    return 'other'


def clean_field_name(name: str) -> str:
    """Clean and standardize field names."""
    if not name:
        return ""
    
    # Remove extra whitespace and special characters
    cleaned = re.sub(r'\s+', '_', name.strip())
    cleaned = re.sub(r'[^\w\s-]', '', cleaned)
    cleaned = re.sub(r'[-_]+', '_', cleaned)
    cleaned = cleaned.strip('_')
    
    # Handle special cases
    if cleaned.lower() == 'total':
        return 'Total'
    if cleaned.lower() == 'other':
        return 'Other'
    
    # Capitalize appropriately
    if cleaned:
        words = cleaned.split('_')
        cleaned = '_'.join(word.capitalize() for word in words)
    
    return cleaned


def clean_quarter_header(header: str) -> str:
    """Clean quarter headers for column names."""
    header_str = str(header).strip()
    
    # Handle patterns like "1Q07", "2Q08", etc.
    if re.match(r'\d[QH]\d{2}', header_str):
        quarter = header_str[0]
        year_short = header_str[2:]
        year_full = f"20{year_short}" if int(year_short) < 50 else f"19{year_short}"
        return f"{year_full}_Q{quarter}"
    
    # Handle other date formats
    if re.match(r'\d{4}', header_str):
        return header_str
    
    return header_str.replace('-', '_').replace(' ', '_')


def build_reported_scoped_name(major_section: str, section: str, subsection: str, field: str) -> str:
    """Build scoped name for Reported tab fields."""
    components = ['Reported']
    
    if major_section:
        components.append(major_section)
    if section:
        components.append(section)
    if subsection:
        components.append(subsection)
    if field:
        components.append(field)
    
    return '.'.join(components)


def apply_reported_enhanced_scoping(scoped_name: str, field_name: str, 
                                   major_section: str, section: str) -> str:
    """Apply enhanced scoping rules specific to Reported tab."""
    field_lower = field_name.lower()
    
    # Geographic regions - highest priority
    geographic_mapping = {
        'north_america': 'North_America',
        'united_states': 'United_States',
        'germany': 'Germany', 
        'other_europe': 'Other_Europe',
        'china': 'China',
        'japan': 'Japan',
        'other_asian_countries': 'Other_Asia',
        'korea': 'Korea',
        'rest_of_world': 'Rest_Of_World'
    }
    
    for geo_key, geo_name in geographic_mapping.items():
        if (geo_key.replace('_', ' ') in field_lower or 
            geo_key in field_lower or
            field_lower == geo_key.split('_')[-1]):  # Match just "china", "japan", etc.
            
            if section and ('revenue' in section.lower() or 'segment' in section.lower()):
                return f"Revenue_Statement.Geographic_Breakdown.{geo_name}"
            elif major_section and 'balance' in major_section.lower():
                return f"Balance_Sheet.Geographic_Assets.{geo_name}"
            else:
                return f"Financial_Statements.Geographic_Data.{geo_name}"
    
    # Product/Application categories
    application_mapping = {
        'materials_processing': 'Materials_Processing',
        'other_application': 'Other_Applications',
        'advanced_applications': 'Advanced_Applications',
        'communications': 'Communications',
        'medical': 'Medical'
    }
    
    for app_key, app_name in application_mapping.items():
        if app_key.replace('_', ' ') in field_lower:
            return f"Revenue_Statement.Application_Breakdown.{app_name}"
    
    # Product type categories
    product_mapping = {
        'high_power_cw_lasers': 'High_Power_CW_Lasers',
        'pulsed_lasers': 'Pulsed_Lasers',
        'qcw_lasers': 'QCW_Lasers',
        'systems': 'Systems_And_Other'
    }
    
    for prod_key, prod_name in product_mapping.items():
        if prod_key.replace('_', ' ') in field_lower or 'high-power cw' in field_lower:
            return f"Revenue_Statement.Product_Type_Breakdown.{prod_name}"
    
    # Income Statement items
    if major_section and ('income' in major_section.lower() or 'segment' in major_section.lower()):
        if any(term in field_lower for term in ['revenue', 'sales', 'net_sales']):
            return f"Income_Statement.Revenue.{field_name}"
        elif any(term in field_lower for term in ['cost', 'expense']):
            return f"Income_Statement.Expenses.{field_name}"
        elif any(term in field_lower for term in ['income', 'profit', 'margin']):
            return f"Income_Statement.Profitability.{field_name}"
        elif 'total' in field_lower:
            return f"Income_Statement.Totals.{field_name}"
    
    # Balance Sheet items
    elif major_section and 'balance' in major_section.lower():
        if any(term in field_lower for term in ['asset', 'cash', 'inventory', 'receivable']):
            return f"Balance_Sheet.Assets.{field_name}"
        elif any(term in field_lower for term in ['liability', 'payable', 'debt']):
            return f"Balance_Sheet.Liabilities.{field_name}"
        elif any(term in field_lower for term in ['equity', 'retained', 'capital']):
            return f"Balance_Sheet.Equity.{field_name}"
    
    # Cash Flow items
    elif major_section and 'cash' in major_section.lower():
        if section and 'operating' in section.lower():
            return f"CashFlow_Statement.Operating_Activities.{field_name}"
        elif section and 'investing' in section.lower():
            return f"CashFlow_Statement.Investing_Activities.{field_name}"
        elif section and 'financing' in section.lower():
            return f"CashFlow_Statement.Financing_Activities.{field_name}"
    
    # Financial ratios and totals
    if 'total' in field_lower:
        if section and 'revenue' in section.lower():
            return f"Revenue_Statement.Totals.{field_name}"
        else:
            return f"Financial_Statements.Totals.{field_name}"
    
    # Default enhanced scoping
    return scoped_name


def save_reported_mapping(field_mappings: List[Dict], column_headers: List[Dict], output_file: str):
    """Save comprehensive Reported tab mapping to CSV."""
    
    # Prepare dynamic column names for all quarter columns
    quarter_columns = [h['clean_header'] for h in column_headers]
    
    with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
        # Define fieldnames
        base_fieldnames = [
            'Row_Number',
            'Original_Field_Name',
            'Cleaned_Field_Name',
            'Major_Section_Context',
            'Section_Context',
            'Subsection_Context', 
            'Basic_Scoped_Name',
            'Enhanced_Scoped_Name',
            'Row_Type',
            'Has_Data'
        ]
        
        # Add all quarter columns (limit to first 20 to keep manageable)
        limited_quarters = quarter_columns[:20]
        fieldnames = base_fieldnames + limited_quarters
        
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        
        for mapping in field_mappings:
            # Build row data
            row_data = {
                'Row_Number': mapping['row_number'],
                'Original_Field_Name': mapping['original_field_name'],
                'Cleaned_Field_Name': mapping['cleaned_field_name'],
                'Major_Section_Context': mapping['major_section_context'],
                'Section_Context': mapping['section_context'],
                'Subsection_Context': mapping['subsection_context'],
                'Basic_Scoped_Name': mapping['basic_scoped_name'],
                'Enhanced_Scoped_Name': mapping['enhanced_scoped_name'],
                'Row_Type': mapping['row_type'],
                'Has_Data': 'Yes' if mapping['has_data'] else 'No'
            }
            
            # Add values for each quarter column (limited)
            for quarter_col in limited_quarters:
                value = mapping['values'].get(quarter_col)
                if value is not None:
                    row_data[quarter_col] = str(value)
                else:
                    row_data[quarter_col] = ''
            
            writer.writerow(row_data)


def main():
    """Main entry point."""
    target_file = "/Users/michaelkim/code/Bernstein/20240725_IPGP.US-IPG Photonics.xlsx"
    output_file = "/Users/michaelkim/code/Bernstein/reported_tab_comprehensive_mapping.csv"
    
    print("="*80)
    print("REPORTED TAB COMPREHENSIVE FIELD MAPPING")
    print("="*80)
    print(f"Target file: {target_file}")
    print(f"Output file: {output_file}")
    
    if not Path(target_file).exists():
        print(f"ERROR: Target file not found: {target_file}")
        return
    
    try:
        # Analyze Reported tab
        field_mappings, column_headers = analyze_reported_tab_structure(target_file)
        
        if not field_mappings:
            print("ERROR: No field mappings generated")
            return
        
        print(f"\n" + "="*50)
        print("RESULTS SUMMARY")
        print("="*50)
        print(f"Total fields mapped: {len(field_mappings)}")
        print(f"Data columns found: {len(column_headers)}")
        
        # Count by major section
        major_sections = {}
        sections = {}
        for mapping in field_mappings:
            major_section = mapping['major_section_context'] or 'No_Major_Section'
            section = mapping['section_context'] or 'No_Section'
            major_sections[major_section] = major_sections.get(major_section, 0) + 1
            sections[section] = sections.get(section, 0) + 1
        
        print("\nFields by major section:")
        for major_section, count in sorted(major_sections.items()):
            print(f"  {major_section}: {count} fields")
        
        print("\nTop sections:")
        for section, count in sorted(sections.items(), key=lambda x: x[1], reverse=True)[:10]:
            if section != 'No_Section':
                print(f"  {section}: {count} fields")
        
        # Save to CSV
        print(f"\n" + "="*50)
        print("SAVING TO CSV")
        print("="*50)
        save_reported_mapping(field_mappings, column_headers, output_file)
        
        print(f"✓ Reported tab comprehensive mapping saved to: {output_file}")
        print(f"✓ Contains {len(field_mappings)} fields with full hierarchical context")
        print(f"✓ Includes data values across {min(20, len(column_headers))} time periods")
        
        print(f"\n" + "="*80)
        print("PROCESSING COMPLETE!")
        print("="*80)
        print(f"Review the comprehensive Reported tab mapping in: {output_file}")
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
