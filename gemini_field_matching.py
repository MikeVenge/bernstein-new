#!/usr/bin/env python3
"""
Gemini 2.5 Pro Field Matching via L2M2

Uses Gemini 2.5 Pro to intelligently match source and destination fields
by sending the fully scoped field lists and asking for smart matching.

Author: AI Assistant
Date: October 2025
"""

import pandas as pd
import requests
import json
import csv
from pathlib import Path
import os
from typing import Dict, List, Tuple, Optional


def load_source_and_destination_fields() -> Tuple[List[Dict], List[Dict]]:
    """Load the fully scoped fields from both source and destination."""
    print("Loading fully scoped fields for Gemini matching...")
    
    # Load source fields (Key Metrics with enhanced scoping)
    source_file = "/Users/michaelkim/code/Bernstein/final_improved_key_metrics_mapping.csv"
    source_fields = []
    
    if Path(source_file).exists():
        source_df = pd.read_csv(source_file)
        for idx, row in source_df.iterrows():
            source_fields.append({
                'source_id': f"SRC_{idx+1}",
                'row_number': row['Row_Number'],
                'original_field_name': row['Original_Field_Name'],
                'section_context': row['Section_Context'] if pd.notna(row['Section_Context']) else '',
                'enhanced_scoped_name': row['Enhanced_Scoped_Name'],
                'q1_2024_value': row['Q1_2024_Value'] if pd.notna(row['Q1_2024_Value']) else None,
                'q2_2024_value': row['Q2_2024_Value'] if pd.notna(row['Q2_2024_Value']) else None,
                'is_percentage': row['Is_Percentage_Section'] if 'Is_Percentage_Section' in row else False
            })
        print(f"Loaded {len(source_fields)} source fields")
    
    # Load destination fields (Reported with enhanced scoping)
    dest_file = "/Users/michaelkim/code/Bernstein/reported_tab_comprehensive_mapping.csv"
    dest_fields = []
    
    if Path(dest_file).exists():
        dest_df = pd.read_csv(dest_file)
        for idx, row in dest_df.iterrows():
            dest_fields.append({
                'dest_id': f"DEST_{idx+1}",
                'row_number': row['Row_Number'],
                'original_field_name': row['Original_Field_Name'],
                'major_section_context': row['Major_Section_Context'] if pd.notna(row['Major_Section_Context']) else '',
                'section_context': row['Section_Context'] if pd.notna(row['Section_Context']) else '',
                'enhanced_scoped_name': row['Enhanced_Scoped_Name']
            })
        print(f"Loaded {len(dest_fields)} destination fields")
    
    return source_fields, dest_fields


def prepare_gemini_prompt(source_fields: List[Dict], dest_fields: List[Dict]) -> str:
    """Prepare the prompt for Gemini 2.5 Pro."""
    
    prompt = """You are an expert financial data analyst. I need you to match source fields to destination fields based on their semantic meaning and hierarchical context.

TASK: Match each destination field to the most appropriate source field. Return ONLY a CSV with the matched pairs.

IMPORTANT RULES:
1. Each source field should map to AT MOST ONE destination field (one-to-one mapping)
2. Pay attention to percentage vs absolute value contexts:
   - "Revenue by region" (absolute values) should match "Region breakdown" 
   - "Revenue by region (% of total)" (percentages) should match "Region mix (%)"
   - "Revenue by application" should match "End market breakdown"
   - "Revenue by product" should match "Segment breakdown"
3. Geographic terms should match: "North America" ↔ "United States and other North America"
4. Only return matches where you are confident (>70% confidence)
5. Return CSV format with headers: source_id,dest_id,confidence_score,match_reason

SOURCE FIELDS (with enhanced scoping):
"""
    
    # Add source fields
    for field in source_fields:
        prompt += f"{field['source_id']}: {field['original_field_name']} | Section: {field['section_context']} | Scope: {field['enhanced_scoped_name']} | Is_Percentage: {field['is_percentage']}\n"
    
    prompt += "\nDESTINATION FIELDS (with enhanced scoping):\n"
    
    # Add destination fields
    for field in dest_fields:
        prompt += f"{field['dest_id']}: {field['original_field_name']} | Section: {field['section_context']} | Scope: {field['enhanced_scoped_name']}\n"
    
    prompt += """
RETURN ONLY CSV FORMAT:
source_id,dest_id,confidence_score,match_reason

Example:
SRC_6,DEST_12,0.95,North America revenue absolute values match
SRC_16,DEST_48,0.90,North America revenue percentage values match
"""
    
    return prompt


def call_gemini_via_l2m2(prompt: str) -> str:
    """Call Gemini 2.5 Pro via L2M2 API."""
    print("Calling Gemini 2.5 Pro via L2M2...")
    
    # L2M2 API configuration
    l2m2_url = os.getenv('L2M2_API_URL', 'https://l2m2-api.adgo.dev')
    api_token = os.getenv('L2M2_API_TOKEN')
    
    if not api_token:
        print("ERROR: L2M2_API_TOKEN environment variable not set")
        print("Please set your L2M2 API token:")
        print("export L2M2_API_TOKEN='your_token_here'")
        return ""
    
    headers = {
        'Authorization': f'Bearer {api_token}',
        'Content-Type': 'application/json'
    }
    
    payload = {
        'model': 'gemini-2.5-pro',
        'messages': [
            {
                'role': 'user',
                'content': prompt
            }
        ],
        'temperature': 0.1,  # Low temperature for consistent results
        'max_tokens': 4000
    }
    
    try:
        response = requests.post(f'{l2m2_url}/v1/chat/completions', 
                               headers=headers, 
                               json=payload, 
                               timeout=120)
        
        if response.status_code == 200:
            result = response.json()
            return result['choices'][0]['message']['content']
        else:
            print(f"ERROR: API call failed with status {response.status_code}")
            print(f"Response: {response.text}")
            return ""
            
    except Exception as e:
        print(f"ERROR calling L2M2: {e}")
        return ""


def parse_gemini_response(response: str) -> List[Dict]:
    """Parse Gemini's CSV response into structured data."""
    print("Parsing Gemini response...")
    
    if not response:
        return []
    
    # Extract CSV content from response
    lines = response.strip().split('\n')
    csv_lines = []
    
    # Find CSV content (skip any explanatory text)
    in_csv = False
    for line in lines:
        if 'source_id' in line and 'dest_id' in line:
            in_csv = True
        if in_csv and ',' in line:
            csv_lines.append(line)
    
    if not csv_lines:
        print("ERROR: No CSV content found in Gemini response")
        return []
    
    # Parse CSV
    matches = []
    header_processed = False
    
    for line in csv_lines:
        if not header_processed:
            header_processed = True
            continue
            
        parts = [part.strip() for part in line.split(',')]
        if len(parts) >= 4:
            matches.append({
                'source_id': parts[0],
                'dest_id': parts[1],
                'confidence_score': float(parts[2]) if parts[2] else 0.0,
                'match_reason': parts[3] if len(parts) > 3 else ''
            })
    
    print(f"Parsed {len(matches)} matches from Gemini response")
    return matches


def enrich_gemini_matches(matches: List[Dict], source_fields: List[Dict], 
                         dest_fields: List[Dict]) -> List[Dict]:
    """Enrich Gemini matches with actual field data and values."""
    print("Enriching Gemini matches with field data and values...")
    
    # Create lookups
    source_lookup = {field['source_id']: field for field in source_fields}
    dest_lookup = {field['dest_id']: field for field in dest_fields}
    
    # Load destination Q1 values
    dest_q1_values = get_destination_q1_values()
    
    enriched_matches = []
    
    for match in matches:
        source_id = match['source_id']
        dest_id = match['dest_id']
        
        if source_id in source_lookup and dest_id in dest_lookup:
            source_field = source_lookup[source_id]
            dest_field = dest_lookup[dest_id]
            
            dest_q1 = dest_q1_values.get(dest_field['row_number'])
            
            enriched_match = {
                'source_field': source_field,
                'dest_field': dest_field,
                'dest_q1_value': dest_q1,
                'confidence_score': match['confidence_score'],
                'match_reason': match['match_reason'],
                'gemini_matched': True
            }
            
            enriched_matches.append(enriched_match)
    
    return enriched_matches


def get_destination_q1_values() -> Dict:
    """Get Q1 2024 values from destination file (Column BR)."""
    target_file = "/Users/michaelkim/code/Bernstein/20240725_IPGP.US-IPG Photonics.xlsx"
    
    import openpyxl
    wb = openpyxl.load_workbook(target_file, data_only=True)
    sheet = wb['Reported']
    
    dest_values = {}
    for row_idx in range(1, sheet.max_row + 1):
        q1_value = sheet.cell(row_idx, 70).value  # Column BR
        dest_values[row_idx] = q1_value
    
    wb.close()
    return dest_values


def save_gemini_results(enriched_matches: List[Dict], output_file: str):
    """Save Gemini matching results."""
    with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
        fieldnames = [
            'Source_ID',
            'Dest_ID', 
            'Source_Row_Number',
            'Dest_Row_Number',
            'Source_Field_Name',
            'Dest_Field_Name',
            'Source_Section_Context',
            'Dest_Section_Context',
            'Source_Enhanced_Scope',
            'Dest_Enhanced_Scope',
            'Source_Q1_2024_Value',
            'Dest_Q1_2024_Value',
            'Source_Q2_2024_Value',
            'Confidence_Score',
            'Match_Reason',
            'Values_Match'
        ]
        
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        
        for match in enriched_matches:
            source = match['source_field']
            dest = match['dest_field']
            dest_q1 = match['dest_q1_value']
            
            # Check value match
            source_q1 = source['q1_2024_value']
            values_match = 'N/A'
            
            if dest_q1 is not None and source_q1 is not None:
                if dest_q1 == source_q1:
                    values_match = 'Exact Match'
                elif abs(float(dest_q1) - float(source_q1)) / float(source_q1) < 0.05:
                    values_match = 'Close Match'
                else:
                    values_match = 'Different'
            elif dest_q1 is None and source_q1 is None:
                values_match = 'Both Empty'
            else:
                values_match = 'One Empty'
            
            writer.writerow({
                'Source_ID': source['source_id'],
                'Dest_ID': dest['dest_id'],
                'Source_Row_Number': source['row_number'],
                'Dest_Row_Number': dest['row_number'],
                'Source_Field_Name': source['original_field_name'],
                'Dest_Field_Name': dest['original_field_name'],
                'Source_Section_Context': source['section_context'],
                'Dest_Section_Context': dest['section_context'],
                'Source_Enhanced_Scope': source['enhanced_scoped_name'],
                'Dest_Enhanced_Scope': dest['enhanced_scoped_name'],
                'Source_Q1_2024_Value': source_q1 if source_q1 is not None else '',
                'Dest_Q1_2024_Value': dest_q1 if dest_q1 is not None else '',
                'Source_Q2_2024_Value': source['q2_2024_value'] if source['q2_2024_value'] is not None else '',
                'Confidence_Score': match['confidence_score'],
                'Match_Reason': match['match_reason'],
                'Values_Match': values_match
            })


def main():
    """Main entry point."""
    output_file = "/Users/michaelkim/code/Bernstein/gemini_field_matching_results.csv"
    
    print("="*80)
    print("GEMINI 2.5 PRO FIELD MATCHING VIA L2M2")
    print("="*80)
    print("Using AI to intelligently match fully scoped fields")
    print(f"Output file: {output_file}")
    
    try:
        # Load source and destination fields
        source_fields, dest_fields = load_source_and_destination_fields()
        
        if not source_fields or not dest_fields:
            print("ERROR: Could not load field mappings")
            return
        
        # Prepare prompt for Gemini
        prompt = prepare_gemini_prompt(source_fields, dest_fields)
        
        print(f"Sending {len(source_fields)} source and {len(dest_fields)} destination fields to Gemini...")
        
        # Call Gemini via L2M2
        gemini_response = call_gemini_via_l2m2(prompt)
        
        if not gemini_response:
            print("ERROR: No response from Gemini")
            return
        
        print("Gemini response received!")
        
        # Parse Gemini response
        matches = parse_gemini_response(gemini_response)
        
        if not matches:
            print("ERROR: Could not parse matches from Gemini response")
            print("Gemini response:")
            print(gemini_response[:1000] + "..." if len(gemini_response) > 1000 else gemini_response)
            return
        
        # Enrich with actual data
        enriched_matches = enrich_gemini_matches(matches, source_fields, dest_fields)
        
        # Save results
        save_gemini_results(enriched_matches, output_file)
        
        # Summary
        exact_matches = [m for m in enriched_matches if 
                        m['dest_q1_value'] is not None and 
                        m['source_field']['q1_2024_value'] is not None and
                        m['dest_q1_value'] == m['source_field']['q1_2024_value']]
        
        print(f"\n" + "="*80)
        print("GEMINI FIELD MATCHING RESULTS")
        print("="*80)
        print(f"Total matches found: {len(enriched_matches)}")
        print(f"Exact value matches: {len(exact_matches)}")
        print(f"Average confidence: {sum(m['confidence_score'] for m in enriched_matches)/len(enriched_matches):.3f}" if enriched_matches else "N/A")
        
        # Show top matches
        print(f"\nTop exact value matches:")
        for match in exact_matches[:10]:
            source = match['source_field']
            dest = match['dest_field']
            value = match['dest_q1_value']
            print(f"  {source['original_field_name']} → {dest['original_field_name']}: {value:,}")
        
        print(f"\nResults saved to: {output_file}")
        print("Gemini has provided intelligent field matching!")
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
