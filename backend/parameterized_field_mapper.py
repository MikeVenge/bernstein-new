#!/usr/bin/env python3
"""
Parameterized Field Mapper - Backend Copy

Copy of the parameterized field mapper for use in the backend API.
"""

import openpyxl
import pandas as pd
import csv
from pathlib import Path
from typing import Dict, List, Optional, Tuple


class ParameterizedFieldMapper:
    """Parameterized field mapping and population engine."""
    
    def __init__(self, source_file: str, destination_file: str, mapping_file: str, 
                 target_column: int, data_column: str = "CO", output_file: Optional[str] = None, 
                 audit_file: Optional[str] = None):
        """Initialize the parameterized field mapper."""
        
        self.source_file = Path(source_file)
        self.destination_file = Path(destination_file)
        self.mapping_file = Path(mapping_file)
        self.target_column = target_column
        self.source_tracking_column = target_column + 1
        self.data_column = data_column
        
        # Convert column reference to number
        self.source_column_num = self._column_ref_to_number(data_column)
        
        if output_file:
            self.output_file = Path(output_file)
        else:
            self.output_file = self.destination_file.parent / f"populated_{self.destination_file.name}"
        
        if audit_file:
            self.audit_file = Path(audit_file)
        else:
            self.audit_file = self.destination_file.parent / f"audit_trail_{self.destination_file.stem}.csv"
        
        self.stats = {
            'mappings_processed': 0,
            'values_populated': 0,
            'source_tracking_added': 0,
            'errors': [],
            'sheet_stats': {},
            'method_stats': {}
        }
    
    def _column_ref_to_number(self, column_ref: str) -> int:
        """Convert column reference to number."""
        if column_ref == 'CO':
            return 93
        elif column_ref == 'BR':
            return 70
        elif column_ref == 'BS':
            return 71
        elif column_ref == 'BT':
            return 72
        elif len(column_ref) == 2:
            first = ord(column_ref[0]) - ord('A') + 1
            second = ord(column_ref[1]) - ord('A') + 1
            return first * 26 + second
        else:
            return 93
    
    def load_mapping_file(self) -> List[Dict]:
        """Load and validate the mapping file."""
        
        if not self.mapping_file.exists():
            raise FileNotFoundError(f"Mapping file not found: {self.mapping_file}")
        
        mappings = []
        with open(self.mapping_file, 'r', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            mappings = list(reader)
        
        return mappings
    
    def validate_files(self) -> Tuple[bool, List[str]]:
        """Validate that all required files exist and have required sheets."""
        
        errors = []
        
        if not self.source_file.exists():
            errors.append(f"Source file not found: {self.source_file}")
            return False, errors
        
        if not self.destination_file.exists():
            errors.append(f"Destination file not found: {self.destination_file}")
            return False, errors
        
        if not self.mapping_file.exists():
            errors.append(f"Mapping file not found: {self.mapping_file}")
            return False, errors
        
        if self.target_column < 1:
            errors.append(f"Invalid target column: {self.target_column}")
        
        # Validate destination file has 'Reported' sheet
        try:
            dest_wb = openpyxl.load_workbook(self.destination_file, data_only=False, read_only=True)
            if 'Reported' not in dest_wb.sheetnames:
                available_sheets = ", ".join(dest_wb.sheetnames)
                errors.append(
                    f"DESTINATION FILE ERROR: Required sheet 'Reported' not found. "
                    f"Available sheets: [{available_sheets}]. "
                    f"Please ensure your destination file has a sheet named 'Reported'."
                )
            dest_wb.close()
        except Exception as e:
            errors.append(f"Failed to validate destination file: {str(e)}")
        
        # Validate source file has required sheets from mapping
        try:
            source_wb = openpyxl.load_workbook(self.source_file, data_only=True, read_only=True)
            source_sheets = set(source_wb.sheetnames)
            
            # Load mapping to check required source sheets
            required_sheets = set()
            with open(self.mapping_file, 'r', encoding='utf-8') as f:
                import csv
                reader = csv.DictReader(f)
                for row in reader:
                    sheet_name = row.get('Source_Sheet', row.get('Source_Sheet_Name', ''))
                    if sheet_name:
                        required_sheets.add(sheet_name)
            
            missing_sheets = required_sheets - source_sheets
            if missing_sheets:
                available_sheets = ", ".join(sorted(source_sheets))
                missing_list = ", ".join(sorted(missing_sheets))
                errors.append(
                    f"SOURCE FILE ERROR: Required sheet(s) not found: [{missing_list}]. "
                    f"Available sheets: [{available_sheets}]. "
                    f"Please ensure your source file has all required sheets."
                )
            
            source_wb.close()
        except Exception as e:
            errors.append(f"Failed to validate source file: {str(e)}")
        
        return len(errors) == 0, errors
    
    def populate_fields(self, mappings: List[Dict]) -> List[Dict]:
        """Populate destination fields using the mappings."""
        
        source_wb = openpyxl.load_workbook(self.source_file, data_only=True)
        dest_wb = openpyxl.load_workbook(self.destination_file, data_only=False)
        dest_sheet = dest_wb['Reported']
        
        population_results = []
        
        for mapping in mappings:
            result = self._process_single_mapping(mapping, source_wb, dest_sheet)
            population_results.append(result)
        
        dest_wb.save(self.output_file)
        source_wb.close()
        dest_wb.close()
        
        self.stats['mappings_processed'] = len(mappings)
        return population_results
    
    def _process_single_mapping(self, mapping: Dict, source_wb: openpyxl.Workbook, 
                               dest_sheet) -> Dict:
        """Process a single field mapping."""
        
        # Handle both generic format and original format
        dest_row = int(mapping.get('Dest_Row', mapping.get('Dest_Row_Number', 0)))
        source_sheet_name = mapping.get('Source_Sheet', mapping.get('Source_Sheet_Name', ''))
        source_row = mapping.get('Source_Row', mapping.get('Source_Row_Number', ''))
        
        result = {
            'Dest_Row': dest_row,
            'Source_Sheet': source_sheet_name,
            'Source_Row': source_row,
            'Status': 'ERROR'
        }
        
        try:
            if source_sheet_name not in source_wb.sheetnames:
                result['Status'] = 'SOURCE_SHEET_NOT_FOUND'
                return result
            
            source_sheet = source_wb[source_sheet_name]
            
            # Handle composite fields
            if '+' in str(source_row):
                composite_rows = [int(r.strip()) for r in str(source_row).split('+')]
                composite_value = 0
                
                for comp_row in composite_rows:
                    comp_value = source_sheet.cell(comp_row, self.source_column_num).value or 0
                    composite_value += comp_value
                
                source_value = composite_value
                source_location = f"{self.source_file.name}|{source_sheet_name}|{source_row}|{self.data_column}"
            else:
                source_row_num = int(source_row)
                source_value = source_sheet.cell(source_row_num, self.source_column_num).value
                source_location = f"{self.source_file.name}|{source_sheet_name}|{source_row_num}|{self.data_column}"
            
            result['Value'] = source_value
            result['Source_Location'] = source_location
            
            if source_value is not None:
                dest_sheet.cell(dest_row, self.target_column).value = source_value
                dest_sheet.cell(dest_row, self.source_tracking_column).value = source_location
                
                self.stats['values_populated'] += 1
                result['Status'] = 'POPULATED'
            else:
                result['Status'] = 'NO_SOURCE_DATA'
                
        except Exception as e:
            result['Status'] = 'ERROR'
            result['Error'] = str(e)
            self.stats['errors'].append(f"Row {dest_row}: {str(e)}")
        
        return result
    
    def generate_audit_trail(self, population_results: List[Dict]):
        """Generate audit trail CSV file."""
        
        with open(self.audit_file, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ['Dest_Row', 'Source_Sheet', 'Source_Row', 'Value', 'Source_Location', 'Status']
            
            if any('Error' in result for result in population_results):
                fieldnames.append('Error')
            
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(population_results)
