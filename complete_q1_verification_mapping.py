#!/usr/bin/env python3
"""
Complete Q1 Value Verification Mapping - All Financial Statements

Extends Q1 verification to cover ALL sections:
1. Segment Information (already done)
2. Consolidated Income Statement (in '000 USD) - YTD
3. Consolidated Balance Sheet (in '000 USD)
4. Consolidated Cash Flow Statement (in '000 USD) - YTD

Uses multiple source sheets: Key Metrics, Income Statement, Balance Sheet, Cash Flows

Author: AI Assistant
Date: October 2025
"""

import openpyxl
import csv
import pandas as pd
from pathlib import Path
import re
from typing import Dict, List, Tuple, Optional


def load_all_source_sheets() -> Dict[str, Dict]:
    """Load enhanced scoping from ALL source sheets."""
    print("Loading enhanced scoping from ALL source sheets...")
    
    source_file = "/Users/michaelkim/code/Bernstein/IPGP-Financial-Data-Workbook-2024-Q2.xlsx"
    
    # Load Key Metrics (already have this)
    key_metrics_scoping = {}
    km_file = "/Users/michaelkim/code/Bernstein/final_improved_key_metrics_mapping.csv"
    if Path(km_file).exists():
        km_df = pd.read_csv(km_file)
        for _, row in km_df.iterrows():
            row_num = row['Row_Number']
            key_metrics_scoping[row_num] = {
                'sheet_name': 'Key Metrics',
                'original_field_name': row['Original_Field_Name'],
                'enhanced_scoped_name': row['Enhanced_Scoped_Name'],
                'section_context': row['Section_Context'] if pd.notna(row['Section_Context']) else '',
                'q1_2024_value': row['Q1_2024_Value'] if pd.notna(row['Q1_2024_Value']) else None,
                'q2_2024_value': row['Q2_2024_Value'] if pd.notna(row['Q2_2024_Value']) else None
            }
        print(f"Loaded Key Metrics scoping: {len(key_metrics_scoping)} fields")
    
    # Load other sheets with enhanced scoping
    other_sheets_scoping = {}
    wb = openpyxl.load_workbook(source_file, data_only=True)
    
    for sheet_name in ['Income Statement', 'Balance Sheet', 'Cash Flows']:
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"Processing {sheet_name} sheet...")
            
            # Find Q1 and Q2 2024 columns
            q1_col, q2_col = find_2024_columns(sheet)
            
            if q1_col and q2_col:
                print(f"  Found Q1 column: {q1_col}, Q2 column: {q2_col}")
                
                for row_idx in range(1, min(sheet.max_row + 1, 200)):
                    first_col = sheet.cell(row_idx, 1).value
                    if first_col and isinstance(first_col, str) and len(first_col.strip()) > 2:
                        q1_value = sheet.cell(row_idx, q1_col).value
                        q2_value = sheet.cell(row_idx, q2_col).value
                        
                        if q1_value is not None:  # Only include rows with Q1 data
                            enhanced_scope = create_enhanced_scope_for_sheet(
                                sheet_name, first_col.strip(), row_idx, sheet
                            )
                            
                            other_sheets_scoping[f"{sheet_name}_{row_idx}"] = {
                                'sheet_name': sheet_name,
                                'row_number': row_idx,
                                'original_field_name': first_col.strip(),
                                'enhanced_scoped_name': enhanced_scope,
                                'section_context': sheet_name,
                                'q1_2024_value': q1_value,
                                'q2_2024_value': q2_value
                            }
            else:
                print(f"  Warning: Could not find 2024 Q1/Q2 columns in {sheet_name}")
    
    wb.close()
    print(f"Loaded other sheets scoping: {len(other_sheets_scoping)} fields")
    
    return {
        'Key Metrics': key_metrics_scoping,
        'Other Sheets': other_sheets_scoping
    }


def find_2024_columns(sheet) -> Tuple[Optional[int], Optional[int]]:
    """Find Q1 and Q2 2024 columns in a sheet."""
    q1_col = None
    q2_col = None
    
    # Check header rows (typically row 4 or 5)
    for header_row in [4, 5]:
        for col_idx in range(80, min(sheet.max_column + 1, 100)):
            header_val = sheet.cell(header_row, col_idx).value
            if header_val:
                header_str = str(header_val)
                if '2024-03-31' in header_str:  # Q1 2024
                    q1_col = col_idx
                elif '2024-06-30' in header_str:  # Q2 2024
                    q2_col = col_idx
    
    return q1_col, q2_col


def create_enhanced_scope_for_sheet(sheet_name: str, field_name: str, row_idx: int, sheet) -> str:
    """Create enhanced scope for fields from Income Statement, Balance Sheet, Cash Flows."""
    
    # Build hierarchical context by looking backwards for section headers
    current_section = None
    current_subsection = None
    
    for check_row in range(row_idx - 1, max(0, row_idx - 15), -1):
        check_label = sheet.cell(check_row, 1).value
        if check_label and isinstance(check_label, str):
            check_label = check_label.strip()
            
            # Look for section headers
            if any(pattern in check_label.lower() for pattern in [
                'assets', 'current assets', 'non-current assets',
                'liabilities', 'current liabilities', 'equity',
                'operating activities', 'investing activities', 'financing activities',
                'revenue', 'expenses', 'income', 'comprehensive income'
            ]):
                if not current_section:
                    current_section = clean_field_name(check_label.rstrip(':'))
                elif not current_subsection:
                    current_subsection = clean_field_name(check_label.rstrip(':'))
    
    # Build enhanced scope
    components = [clean_field_name(sheet_name)]
    
    if current_section:
        components.append(current_section)
    if current_subsection:
        components.append(current_subsection)
    
    components.append(clean_field_name(field_name))
    
    return '.'.join(components)


def clean_field_name(name: str) -> str:
    """Clean field names for scoping."""
    if not name:
        return ""
    
    cleaned = re.sub(r'\s+', '_', name.strip())
    cleaned = re.sub(r'[^\w\s-]', '', cleaned)
    cleaned = re.sub(r'[-_]+', '_', cleaned)
    cleaned = cleaned.strip('_')
    
    if cleaned:
        words = cleaned.split('_')
        cleaned = '_'.join(word.capitalize() for word in words)
    
    return cleaned


def perform_complete_q1_verification() -> List[Dict]:
    """
    Perform Q1 verification across ALL financial statement sections.
    """
    print("="*80)
    print("COMPLETE Q1 VALUE VERIFICATION - ALL FINANCIAL STATEMENTS")
    print("="*80)
    
    # Load destination enhanced scoping
    dest_file = "/Users/michaelkim/code/Bernstein/reported_tab_comprehensive_mapping.csv"
    dest_scoping = {}
    
    if Path(dest_file).exists():
        dest_df = pd.read_csv(dest_file)
        for _, row in dest_df.iterrows():
            row_num = row['Row_Number']
            dest_scoping[row_num] = {
                'original_field_name': row['Original_Field_Name'],
                'enhanced_scoped_name': row['Enhanced_Scoped_Name'],
                'major_section_context': row['Major_Section_Context'] if pd.notna(row['Major_Section_Context']) else '',
                'section_context': row['Section_Context'] if pd.notna(row['Section_Context']) else ''
            }
        print(f"Loaded destination scoping: {len(dest_scoping)} fields")
    
    # Load all source sheets
    all_source_scoping = load_all_source_sheets()
    
    # Combine all source fields into one lookup
    all_source_fields = {}
    for sheet_type, fields in all_source_scoping.items():
        all_source_fields.update(fields)
    
    print(f"Total source fields available: {len(all_source_fields)}")
    
    # Load destination Q1 values
    dest_file_path = "/Users/michaelkim/code/Bernstein/20240725_IPGP.US-IPG Photonics.xlsx"
    dest_wb = openpyxl.load_workbook(dest_file_path, data_only=True)
    dest_sheet = dest_wb['Reported']
    
    dest_q1_data = {}
    for row_idx in range(1, dest_sheet.max_row + 1):
        q1_value = dest_sheet.cell(row_idx, 70).value  # Column BR
        if q1_value is not None:
            dest_q1_data[row_idx] = q1_value
    
    dest_wb.close()
    print(f"Loaded destination Q1 data: {len(dest_q1_data)} rows")
    
    # Perform comprehensive Q1 verification
    print("\nPerforming comprehensive Q1 value verification...")
    
    matches = []
    used_source_keys = set()
    
    for dest_row, dest_q1_value in dest_q1_data.items():
        if dest_row not in dest_scoping:
            continue
            
        dest_field_info = dest_scoping[dest_row]
        dest_major_section = dest_field_info['major_section_context'].lower()
        
        print(f"\nDEST Row {dest_row}: {dest_field_info['original_field_name']}")
        print(f"  Section: {dest_major_section}")
        print(f"  Q1 value: {dest_q1_value}")
        
        # Find matching source field with same Q1 value
        best_match = None
        
        for source_key, source_field in all_source_fields.items():
            if (source_key in used_source_keys or 
                source_field['q1_2024_value'] is None or
                source_field['q1_2024_value'] != dest_q1_value):
                continue
            
            # Found exact Q1 match
            match = {
                'Dest_Row_Number': dest_row,
                'Dest_Field_Name': dest_field_info['original_field_name'],
                'Dest_Enhanced_Scope': dest_field_info['enhanced_scoped_name'],
                'Dest_Section_Context': dest_field_info['section_context'],
                'Dest_Major_Section_Context': dest_field_info['major_section_context'],
                
                'Source_Sheet_Name': source_field['sheet_name'],
                'Source_Row_Number': source_field.get('row_number', str(source_key).split('_')[1] if '_' in str(source_key) else ''),
                'Source_Field_Name': source_field['original_field_name'],
                'Source_Enhanced_Scope': source_field['enhanced_scoped_name'],
                'Source_Section_Context': source_field['section_context'],
                
                'Q1_Verification_Value': dest_q1_value,
                'Source_Q2_Value': source_field['q2_2024_value'],
                'Match_Method': 'Q1_Value_Verification_Complete',
                'Match_Confidence': 1.0
            }
            
            matches.append(match)
            used_source_keys.add(source_key)
            
            print(f"  ✓ MATCHED to {source_field['sheet_name']} Row {source_field.get('row_number', 'N/A')}: {source_field['original_field_name']}")
            print(f"    Q1 verification: {dest_q1_value} = {source_field['q1_2024_value']} ✓")
            print(f"    Q2 available: {source_field['q2_2024_value']}")
            
            break
        
        if not best_match:
            print(f"  ❌ No Q1 match found")
    
    print(f"\nComplete Q1 verification matching results:")
    print(f"  Total matches found: {len(matches)}")
    print(f"  Source fields used: {len(used_source_keys)}")
    
    # Show breakdown by sheet
    sheet_breakdown = {}
    for match in matches:
        sheet = match['Source_Sheet_Name']
        sheet_breakdown[sheet] = sheet_breakdown.get(sheet, 0) + 1
    
    print(f"  Matches by source sheet:")
    for sheet, count in sheet_breakdown.items():
        print(f"    {sheet}: {count} matches")
    
    return matches


def save_complete_mapping_file(matches: List[Dict], output_file: str):
    """Save the complete Q1 verification mapping file."""
    print(f"\nSaving complete mapping file to: {output_file}")
    
    with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
        fieldnames = [
            'Dest_Row_Number',
            'Dest_Field_Name', 
            'Dest_Enhanced_Scope',
            'Dest_Section_Context',
            'Dest_Major_Section_Context',
            'Source_Sheet_Name',
            'Source_Row_Number',
            'Source_Field_Name',
            'Source_Enhanced_Scope', 
            'Source_Section_Context',
            'Q1_Verification_Value',
            'Source_Q2_Value',
            'Match_Method',
            'Match_Confidence'
        ]
        
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(matches)
    
    print(f"✓ Complete mapping file saved with {len(matches)} verified matches")


def main():
    """Main entry point for complete Q1 verification mapping."""
    output_file = "/Users/michaelkim/code/Bernstein/complete_q1_verified_mapping.csv"
    
    print("="*80)
    print("COMPLETE Q1 VALUE VERIFICATION - ALL FINANCIAL STATEMENTS")
    print("="*80)
    print("Covering ALL sections:")
    print("  1. Segment Information")
    print("  2. Consolidated Income Statement (in '000 USD) - YTD")
    print("  3. Consolidated Balance Sheet (in '000 USD)")
    print("  4. Consolidated Cash Flow Statement (in '000 USD) - YTD")
    print(f"Output mapping file: {output_file}")
    
    try:
        # Perform complete Q1 verification
        matches = perform_complete_q1_verification()
        
        if not matches:
            print("ERROR: No Q1 verification matches found")
            return
        
        # Save complete mapping file
        save_complete_mapping_file(matches, output_file)
        
        # Summary by destination section
        section_breakdown = {}
        for match in matches:
            section = match['Dest_Major_Section_Context'] or 'Unknown'
            section_breakdown[section] = section_breakdown.get(section, 0) + 1
        
        print(f"\n" + "="*80)
        print("COMPLETE Q1 VERIFICATION RESULTS")
        print("="*80)
        print(f"Total Q1 verified matches: {len(matches)}")
        
        print(f"\nMatches by destination section:")
        for section, count in sorted(section_breakdown.items()):
            print(f"  {section}: {count} matches")
        
        # Show source coverage
        source_breakdown = {}
        for match in matches:
            source_sheet = match['Source_Sheet_Name']
            source_breakdown[source_sheet] = source_breakdown.get(source_sheet, 0) + 1
        
        print(f"\nMatches by source sheet:")
        for sheet, count in sorted(source_breakdown.items()):
            print(f"  {sheet}: {count} matches")
        
        print(f"\nComplete mapping file ready: {output_file}")
        print("This file contains Q1 verified matches from ALL financial statement sections")
        print("Ready for complete population of destination file!")
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
