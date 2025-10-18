#!/usr/bin/env python3
"""
Comprehensive Field Mapping for Key Metrics Tab

This script creates a detailed mapping of all fields in the Key Metrics tab
with full hierarchical scoping and enhanced context information.

Author: AI Assistant
Date: October 2025
"""

import openpyxl
import csv
from pathlib import Path
import re
from typing import Dict, List, Tuple, Optional


def analyze_key_metrics_structure(source_file: str) -> List[Dict]:
    """
    Analyze the Key Metrics sheet and create comprehensive field mappings.
    """
    print("Loading workbook and analyzing Key Metrics sheet...")
    wb = openpyxl.load_workbook(source_file, data_only=True)
    
    if 'Key Metrics' not in wb.sheetnames:
        print(f"ERROR: Key Metrics sheet not found. Available sheets: {wb.sheetnames}")
        wb.close()
        return []
    
    sheet = wb['Key Metrics']
    print(f"Key Metrics sheet has {sheet.max_row} rows and {sheet.max_column} columns")
    
    # Get all column headers (dates/periods)
    header_row = 4  # Based on the structure we saw earlier
    column_headers = []
    for col_idx in range(2, min(sheet.max_column + 1, 20)):  # Check first 20 columns
        header_val = sheet.cell(header_row, col_idx).value
        if header_val:
            column_headers.append({
                'col_idx': col_idx,
                'header': str(header_val),
                'clean_header': clean_date_header(str(header_val))
            })
    
    print(f"Found {len(column_headers)} data columns: {[h['clean_header'] for h in column_headers[:5]]}...")
    
    # Process all rows to build comprehensive mapping
    field_mappings = []
    
    # Context tracking
    current_section = None
    current_subsection = None
    
    for row_idx in range(1, sheet.max_row + 1):
        first_col = sheet.cell(row_idx, 1).value
        if not first_col:
            continue
            
        first_col_str = str(first_col).strip()
        
        # Skip header rows
        if (row_idx <= 4 or 
            first_col_str.lower().startswith('ipg photonics') or
            first_col_str.lower() in ['key metrics', 'quarter ended', 'cumulative quarter ended']):
            continue
        
        # Classify row type and update context
        row_type = classify_key_metrics_row(first_col_str, row_idx)
        
        if row_type == 'section_header':
            current_section = clean_field_name(first_col_str.rstrip(':'))
            current_subsection = None
            print(f"Section: {current_section}")
            
        elif row_type == 'subsection_header':
            current_subsection = clean_field_name(first_col_str.rstrip(':'))
            print(f"  Subsection: {current_subsection}")
            
        elif row_type == 'data_field':
            # Get all values for this row across columns
            row_values = {}
            for col_info in column_headers:
                col_idx = col_info['col_idx']
                value = sheet.cell(row_idx, col_idx).value
                row_values[col_info['clean_header']] = value
            
            # Create comprehensive field mapping
            field_name = clean_field_name(first_col_str)
            scoped_name = build_key_metrics_scoped_name(
                current_section, current_subsection, field_name
            )
            enhanced_scoped_name = apply_key_metrics_enhanced_scoping(
                scoped_name, field_name, current_section
            )
            
            field_mapping = {
                'row_number': row_idx,
                'original_field_name': first_col_str,
                'cleaned_field_name': field_name,
                'section_context': current_section or '',
                'subsection_context': current_subsection or '',
                'basic_scoped_name': scoped_name,
                'enhanced_scoped_name': enhanced_scoped_name,
                'row_type': row_type,
                'values': row_values,
                'has_data': any(v is not None and str(v).strip() for v in row_values.values())
            }
            
            field_mappings.append(field_mapping)
            
            # Print key fields for monitoring
            if any(keyword in first_col_str.lower() for keyword in ['total', 'revenue', 'china', 'germany', 'north america']):
                print(f"    Field: {first_col_str} → {enhanced_scoped_name}")
    
    wb.close()
    return field_mappings, column_headers


def classify_key_metrics_row(first_col: str, row_idx: int) -> str:
    """
    Classify Key Metrics row types.
    """
    first_col_lower = first_col.lower()
    
    # Section headers (major groupings)
    section_patterns = [
        r'revenue by', r'by region', r'by application', r'by product',
        r'long.?lived assets', r'employees by', r'backlog', r'customer'
    ]
    
    for pattern in section_patterns:
        if re.search(pattern, first_col_lower):
            return 'section_header'
    
    # Subsection headers (lines ending with colon)
    if first_col.endswith(':'):
        return 'subsection_header'
    
    # Data fields - anything else with meaningful content
    if (len(first_col) > 1 and 
        not first_col_lower.startswith('ipg photonics') and
        first_col_lower not in ['quarter ended', 'cumulative quarter ended']):
        return 'data_field'
    
    return 'other'


def clean_field_name(name: str) -> str:
    """Clean and standardize field names."""
    if not name:
        return ""
    
    # Remove extra whitespace and special characters
    cleaned = re.sub(r'\s+', '_', name.strip())
    cleaned = re.sub(r'[^\w\s-]', '', cleaned)
    cleaned = re.sub(r'[-_]+', '_', cleaned)
    cleaned = cleaned.strip('_')
    
    # Capitalize appropriately
    if cleaned:
        words = cleaned.split('_')
        cleaned = '_'.join(word.capitalize() for word in words)
    
    return cleaned


def clean_date_header(header: str) -> str:
    """Clean date headers for column names."""
    # Convert dates to readable format
    if 'datetime' in str(type(header)).lower():
        return header.strftime('%Y_Q%m')
    
    # Handle string dates
    header_str = str(header).strip()
    if re.match(r'\d{4}-\d{2}-\d{2}', header_str):
        year = header_str[:4]
        month = header_str[5:7]
        quarter = f"Q{(int(month) - 1) // 3 + 1}"
        return f"{year}_{quarter}"
    
    return header_str.replace('-', '_').replace(' ', '_')


def build_key_metrics_scoped_name(section: str, subsection: str, field: str) -> str:
    """Build scoped name for Key Metrics fields."""
    components = ['Key_Metrics']
    
    if section:
        components.append(section)
    if subsection:
        components.append(subsection)
    if field:
        components.append(field)
    
    return '.'.join(components)


def apply_key_metrics_enhanced_scoping(scoped_name: str, field_name: str, section: str) -> str:
    """Apply enhanced scoping rules specific to Key Metrics."""
    field_lower = field_name.lower()
    
    # Geographic regions - highest priority
    geographic_mapping = {
        'north_america': 'North_America',
        'united_states': 'United_States',
        'germany': 'Germany', 
        'other_europe': 'Other_Europe',
        'china': 'China',
        'japan': 'Japan',
        'other_asia': 'Other_Asia',
        'korea': 'Korea',
        'rest_of_world': 'Rest_Of_World'
    }
    
    for geo_key, geo_name in geographic_mapping.items():
        if geo_key.replace('_', ' ') in field_lower or geo_key in field_lower:
            if section and 'revenue' in section.lower():
                return f"Revenue_Statement.Geographic_Breakdown.{geo_name}"
            elif section and 'assets' in section.lower():
                return f"Balance_Sheet.Geographic_Assets.{geo_name}"
            elif section and 'employee' in section.lower():
                return f"Operations.Employee_Distribution.{geo_name}"
            else:
                return f"Key_Metrics.Geographic_Data.{geo_name}"
    
    # Product/Application categories
    product_mapping = {
        'materials_processing': 'Materials_Processing',
        'other_applications': 'Other_Applications',
        'advanced_applications': 'Advanced_Applications',
        'communications': 'Communications',
        'medical': 'Medical'
    }
    
    for prod_key, prod_name in product_mapping.items():
        if prod_key.replace('_', ' ') in field_lower:
            return f"Revenue_Statement.Application_Breakdown.{prod_name}"
    
    # Product type categories
    laser_mapping = {
        'high_power': 'High_Power_CW_Lasers',
        'pulsed': 'Pulsed_Lasers', 
        'qcw': 'QCW_Lasers',
        'systems': 'Systems_And_Other'
    }
    
    for laser_key, laser_name in laser_mapping.items():
        if laser_key.replace('_', ' ') in field_lower:
            return f"Revenue_Statement.Product_Type_Breakdown.{laser_name}"
    
    # Financial metrics
    if 'total' in field_lower:
        if section and 'revenue' in section.lower():
            return f"Revenue_Statement.Totals.{field_name}"
        elif section and 'assets' in section.lower():
            return f"Balance_Sheet.Asset_Totals.{field_name}"
        elif section and 'employee' in section.lower():
            return f"Operations.Employee_Totals.{field_name}"
        else:
            return f"Key_Metrics.Totals.{field_name}"
    
    # Customer and backlog metrics
    if any(term in field_lower for term in ['customer', 'backlog']):
        return f"Operations.Customer_Metrics.{field_name}"
    
    # Revenue growth and ratios
    if any(term in field_lower for term in ['growth', 'ratio', '%', 'percent']):
        return f"Financial_Ratios.Growth_Metrics.{field_name}"
    
    # Default enhanced scoping
    return scoped_name


def save_key_metrics_mapping(field_mappings: List[Dict], column_headers: List[Dict], output_file: str):
    """Save comprehensive Key Metrics mapping to CSV."""
    
    # Prepare dynamic column names for all date columns
    date_columns = [h['clean_header'] for h in column_headers]
    
    with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
        # Define fieldnames
        base_fieldnames = [
            'Row_Number',
            'Original_Field_Name',
            'Cleaned_Field_Name',
            'Section_Context',
            'Subsection_Context', 
            'Basic_Scoped_Name',
            'Enhanced_Scoped_Name',
            'Row_Type',
            'Has_Data'
        ]
        
        # Add all date columns
        fieldnames = base_fieldnames + date_columns
        
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        
        for mapping in field_mappings:
            # Build row data
            row_data = {
                'Row_Number': mapping['row_number'],
                'Original_Field_Name': mapping['original_field_name'],
                'Cleaned_Field_Name': mapping['cleaned_field_name'],
                'Section_Context': mapping['section_context'],
                'Subsection_Context': mapping['subsection_context'],
                'Basic_Scoped_Name': mapping['basic_scoped_name'],
                'Enhanced_Scoped_Name': mapping['enhanced_scoped_name'],
                'Row_Type': mapping['row_type'],
                'Has_Data': 'Yes' if mapping['has_data'] else 'No'
            }
            
            # Add values for each date column
            for date_col in date_columns:
                value = mapping['values'].get(date_col)
                if value is not None:
                    row_data[date_col] = str(value)
                else:
                    row_data[date_col] = ''
            
            writer.writerow(row_data)


def main():
    """Main entry point."""
    source_file = "/Users/michaelkim/code/Bernstein/IPGP-Financial-Data-Workbook-2024-Q2.xlsx"
    output_file = "/Users/michaelkim/code/Bernstein/key_metrics_comprehensive_mapping.csv"
    
    print("="*80)
    print("KEY METRICS COMPREHENSIVE FIELD MAPPING")
    print("="*80)
    print(f"Source file: {source_file}")
    print(f"Output file: {output_file}")
    
    if not Path(source_file).exists():
        print(f"ERROR: Source file not found: {source_file}")
        return
    
    try:
        # Analyze Key Metrics sheet
        field_mappings, column_headers = analyze_key_metrics_structure(source_file)
        
        if not field_mappings:
            print("ERROR: No field mappings generated")
            return
        
        print(f"\n" + "="*50)
        print("RESULTS SUMMARY")
        print("="*50)
        print(f"Total fields mapped: {len(field_mappings)}")
        print(f"Data columns found: {len(column_headers)}")
        
        # Count by section
        sections = {}
        for mapping in field_mappings:
            section = mapping['section_context'] or 'No_Section'
            sections[section] = sections.get(section, 0) + 1
        
        print("\nFields by section:")
        for section, count in sorted(sections.items()):
            print(f"  {section}: {count} fields")
        
        # Save to CSV
        print(f"\n" + "="*50)
        print("SAVING TO CSV")
        print("="*50)
        save_key_metrics_mapping(field_mappings, column_headers, output_file)
        
        print(f"✓ Key Metrics comprehensive mapping saved to: {output_file}")
        print(f"✓ Contains {len(field_mappings)} fields with full hierarchical context")
        print(f"✓ Includes data values across {len(column_headers)} time periods")
        
        print(f"\n" + "="*80)
        print("PROCESSING COMPLETE!")
        print("="*80)
        print(f"Review the comprehensive Key Metrics mapping in: {output_file}")
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
