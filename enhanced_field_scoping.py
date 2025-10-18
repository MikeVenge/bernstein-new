#!/usr/bin/env python3
"""
Enhanced Field Scoping Processor for IPGP Financial Data

This script creates fully contextualized field names with proper hierarchical scoping.
For example: "Total" becomes "CashFlow_Statement.Global_Revenues.Total"

Strategy:
1. Parse the Source tab to understand hierarchical structure
2. Build context hierarchy (Statement -> Section -> Subsection -> Field)
3. Apply sophisticated scoping rules based on financial statement structure
4. Output fully scoped field names to CSV for review

Author: AI Assistant
Date: October 2025
"""

import openpyxl
import csv
from pathlib import Path
import re
from typing import Dict, List, Tuple, Optional


def analyze_source_tab_structure(source_file: str) -> Dict:
    """
    Analyze all financial statement sheets to understand the hierarchical structure.
    Returns a mapping of the document structure.
    """
    print("Loading source workbook...")
    wb = openpyxl.load_workbook(source_file, data_only=True)
    
    # Process all relevant financial statement sheets
    relevant_sheets = ['Income Statement', 'Balance Sheet', 'Cash Flows', 'Key Metrics']
    available_sheets = [name for name in relevant_sheets if name in wb.sheetnames]
    
    print(f"Available financial statement sheets: {available_sheets}")
    
    if not available_sheets:
        print(f"No relevant sheets found. Available sheets: {wb.sheetnames}")
        return {'sheets': []}
    
    # Analyze all relevant sheets
    all_sheets_structure = {'sheets': []}
    
    for sheet_name in available_sheets:
        source_sheet = wb[sheet_name]
        
        # Analyze structure for this sheet
        structure = {
            'sheet_name': source_sheet.title,
            'max_row': source_sheet.max_row,
            'max_col': source_sheet.max_column,
            'rows': []
        }
        
        print(f"Analyzing {sheet_name}: {structure['max_row']} rows...")
        
        for row_idx in range(1, min(structure['max_row'] + 1, 200)):  # Limit to first 200 rows per sheet
            row_data = {
                'row': row_idx,
                'sheet': sheet_name,
                'values': [],
                'has_content': False
            }
            
            # Get first 20 columns to understand structure
            for col_idx in range(1, min(21, structure['max_col'] + 1)):
                cell_value = source_sheet.cell(row_idx, col_idx).value
                row_data['values'].append(cell_value)
                if cell_value is not None and str(cell_value).strip():
                    row_data['has_content'] = True
            
            if row_data['has_content']:
                structure['rows'].append(row_data)
        
        all_sheets_structure['sheets'].append(structure)
    
    wb.close()
    return all_sheets_structure


def classify_row_type(values: List, row_idx: int) -> str:
    """
    Classify what type of row this is based on content patterns.
    Returns: 'statement_header', 'section_header', 'subsection_header', 'data_field', 'empty', 'other'
    """
    if not values or all(v is None or str(v).strip() == "" for v in values):
        return 'empty'
    
    first_col = str(values[0]).strip() if values[0] is not None else ""
    second_col = str(values[1]).strip() if len(values) > 1 and values[1] is not None else ""
    
    # Statement headers (major sections)
    statement_patterns = [
        r'income\s+statement', r'statement\s+of\s+operations', r'profit\s+and\s+loss',
        r'balance\s+sheet', r'statement\s+of\s+financial\s+position',
        r'cash\s+flow', r'statement\s+of\s+cash\s+flows',
        r'statement\s+of\s+equity', r'shareholders?\s+equity'
    ]
    
    for pattern in statement_patterns:
        if re.search(pattern, first_col.lower()):
            return 'statement_header'
    
    # Section headers (major groupings within statements)
    section_patterns = [
        r'revenue', r'net\s+sales', r'total\s+revenue',
        r'cost\s+of', r'operating\s+expenses', r'operating\s+income',
        r'assets', r'current\s+assets', r'non[-\s]?current\s+assets',
        r'liabilities', r'current\s+liabilities', r'equity',
        r'operating\s+activities', r'investing\s+activities', r'financing\s+activities',
        r'by\s+region', r'by\s+product', r'by\s+application', r'by\s+market'
    ]
    
    for pattern in section_patterns:
        if re.search(pattern, first_col.lower()):
            return 'section_header'
    
    # Subsection headers (typically end with colon or are indented)
    if first_col.endswith(':') or (len(first_col) > 5 and first_col[0].isupper() and not any(char.isdigit() for char in first_col)):
        return 'subsection_header'
    
    # Data fields - be more liberal in detection
    # 1. Fields that look like financial metrics
    if first_col and any(keyword in first_col.lower() for keyword in [
        'revenue', 'sales', 'income', 'expense', 'cost', 'asset', 'liability', 
        'cash', 'total', 'north america', 'germany', 'china', 'japan', 'europe',
        'materials processing', 'other', 'communications', 'medical', 'advanced',
        'high-power', 'pulsed', 'qcw', 'systems', 'basic', 'diluted'
    ]):
        return 'data_field'
    
    # 2. Fields with numerical values in subsequent columns (original logic)
    has_numbers = False
    for val in values[1:]:
        if val is not None:
            val_str = str(val).strip()
            if val_str and (isinstance(val, (int, float)) or re.search(r'[\d,]+', val_str)):
                has_numbers = True
                break
    
    if has_numbers and first_col:
        return 'data_field'
    
    # 3. Any row with a meaningful first column that's not clearly a header
    if (first_col and len(first_col) > 2 and 
        not first_col.endswith(':') and 
        not first_col.lower().startswith('ipg photonics') and
        not first_col.lower() in ['quarter ended', 'cumulative quarter ended']):
        return 'data_field'
    
    return 'other'


def build_hierarchical_context(structure: Dict) -> List[Dict]:
    """
    Build hierarchical context for each row based on the structure analysis.
    """
    contextualized_rows = []
    
    print("Building hierarchical context...")
    
    # Process each sheet
    for sheet_structure in structure['sheets']:
        sheet_name = sheet_structure['sheet_name']
        print(f"Processing sheet: {sheet_name}")
        
        # Current context stack for this sheet
        current_statement = clean_field_name(sheet_name)  # Use sheet name as statement
        current_section = None
        current_subsection = None
        
        for row_data in sheet_structure['rows']:
            row_idx = row_data['row']
            values = row_data['values']
            row_type = classify_row_type(values, row_idx)
            
            first_col = str(values[0]).strip() if values[0] is not None else ""
            
            # Update context based on row type
            if row_type == 'statement_header':
                current_statement = clean_field_name(first_col)
                current_section = None
                current_subsection = None
            elif row_type == 'section_header':
                current_section = clean_field_name(first_col)
                current_subsection = None
            elif row_type == 'subsection_header':
                current_subsection = clean_field_name(first_col.rstrip(':'))
            
            # Build scoped name for data fields
            if row_type == 'data_field' and first_col:
                field_name = clean_field_name(first_col)
                scoped_name = build_scoped_name(
                    current_statement, 
                    current_section, 
                    current_subsection, 
                    field_name
                )
                
                contextualized_row = {
                    'row': row_idx,
                    'sheet': sheet_name,
                    'original_name': first_col,
                    'field_name': field_name,
                    'scoped_name': scoped_name,
                    'statement': current_statement,
                    'section': current_section,
                    'subsection': current_subsection,
                    'row_type': row_type,
                    'values': values
                }
                contextualized_rows.append(contextualized_row)
    
    return contextualized_rows


def clean_field_name(name: str) -> str:
    """Clean and standardize field names for scoping."""
    if not name:
        return ""
    
    # Remove extra whitespace and special characters
    cleaned = re.sub(r'\s+', '_', name.strip())
    cleaned = re.sub(r'[^\w\s-]', '', cleaned)
    cleaned = re.sub(r'[-_]+', '_', cleaned)
    
    # Remove trailing underscores
    cleaned = cleaned.strip('_')
    
    # Capitalize appropriately
    if cleaned:
        cleaned = '_'.join(word.capitalize() for word in cleaned.split('_'))
    
    return cleaned


def build_scoped_name(statement: str, section: str, subsection: str, field: str) -> str:
    """
    Build a fully scoped field name with hierarchical context.
    Example: CashFlow_Statement.Global_Revenues.Total
    """
    components = []
    
    # Add statement level
    if statement:
        components.append(statement)
    
    # Add section level
    if section:
        components.append(section)
    
    # Add subsection level
    if subsection:
        components.append(subsection)
    
    # Add field name
    if field:
        components.append(field)
    
    # If no context, just use the field name
    if not components:
        return field if field else "Unknown_Field"
    
    return '.'.join(components)


def enhance_scoping_with_patterns(contextualized_rows: List[Dict]) -> List[Dict]:
    """
    Apply additional scoping rules based on financial statement patterns.
    """
    print("Applying enhanced scoping patterns...")
    
    enhanced_rows = []
    
    for row in contextualized_rows:
        enhanced_row = row.copy()
        original_scoped = row['scoped_name']
        field_name = row['field_name'].lower()
        
        # Apply specific financial statement scoping rules
        enhanced_scoped = apply_financial_scoping_rules(
            original_scoped, 
            field_name, 
            row['statement'], 
            row['section'], 
            row['subsection']
        )
        
        enhanced_row['enhanced_scoped_name'] = enhanced_scoped
        enhanced_rows.append(enhanced_row)
    
    return enhanced_rows


def apply_financial_scoping_rules(scoped_name: str, field_name: str, statement: str, section: str, subsection: str) -> str:
    """
    Apply specific financial statement scoping rules.
    """
    # Geographic regions
    geographic_terms = ['north_america', 'united_states', 'germany', 'china', 'japan', 'korea', 'europe', 'asia']
    if any(geo in field_name for geo in geographic_terms):
        if 'revenue' in scoped_name.lower() or 'sales' in scoped_name.lower():
            return f"Revenue_Statement.Geographic_Breakdown.{field_name.title().replace('_', '_')}"
    
    # Product categories
    product_terms = ['materials_processing', 'high_power', 'pulsed', 'qcw', 'systems']
    if any(prod in field_name for prod in product_terms):
        return f"Revenue_Statement.Product_Breakdown.{field_name.title().replace('_', '_')}"
    
    # Cash flow activities
    if statement and 'cash' in statement.lower():
        if 'operating' in scoped_name.lower():
            return f"CashFlow_Statement.Operating_Activities.{field_name.title().replace('_', '_')}"
        elif 'investing' in scoped_name.lower():
            return f"CashFlow_Statement.Investing_Activities.{field_name.title().replace('_', '_')}"
        elif 'financing' in scoped_name.lower():
            return f"CashFlow_Statement.Financing_Activities.{field_name.title().replace('_', '_')}"
    
    # Balance sheet items
    if statement and 'balance' in statement.lower():
        if any(term in field_name for term in ['asset', 'cash', 'inventory', 'receivable']):
            return f"Balance_Sheet.Assets.{field_name.title().replace('_', '_')}"
        elif any(term in field_name for term in ['liability', 'payable', 'debt']):
            return f"Balance_Sheet.Liabilities.{field_name.title().replace('_', '_')}"
        elif any(term in field_name for term in ['equity', 'retained', 'capital']):
            return f"Balance_Sheet.Equity.{field_name.title().replace('_', '_')}"
    
    # Income statement items
    if statement and ('income' in statement.lower() or 'operations' in statement.lower()):
        if any(term in field_name for term in ['revenue', 'sales', 'net_sales']):
            return f"Income_Statement.Revenue.{field_name.title().replace('_', '_')}"
        elif any(term in field_name for term in ['cost', 'expense']):
            return f"Income_Statement.Expenses.{field_name.title().replace('_', '_')}"
        elif 'total' in field_name:
            return f"Income_Statement.Totals.{field_name.title().replace('_', '_')}"
    
    # Default: return enhanced version of original
    return scoped_name


def process_source_tab(source_file: str, output_file: str):
    """
    Main processing function to create fully scoped field names from Source tab.
    """
    print("="*80)
    print("ENHANCED FIELD SCOPING PROCESSOR")
    print("="*80)
    print(f"Source file: {source_file}")
    print(f"Output file: {output_file}")
    
    # Step 1: Analyze structure
    print("\n" + "="*50)
    print("STEP 1: ANALYZING SOURCE TAB STRUCTURE")
    print("="*50)
    structure = analyze_source_tab_structure(source_file)
    total_rows = sum(len(sheet['rows']) for sheet in structure['sheets'])
    print(f"Found {len(structure['sheets'])} sheets with {total_rows} total content rows")
    
    # Step 2: Build hierarchical context
    print("\n" + "="*50)
    print("STEP 2: BUILDING HIERARCHICAL CONTEXT")
    print("="*50)
    contextualized_rows = build_hierarchical_context(structure)
    print(f"Identified {len(contextualized_rows)} data fields")
    
    # Step 3: Apply enhanced scoping
    print("\n" + "="*50)
    print("STEP 3: APPLYING ENHANCED SCOPING RULES")
    print("="*50)
    enhanced_rows = enhance_scoping_with_patterns(contextualized_rows)
    
    # Step 4: Output to CSV
    print("\n" + "="*50)
    print("STEP 4: GENERATING CSV OUTPUT")
    print("="*50)
    
    with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
        fieldnames = [
            'Sheet_Name',
            'Row_Number',
            'Original_Field_Name',
            'Cleaned_Field_Name', 
            'Basic_Scoped_Name',
            'Enhanced_Scoped_Name',
            'Statement_Context',
            'Section_Context',
            'Subsection_Context',
            'Row_Type',
            'Sample_Values'
        ]
        
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        
        for row in enhanced_rows:
            # Get sample values (first few non-empty values)
            sample_values = []
            for val in row['values'][1:6]:  # Skip first column (field name), take next 5
                if val is not None and str(val).strip():
                    sample_values.append(str(val).strip())
            
            writer.writerow({
                'Sheet_Name': row['sheet'],
                'Row_Number': row['row'],
                'Original_Field_Name': row['original_name'],
                'Cleaned_Field_Name': row['field_name'],
                'Basic_Scoped_Name': row['scoped_name'],
                'Enhanced_Scoped_Name': row['enhanced_scoped_name'],
                'Statement_Context': row['statement'] or '',
                'Section_Context': row['section'] or '',
                'Subsection_Context': row['subsection'] or '',
                'Row_Type': row['row_type'],
                'Sample_Values': ' | '.join(sample_values[:3])  # First 3 sample values
            })
    
    print(f"✓ Enhanced scoping results saved to: {output_file}")
    print(f"✓ Processed {len(enhanced_rows)} fields with full hierarchical context")
    
    # Print summary statistics
    print("\n" + "="*50)
    print("SUMMARY STATISTICS")
    print("="*50)
    
    statement_counts = {}
    for row in enhanced_rows:
        stmt = row['statement'] or 'Unknown'
        statement_counts[stmt] = statement_counts.get(stmt, 0) + 1
    
    print("Fields by Statement Type:")
    for stmt, count in sorted(statement_counts.items()):
        print(f"  {stmt}: {count} fields")
    
    print(f"\nTotal fields processed: {len(enhanced_rows)}")
    print(f"CSV file ready for review: {output_file}")


def main():
    """Main entry point."""
    # File paths
    source_file = "/Users/michaelkim/code/Bernstein/IPGP-Financial-Data-Workbook-2024-Q2.xlsx"
    output_file = "/Users/michaelkim/code/Bernstein/enhanced_field_scoping_results.csv"
    
    # Check if source file exists
    if not Path(source_file).exists():
        print(f"ERROR: Source file not found: {source_file}")
        print("Available files:")
        for file in Path("/Users/michaelkim/code/Bernstein").glob("*.xlsx"):
            print(f"  {file}")
        return
    
    try:
        process_source_tab(source_file, output_file)
        print("\n" + "="*80)
        print("PROCESSING COMPLETE!")
        print("="*80)
        print(f"Review the results in: {output_file}")
        print("You can now examine the enhanced scoped field names.")
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
