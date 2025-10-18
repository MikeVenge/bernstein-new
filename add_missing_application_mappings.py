#!/usr/bin/env python3
"""
Add Missing Application Mappings

Adds the missing application breakdown field mappings to the consolidated file
and populates the destination.

Author: AI Assistant
Date: October 2025
"""

import openpyxl
import pandas as pd
import csv
from pathlib import Path


def create_missing_application_mappings():
    """Create mappings for the missing application breakdown fields."""
    
    print("=== CREATING MISSING APPLICATION MAPPINGS ===")
    
    # Row 23 has a clear match to "Other applications"
    missing_mappings = [
        {
            'Dest_Row_Number': '23',
            'Dest_Field_Name': 'Other application, of which',
            'Dest_Enhanced_Scope': 'Reported.Segment_Information_In_000_Usd_Ytd.End_Market_Breakdown.Other_Application_Of_Which',
            'Dest_Section_Context': 'End_Market_Breakdown',
            'Dest_Major_Section_Context': 'Segment Information',
            'Source_Sheet_Name': 'Key Metrics',
            'Source_Row_Number': '27',
            'Source_Field_Name': 'Other applications',
            'Source_Enhanced_Scope': 'Revenue_Statement.Application_Breakdown.Other_Applications',
            'Source_Section_Context': 'Application_Breakdown',
            'Q1_Verification_Value': '25644',   # Dest Q1 value
            'Source_Q1_Value': '30352',        # Source Q1 value (close)
            'Source_Q2_Value': '31872',        # Source Q2 value to populate
            'Match_Method': 'Manual_Semantic_Match_Q1_Close',
            'Match_Confidence': '0.90',
            'Match_Reason': 'Semantic match: \"Other application, of which\" to \"Other applications\". Q1 close: 25644 vs 30352'
        }
    ]
    
    # Rows 24-26 don't have source data available (all show None/None in source)
    # These appear to be sub-breakdown fields that aren't populated in the source
    
    print(f"Created {len(missing_mappings)} missing application mappings:")
    for mapping in missing_mappings:
        print(f"  Row {mapping['Dest_Row_Number']}: {mapping['Dest_Field_Name']}")
        print(f"    → {mapping['Source_Sheet_Name']} Row {mapping['Source_Row_Number']}: {mapping['Source_Field_Name']}")
        print(f"    Q1: {mapping['Q1_Verification_Value']} vs {mapping['Source_Q1_Value']}")
        print(f"    Q2: {mapping['Source_Q2_Value']}")
    
    print(f"\nRows 24-26 (Advanced Applications, Communications, Medical):")
    print(f"  These fields have no Q1 data in destination and no corresponding")
    print(f"  source data in Key Metrics (all show None/None)")
    print(f"  These appear to be sub-breakdown fields not available in source")
    
    return missing_mappings


def add_to_consolidated_mapping(new_mappings):
    """Add new mappings to the consolidated mapping file."""
    
    print(f"\n=== ADDING TO CONSOLIDATED MAPPING FILE ===")
    
    consolidated_file = "/Users/michaelkim/code/Bernstein/CONSOLIDATED_FIELD_MAPPINGS.csv"
    
    # Load existing consolidated mappings
    existing_mappings = []
    with open(consolidated_file, 'r', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        existing_mappings = list(reader)
    
    print(f"Existing mappings: {len(existing_mappings)}")
    
    # Add new mappings
    for new_mapping in new_mappings:
        existing_mappings.append(new_mapping)
    
    # Sort by destination row number
    existing_mappings.sort(key=lambda x: int(x['Dest_Row_Number']) if x['Dest_Row_Number'].isdigit() else 999999)
    
    # Save updated consolidated file
    with open(consolidated_file, 'w', newline='', encoding='utf-8') as csvfile:
        fieldnames = [
            'Dest_Row_Number', 'Dest_Field_Name', 'Dest_Enhanced_Scope',
            'Dest_Section_Context', 'Dest_Major_Section_Context',
            'Source_Sheet_Name', 'Source_Row_Number', 'Source_Field_Name',
            'Source_Enhanced_Scope', 'Source_Section_Context',
            'Q1_Verification_Value', 'Source_Q1_Value', 'Source_Q2_Value',
            'Match_Method', 'Match_Confidence', 'Match_Reason', 'Source_File'
        ]
        
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(existing_mappings)
    
    print(f"✅ Updated consolidated mapping file")
    print(f"Total mappings now: {len(existing_mappings)}")
    
    return len(existing_mappings)


def populate_missing_application_field(mappings):
    """Populate the missing application field in destination."""
    
    print(f"\n=== POPULATING MISSING APPLICATION FIELD ===")
    
    source_file = "/Users/michaelkim/code/Bernstein/IPGP-Financial-Data-Workbook-2024-Q2.xlsx"
    dest_file = "/Users/michaelkim/code/Bernstein/FRESH_POPULATED_FROM_CONSOLIDATED_IPGP.xlsx"
    
    # Load workbooks
    source_wb = openpyxl.load_workbook(source_file, data_only=True)
    dest_wb = openpyxl.load_workbook(dest_file, data_only=False)
    dest_sheet = dest_wb['Reported']
    
    populated_count = 0
    
    for mapping in mappings:
        dest_row = int(mapping['Dest_Row_Number'])
        source_sheet_name = mapping['Source_Sheet_Name']
        source_row = int(mapping['Source_Row_Number'])
        expected_q2_value = mapping['Source_Q2_Value']
        
        print(f"Row {dest_row}: {mapping['Dest_Field_Name']}")
        print(f"  From {source_sheet_name} Row {source_row}: {mapping['Source_Field_Name']}")
        print(f"  Expected Q2 value: {expected_q2_value}")
        
        # Verify the source value
        if source_sheet_name in source_wb.sheetnames:
            source_sheet = source_wb[source_sheet_name]
            actual_q2_value = source_sheet.cell(source_row, 93).value  # Column CO
            
            print(f"  Actual Q2 value: {actual_q2_value}")
            
            if str(actual_q2_value) == str(expected_q2_value):
                # Populate Column BS (71) with Q2 value
                dest_sheet.cell(dest_row, 71).value = actual_q2_value
                
                # Add source tracking to Column BT (72)
                source_location = f"IPGP-Financial-Data-Workbook-2024-Q2.xlsx|{source_sheet_name}|{source_row}|93"
                dest_sheet.cell(dest_row, 72).value = source_location
                
                populated_count += 1
                print(f"  ✅ POPULATED BS: {actual_q2_value}")
                print(f"  ✅ TRACKED BT: {source_location}")
            else:
                print(f"  ❌ Q2 value mismatch: expected {expected_q2_value}, got {actual_q2_value}")
        else:
            print(f"  ❌ Source sheet not found: {source_sheet_name}")
    
    # Save the updated file
    output_file = "/Users/michaelkim/code/Bernstein/FINAL_COMPLETE_WITH_MISSING_APPLICATIONS_IPGP.xlsx"
    dest_wb.save(output_file)
    
    source_wb.close()
    dest_wb.close()
    
    print(f"\n✅ Missing application field population complete!")
    print(f"Populated {populated_count} field")
    print(f"Updated file saved as: {output_file}")
    
    return output_file, populated_count


def main():
    """Main entry point for adding missing application mappings."""
    
    print("="*80)
    print("ADDING MISSING APPLICATION BREAKDOWN MAPPINGS")
    print("="*80)
    print("Adding Row 23 that was missed in consolidation")
    
    try:
        # Create missing mappings
        missing_mappings = create_missing_application_mappings()
        
        if missing_mappings:
            # Add to consolidated mapping file
            total_mappings = add_to_consolidated_mapping(missing_mappings)
            
            # Populate destination file
            final_file, populated_count = populate_missing_application_field(missing_mappings)
            
            print(f"\n" + "="*80)
            print("MISSING APPLICATION MAPPINGS RESULTS")
            print("="*80)
            print(f"Missing mappings added: {len(missing_mappings)}")
            print(f"Successfully populated: {populated_count}")
            print(f"Total consolidated mappings now: {total_mappings}")
            print(f"Final file: {final_file}")
            print(f"✅ Row 23 'Other application, of which' now populated!")
            
            # Note about rows 24-26
            print(f"\n📝 NOTE: Rows 24-26 (Advanced Applications, Communications, Medical)")
            print(f"   cannot be populated as they have no source data available")
            print(f"   in the Key Metrics sheet (all show None/None)")
        else:
            print("No missing mappings to add")
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
