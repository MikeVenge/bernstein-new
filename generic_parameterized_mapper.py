#!/usr/bin/env python3
"""
Generic Parameterized Field Mapper

A flexible field mapper that works with generic mapping files.
Supports different source files, destination files, time periods, and companies.

Usage:
    python generic_parameterized_mapper.py --source SOURCE_FILE --destination DEST_FILE 
                                          --mapping GENERIC_MAPPING_FILE --column COLUMN_NUMBER
                                          [--data-column DATA_COLUMN] [--output OUTPUT_FILE]

Author: AI Assistant
Date: October 2025
"""

import openpyxl
import pandas as pd
import csv
import argparse
from pathlib import Path
from typing import Dict, List, Optional, Tuple


class GenericParameterizedMapper:
    """Generic parameterized field mapper that works with any dataset."""
    
    def __init__(self, source_file: str, destination_file: str, mapping_file: str, 
                 target_column: int, data_column: str = "CO", output_file: Optional[str] = None):
        """
        Initialize the generic parameterized mapper.
        
        Args:
            source_file: Path to source Excel file
            destination_file: Path to destination Excel file  
            mapping_file: Path to generic CSV mapping file
            target_column: DESTINATION column number to populate (1-based)
            data_column: SOURCE column reference to read data from (default: "CO")
            output_file: Optional output file path (default: auto-generated)
        """
        self.source_file = Path(source_file)
        self.destination_file = Path(destination_file)
        self.mapping_file = Path(mapping_file)
        self.target_column = target_column
        self.source_tracking_column = target_column + 1
        self.data_column = data_column
        
        # Convert column reference to number (CO = 93)
        self.source_column_num = self._column_ref_to_number(data_column)
        
        if output_file:
            self.output_file = Path(output_file)
        else:
            self.output_file = self.destination_file.parent / f"generic_populated_{self.destination_file.name}"
        
        self.stats = {
            'mappings_processed': 0,
            'values_populated': 0,
            'transformations_applied': 0,
            'errors': []
        }
    
    def _column_ref_to_number(self, column_ref: str) -> int:
        """Convert column reference (like 'CO') to column number."""
        if column_ref == 'CO':
            return 93
        elif column_ref == 'BR':
            return 70
        elif len(column_ref) == 2:
            # Convert two-letter column reference to number
            first = ord(column_ref[0]) - ord('A') + 1
            second = ord(column_ref[1]) - ord('A') + 1
            return first * 26 + second
        else:
            return 93  # Default to CO
    
    def load_generic_mapping(self) -> List[Dict]:
        """Load generic mapping file."""
        
        print(f"=== LOADING GENERIC MAPPING ===")
        print(f"Mapping file: {self.mapping_file}")
        
        if not self.mapping_file.exists():
            raise FileNotFoundError(f"Generic mapping file not found: {self.mapping_file}")
        
        mappings = []
        with open(self.mapping_file, 'r', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            mappings = list(reader)
        
        print(f"Loaded {len(mappings)} generic mappings")
        
        # Show breakdown
        mapping_types = {}
        transformations = {}
        
        for mapping in mappings:
            map_type = mapping.get('Mapping_Type', 'Unknown')
            transform = mapping.get('Transformation', 'Unknown')
            
            mapping_types[map_type] = mapping_types.get(map_type, 0) + 1
            transformations[transform] = transformations.get(transform, 0) + 1
        
        print(f"Mapping types:")
        for map_type, count in sorted(mapping_types.items()):
            print(f"  {map_type}: {count}")
        
        print(f"Transformations:")
        for transform, count in sorted(transformations.items()):
            print(f"  {transform}: {count}")
        
        return mappings
    
    def apply_transformation(self, value, transformation_type: str) -> any:
        """Apply transformation to the value based on type."""
        
        if transformation_type == 'DIRECT_COPY':
            return value
        elif transformation_type == 'PERCENTAGE_VALUE':
            return value  # Keep as decimal
        elif transformation_type == 'SUM_FIELDS':
            return value  # Already summed in processing
        elif transformation_type == 'ZERO_VALUE':
            return 0 if value is None else value
        else:
            return value
    
    def process_mapping(self, mapping: Dict, source_wb: openpyxl.Workbook, 
                       dest_sheet: openpyxl.worksheet.worksheet.Worksheet) -> Dict:
        """Process a single generic mapping."""
        
        dest_row = int(mapping['Dest_Row'])
        source_sheet_name = mapping['Source_Sheet']
        source_row = mapping['Source_Row']
        transformation_type = mapping.get('Transformation', 'DIRECT_COPY')
        
        result = {
            'Dest_Row': dest_row,
            'Dest_Field': mapping.get('Dest_Field', ''),
            'Source_Sheet': source_sheet_name,
            'Source_Row': source_row,
            'Source_Field': mapping.get('Source_Field', ''),
            'Mapping_Type': mapping.get('Mapping_Type', ''),
            'Transformation': transformation_type,
            'Status': 'ERROR'
        }
        
        try:
            # Validate source sheet
            if source_sheet_name not in source_wb.sheetnames:
                result['Status'] = 'SOURCE_SHEET_NOT_FOUND'
                return result
            
            source_sheet = source_wb[source_sheet_name]
            
            # Handle composite fields (multiple source rows)
            if '+' in str(source_row):
                composite_rows = [int(r.strip()) for r in str(source_row).split('+')]
                composite_value = 0
                
                for comp_row in composite_rows:
                    comp_value = source_sheet.cell(comp_row, self.source_column_num).value or 0
                    composite_value += comp_value
                
                source_value = composite_value
                source_location = f"{self.source_file.name}|{source_sheet_name}|{source_row}|{self.data_column}"
                result['Status'] = 'COMPOSITE_PROCESSED'
                self.stats['transformations_applied'] += 1
            else:
                # Single source row
                source_row_num = int(source_row)
                source_value = source_sheet.cell(source_row_num, self.source_column_num).value
                source_location = f"{self.source_file.name}|{source_sheet_name}|{source_row_num}|{self.data_column}"
                result['Status'] = 'PROCESSED'
            
            # Apply transformation
            transformed_value = self.apply_transformation(source_value, transformation_type)
            result['Value'] = transformed_value
            result['Source_Location'] = source_location
            
            # Populate destination
            if transformed_value is not None:
                dest_sheet.cell(dest_row, self.target_column).value = transformed_value
                dest_sheet.cell(dest_row, self.source_tracking_column).value = source_location
                
                self.stats['values_populated'] += 1
                result['Status'] = 'POPULATED'
            else:
                result['Status'] = 'NO_SOURCE_DATA'
                
        except Exception as e:
            result['Status'] = 'ERROR'
            result['Error'] = str(e)
            self.stats['errors'].append(f"Row {dest_row}: {str(e)}")
        
        return result
    
    def run_generic_mapping(self) -> bool:
        """Run the complete generic field mapping process."""
        
        print("="*80)
        print("GENERIC PARAMETERIZED FIELD MAPPER")
        print("="*80)
        print("Reusable mapper for any dataset with same field relationships")
        
        try:
            # Load generic mapping
            mappings = self.load_generic_mapping()
            
            print(f"\n=== PROCESSING MAPPINGS ===")
            print(f"Source: {self.source_file}")
            print(f"Destination: {self.destination_file}")
            print(f"Target column: {self.target_column}")
            print(f"Data column: {self.data_column} (column {self.source_column_num})")
            
            # Load workbooks
            source_wb = openpyxl.load_workbook(self.source_file, data_only=True)
            dest_wb = openpyxl.load_workbook(self.destination_file, data_only=False)
            dest_sheet = dest_wb['Reported']  # Assume Reported sheet
            
            results = []
            
            for i, mapping in enumerate(mappings, 1):
                result = self.process_mapping(mapping, source_wb, dest_sheet)
                results.append(result)
                
                if i % 20 == 0 or i <= 5:
                    status_icon = "✅" if result['Status'] == 'POPULATED' else "❌"
                    print(f"[{i}/{len(mappings)}] {status_icon} Row {result['Dest_Row']}: {result.get('Value', 'N/A')}")
            
            # Save output file
            dest_wb.save(self.output_file)
            source_wb.close()
            dest_wb.close()
            
            self.stats['mappings_processed'] = len(mappings)
            
            # Generate summary
            self._print_summary(results)
            
            return True
            
        except Exception as e:
            print(f"ERROR: {e}")
            return False
    
    def _print_summary(self, results: List[Dict]):
        """Print execution summary."""
        
        print(f"\n" + "="*80)
        print("GENERIC MAPPING EXECUTION RESULTS")
        print("="*80)
        
        print(f"PARAMETERS:")
        print(f"  Source file: {self.source_file}")
        print(f"  Destination file: {self.destination_file}")
        print(f"  Target column: {self.target_column}")
        print(f"  Data column: {self.data_column}")
        
        print(f"\nRESULTS:")
        print(f"  Mappings processed: {self.stats['mappings_processed']}")
        print(f"  Values populated: {self.stats['values_populated']}")
        print(f"  Transformations applied: {self.stats['transformations_applied']}")
        
        if self.stats['mappings_processed'] > 0:
            success_rate = self.stats['values_populated'] / self.stats['mappings_processed'] * 100
            print(f"  Success rate: {success_rate:.1f}%")
        
        # Status breakdown
        status_counts = {}
        for result in results:
            status = result['Status']
            status_counts[status] = status_counts.get(status, 0) + 1
        
        print(f"\nStatus breakdown:")
        for status, count in sorted(status_counts.items()):
            print(f"  {status}: {count}")
        
        if self.stats['errors']:
            print(f"\nErrors: {len(self.stats['errors'])}")
            for error in self.stats['errors'][:3]:
                print(f"  {error}")
        
        print(f"\nOutput file: {self.output_file}")


def parse_arguments():
    """Parse command line arguments for generic mapper."""
    
    parser = argparse.ArgumentParser(
        description='Generic Parameterized Field Mapper - Works with any dataset',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    
    parser.add_argument('--source', '-s', required=True, help='Source Excel file')
    parser.add_argument('--destination', '-d', required=True, help='Destination Excel file')
    parser.add_argument('--mapping', '-m', required=True, help='Generic mapping CSV file')
    parser.add_argument('--column', '-c', type=int, required=True, help='DESTINATION column number to populate (1-based)')
    parser.add_argument('--data-column', default='CO', help='SOURCE column reference to read data from (default: CO)')
    parser.add_argument('--output', '-o', help='Output file path')
    
    return parser.parse_args()


def main():
    """Main entry point for generic parameterized mapper."""
    
    args = parse_arguments()
    
    mapper = GenericParameterizedMapper(
        source_file=args.source,
        destination_file=args.destination,
        mapping_file=args.mapping,
        target_column=args.column,
        data_column=args.data_column,
        output_file=args.output
    )
    
    success = mapper.run_generic_mapping()
    
    if success:
        print(f"\n🎉 SUCCESS: Generic field mapping completed!")
    else:
        print(f"\n❌ FAILURE: Generic field mapping failed!")
        exit(1)


if __name__ == "__main__":
    main()
