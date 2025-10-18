#!/usr/bin/env python3
"""
Precision-Adjusted Q1 Value Verification

Adjusts precision for percentage fields to enable matching.
Rounds values to 6 decimal places for comparison to handle floating-point precision differences.

Author: AI Assistant
Date: October 2025
"""

import openpyxl
import csv
import pandas as pd
from pathlib import Path
from typing import Dict, List, Tuple, Optional


def round_for_comparison(value, precision=6):
    """Round value for comparison, handling None values."""
    if value is None:
        return None
    try:
        return round(float(value), precision)
    except:
        return value


def perform_precision_adjusted_q1_verification() -> List[Dict]:
    """
    Perform Q1 verification with precision adjustment for percentage fields.
    """
    print("="*80)
    print("PRECISION-ADJUSTED Q1 VALUE VERIFICATION")
    print("="*80)
    print("Using 6 decimal place precision for percentage field matching")
    
    # Load Key Metrics enhanced scoping
    source_scoping = {}
    km_file = "/Users/michaelkim/code/Bernstein/final_improved_key_metrics_mapping.csv"
    if Path(km_file).exists():
        km_df = pd.read_csv(km_file)
        for _, row in km_df.iterrows():
            row_num = row['Row_Number']
            source_scoping[row_num] = {
                'sheet_name': 'Key Metrics',
                'row_number': row_num,
                'original_field_name': row['Original_Field_Name'],
                'enhanced_scoped_name': row['Enhanced_Scoped_Name'],
                'section_context': row['Section_Context'] if pd.notna(row['Section_Context']) else '',
                'q1_2024_value': row['Q1_2024_Value'] if pd.notna(row['Q1_2024_Value']) else None,
                'q2_2024_value': row['Q2_2024_Value'] if pd.notna(row['Q2_2024_Value']) else None,
                'q1_rounded': round_for_comparison(row['Q1_2024_Value'])
            }
        print(f"Loaded Key Metrics scoping: {len(source_scoping)} fields")
    
    # Load destination scoping
    dest_file = "/Users/michaelkim/code/Bernstein/reported_tab_comprehensive_mapping.csv"
    dest_scoping = {}
    
    if Path(dest_file).exists():
        dest_df = pd.read_csv(dest_file)
        for _, row in dest_df.iterrows():
            row_num = row['Row_Number']
            dest_scoping[row_num] = {
                'original_field_name': row['Original_Field_Name'],
                'enhanced_scoped_name': row['Enhanced_Scoped_Name'],
                'major_section_context': row['Major_Section_Context'] if pd.notna(row['Major_Section_Context']) else '',
                'section_context': row['Section_Context'] if pd.notna(row['Section_Context']) else ''
            }
        print(f"Loaded destination scoping: {len(dest_scoping)} fields")
    
    # Load destination Q1 values
    dest_file_path = "/Users/michaelkim/code/Bernstein/20240725_IPGP.US-IPG Photonics.xlsx"
    dest_wb = openpyxl.load_workbook(dest_file_path, data_only=True)
    dest_sheet = dest_wb['Reported']
    
    dest_q1_data = {}
    for row_idx in range(1, dest_sheet.max_row + 1):
        q1_value = dest_sheet.cell(row_idx, 70).value  # Column BR
        if q1_value is not None:
            dest_q1_data[row_idx] = {
                'original': q1_value,
                'rounded': round_for_comparison(q1_value)
            }
    
    dest_wb.close()
    print(f"Loaded destination Q1 data: {len(dest_q1_data)} rows")
    
    # Perform precision-adjusted Q1 verification
    print(f"\nPerforming precision-adjusted Q1 verification...")
    
    matches = []
    used_source_rows = set()  # Track to prefer unique mappings when possible
    
    for dest_row, dest_q1_info in dest_q1_data.items():
        if dest_row not in dest_scoping:
            continue
            
        dest_field_info = dest_scoping[dest_row]
        dest_q1_original = dest_q1_info['original']
        dest_q1_rounded = dest_q1_info['rounded']
        
        print(f"\nDEST Row {dest_row}: {dest_field_info['original_field_name']}")
        print(f"  Q1 value: {dest_q1_original} (rounded: {dest_q1_rounded})")
        
        # Find source field with matching Q1 value (using rounded comparison)
        best_match = None
        best_score = 0
        
        for source_row, source_field in source_scoping.items():
            if source_field['q1_rounded'] is None:
                continue
            
            # Use rounded comparison for precision adjustment
            if source_field['q1_rounded'] == dest_q1_rounded:
                # Calculate preference score
                score = 1.0
                
                # Prefer unused source rows when possible
                if source_row not in used_source_rows:
                    score += 0.5
                
                # Prefer better context matches
                if dest_field_info['original_field_name'].lower() == source_field['original_field_name'].lower():
                    score += 0.3
                
                if score > best_score:
                    best_score = score
                    best_match = (source_row, source_field)
        
        if best_match:
            source_row, source_field = best_match
            used_source_rows.add(source_row)
            
            match = {
                'Dest_Row_Number': dest_row,
                'Dest_Field_Name': dest_field_info['original_field_name'],
                'Dest_Enhanced_Scope': dest_field_info['enhanced_scoped_name'],
                'Dest_Section_Context': dest_field_info['section_context'],
                'Dest_Major_Section_Context': dest_field_info['major_section_context'],
                
                'Source_Sheet_Name': source_field['sheet_name'],
                'Source_Row_Number': source_field['row_number'],
                'Source_Field_Name': source_field['original_field_name'],
                'Source_Enhanced_Scope': source_field['enhanced_scoped_name'],
                'Source_Section_Context': source_field['section_context'],
                
                'Q1_Verification_Value_Original': dest_q1_original,
                'Q1_Verification_Value_Rounded': dest_q1_rounded,
                'Source_Q1_Value_Rounded': source_field['q1_rounded'],
                'Source_Q2_Value': source_field['q2_2024_value'],
                'Match_Method': 'Q1_Value_Verification_Precision_Adjusted',
                'Match_Confidence': 1.0,
                'Preference_Score': best_score
            }
            
            matches.append(match)
            
            print(f"  ✓ MATCHED to Row {source_field['row_number']}: {source_field['original_field_name']}")
            print(f"    Q1 verification: {dest_q1_rounded} = {source_field['q1_rounded']} ✓ (precision adjusted)")
            print(f"    Q2 available: {source_field['q2_2024_value']}")
        else:
            print(f"  ❌ No Q1 match found for rounded value: {dest_q1_rounded}")
    
    print(f"\nPrecision-adjusted Q1 verification complete: {len(matches)} matches found")
    return matches


def main():
    """Main entry point for precision-adjusted Q1 verification."""
    output_file = "/Users/michaelkim/code/Bernstein/precision_adjusted_q1_mapping.csv"
    
    print("PRECISION-ADJUSTED Q1 VALUE VERIFICATION")
    print("Rounds values to 6 decimal places to handle floating-point precision")
    print(f"Output file: {output_file}")
    
    try:
        matches = perform_precision_adjusted_q1_verification()
        
        # Save results
        with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = [
                'Dest_Row_Number', 'Dest_Field_Name', 'Dest_Enhanced_Scope',
                'Dest_Section_Context', 'Dest_Major_Section_Context',
                'Source_Sheet_Name', 'Source_Row_Number', 'Source_Field_Name',
                'Source_Enhanced_Scope', 'Source_Section_Context',
                'Q1_Verification_Value_Original', 'Q1_Verification_Value_Rounded',
                'Source_Q1_Value_Rounded', 'Source_Q2_Value',
                'Match_Method', 'Match_Confidence', 'Preference_Score'
            ]
            
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(matches)
        
        print(f"\nPrecision-adjusted mapping saved to: {output_file}")
        print(f"Total matches: {len(matches)}")
        
        # Check specifically for rows 48-54
        rows_48_54 = [m for m in matches if 48 <= m['Dest_Row_Number'] <= 54]
        print(f"\nRows 48-54 found: {len(rows_48_54)}")
        for match in rows_48_54:
            print(f"  Row {match['Dest_Row_Number']}: {match['Dest_Field_Name']} → Row {match['Source_Row_Number']}: {match['Source_Field_Name']}")
        
        # Show improvement
        if len(rows_48_54) > 0:
            print(f"\n✅ SUCCESS: Precision adjustment captured {len(rows_48_54)} previously missing fields!")
        else:
            print(f"\n❌ Still no matches for rows 48-54 - may need different approach")
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
