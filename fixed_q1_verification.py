#!/usr/bin/env python3
"""
Fixed Q1 Value Verification - Allows Multiple Context Mappings

Fixes the issue where percentage context fields (rows 48-54) were not mapped
because source percentage fields were already used for absolute context mappings.

Allows same source field to map to multiple destinations when contexts are different.

Author: AI Assistant
Date: October 2025
"""

import openpyxl
import csv
import pandas as pd
from pathlib import Path
from typing import Dict, List, Tuple, Optional


def perform_fixed_q1_verification() -> List[Dict]:
    """
    Perform Q1 verification that allows multiple context mappings.
    """
    print("="*80)
    print("FIXED Q1 VALUE VERIFICATION - MULTIPLE CONTEXT MAPPINGS")
    print("="*80)
    
    # Load source data from all sheets
    source_file = "/Users/michaelkim/code/Bernstein/IPGP-Financial-Data-Workbook-2024-Q2.xlsx"
    
    # Load Key Metrics enhanced scoping
    source_scoping = {}
    km_file = "/Users/michaelkim/code/Bernstein/final_improved_key_metrics_mapping.csv"
    if Path(km_file).exists():
        km_df = pd.read_csv(km_file)
        for _, row in km_df.iterrows():
            row_num = row['Row_Number']
            source_scoping[f"KM_{row_num}"] = {
                'sheet_name': 'Key Metrics',
                'row_number': row_num,
                'original_field_name': row['Original_Field_Name'],
                'enhanced_scoped_name': row['Enhanced_Scoped_Name'],
                'section_context': row['Section_Context'] if pd.notna(row['Section_Context']) else '',
                'q1_2024_value': row['Q1_2024_Value'] if pd.notna(row['Q1_2024_Value']) else None,
                'q2_2024_value': row['Q2_2024_Value'] if pd.notna(row['Q2_2024_Value']) else None
            }
        print(f"Loaded Key Metrics scoping: {len(source_scoping)} fields")
    
    # Load destination scoping
    dest_file = "/Users/michaelkim/code/Bernstein/reported_tab_comprehensive_mapping.csv"
    dest_scoping = {}
    
    if Path(dest_file).exists():
        dest_df = pd.read_csv(dest_file)
        for _, row in dest_df.iterrows():
            row_num = row['Row_Number']
            dest_scoping[row_num] = {
                'original_field_name': row['Original_Field_Name'],
                'enhanced_scoped_name': row['Enhanced_Scoped_Name'],
                'major_section_context': row['Major_Section_Context'] if pd.notna(row['Major_Section_Context']) else '',
                'section_context': row['Section_Context'] if pd.notna(row['Section_Context']) else ''
            }
        print(f"Loaded destination scoping: {len(dest_scoping)} fields")
    
    # Load destination Q1 values
    dest_file_path = "/Users/michaelkim/code/Bernstein/20240725_IPGP.US-IPG Photonics.xlsx"
    dest_wb = openpyxl.load_workbook(dest_file_path, data_only=True)
    dest_sheet = dest_wb['Reported']
    
    dest_q1_data = {}
    for row_idx in range(1, dest_sheet.max_row + 1):
        q1_value = dest_sheet.cell(row_idx, 70).value  # Column BR
        if q1_value is not None:
            dest_q1_data[row_idx] = q1_value
    
    dest_wb.close()
    
    # Perform Q1 verification WITHOUT restricting source field reuse
    print(f"\nPerforming Q1 verification for ALL destination fields...")
    print(f"Allowing source fields to be used multiple times for different contexts")
    
    matches = []
    
    for dest_row, dest_q1_value in dest_q1_data.items():
        if dest_row not in dest_scoping:
            continue
            
        dest_field_info = dest_scoping[dest_row]
        
        print(f"\nDEST Row {dest_row}: {dest_field_info['original_field_name']}")
        print(f"  Q1 value: {dest_q1_value}")
        
        # Find ALL source fields with matching Q1 value
        matching_sources = []
        
        for source_key, source_field in source_scoping.items():
            if (source_field['q1_2024_value'] is not None and 
                source_field['q1_2024_value'] == dest_q1_value):
                matching_sources.append((source_key, source_field))
        
        if matching_sources:
            # If multiple matches, pick the best one based on context
            best_match = None
            best_score = 0
            
            for source_key, source_field in matching_sources:
                # Score based on field name similarity and context
                score = calculate_context_match_score(dest_field_info, source_field)
                
                if score > best_score:
                    best_score = score
                    best_match = (source_key, source_field)
            
            if best_match:
                source_key, source_field = best_match
                
                match = {
                    'Dest_Row_Number': dest_row,
                    'Dest_Field_Name': dest_field_info['original_field_name'],
                    'Dest_Enhanced_Scope': dest_field_info['enhanced_scoped_name'],
                    'Dest_Section_Context': dest_field_info['section_context'],
                    'Dest_Major_Section_Context': dest_field_info['major_section_context'],
                    
                    'Source_Sheet_Name': source_field['sheet_name'],
                    'Source_Row_Number': source_field['row_number'],
                    'Source_Field_Name': source_field['original_field_name'],
                    'Source_Enhanced_Scope': source_field['enhanced_scoped_name'],
                    'Source_Section_Context': source_field['section_context'],
                    
                    'Q1_Verification_Value': dest_q1_value,
                    'Source_Q2_Value': source_field['q2_2024_value'],
                    'Match_Method': 'Q1_Value_Verification_Fixed',
                    'Match_Confidence': 1.0,
                    'Context_Score': best_score
                }
                
                matches.append(match)
                
                print(f"  ✓ MATCHED to {source_field['sheet_name']} Row {source_field['row_number']}: {source_field['original_field_name']}")
                print(f"    Context score: {best_score:.2f}")
                print(f"    Q1 verification: {dest_q1_value} = {source_field['q1_2024_value']} ✓")
        else:
            print(f"  ❌ No Q1 match found for value: {dest_q1_value}")
    
    print(f"\nFixed Q1 verification complete: {len(matches)} matches found")
    return matches


def calculate_context_match_score(dest_field: Dict, source_field: Dict) -> float:
    """Calculate context match score to pick best source when multiple Q1 matches exist."""
    score = 0.0
    
    dest_name = dest_field['original_field_name'].lower()
    source_name = source_field['original_field_name'].lower()
    
    # Field name similarity (most important)
    if dest_name == source_name:
        score += 0.5
    elif dest_name in source_name or source_name in dest_name:
        score += 0.3
    
    # Section context matching
    dest_section = dest_field.get('section_context', '').lower()
    source_section = source_field.get('section_context', '').lower()
    
    # Prefer percentage contexts for percentage destination sections
    if ('segment_information' in dest_section and 
        '% of total' in source_section):
        score += 0.3
    
    # Prefer absolute contexts for segment breakdown sections  
    if ('segment_breakdown' in dest_section and 
        '% of total' not in source_section):
        score += 0.3
    
    return score


def main():
    """Main entry point for fixed Q1 verification."""
    output_file = "/Users/michaelkim/code/Bernstein/fixed_q1_verified_mapping.csv"
    
    print("Fixed Q1 verification that allows multiple context mappings")
    print(f"Output file: {output_file}")
    
    try:
        matches = perform_fixed_q1_verification()
        
        # Save results
        with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = [
                'Dest_Row_Number', 'Dest_Field_Name', 'Dest_Enhanced_Scope',
                'Dest_Section_Context', 'Dest_Major_Section_Context',
                'Source_Sheet_Name', 'Source_Row_Number', 'Source_Field_Name',
                'Source_Enhanced_Scope', 'Source_Section_Context',
                'Q1_Verification_Value', 'Source_Q2_Value',
                'Match_Method', 'Match_Confidence', 'Context_Score'
            ]
            
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(matches)
        
        print(f"\nFixed mapping saved to: {output_file}")
        print(f"Total matches: {len(matches)}")
        
        # Check specifically for rows 48-54
        rows_48_54 = [m for m in matches if 48 <= m['Dest_Row_Number'] <= 54]
        print(f"\nRows 48-54 found: {len(rows_48_54)}")
        for match in rows_48_54:
            print(f"  Row {match['Dest_Row_Number']}: {match['Dest_Field_Name']} → {match['Source_Field_Name']}")
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
