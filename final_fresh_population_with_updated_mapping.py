#!/usr/bin/env python3
"""
Final Fresh Population with Updated Mapping

Performs a complete fresh population using the updated consolidated mapping file
that now includes all 134 verified mappings including Row 23.

Author: AI Assistant
Date: October 2025
"""

import openpyxl
import pandas as pd
import csv
from pathlib import Path
from typing import Dict, List, Optional


def load_updated_consolidated_mappings() -> List[Dict]:
    """Load all mappings from the updated consolidated mapping file."""
    
    print("=== LOADING UPDATED CONSOLIDATED FIELD MAPPINGS ===")
    
    consolidated_file = "/Users/michaelkim/code/Bernstein/CONSOLIDATED_FIELD_MAPPINGS.csv"
    
    if not Path(consolidated_file).exists():
        print(f"ERROR: Consolidated mapping file not found: {consolidated_file}")
        return []
    
    mappings = []
    with open(consolidated_file, 'r', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        mappings = list(reader)
    
    print(f"Loaded {len(mappings)} consolidated mappings (updated)")
    
    # Show breakdown by source sheet
    sheet_breakdown = {}
    method_breakdown = {}
    
    for mapping in mappings:
        sheet = mapping['Source_Sheet_Name']
        method = mapping['Match_Method']
        
        sheet_breakdown[sheet] = sheet_breakdown.get(sheet, 0) + 1
        method_breakdown[method] = method_breakdown.get(method, 0) + 1
    
    print(f"\nMappings by source sheet:")
    for sheet, count in sorted(sheet_breakdown.items()):
        print(f"  {sheet}: {count} mappings")
    
    print(f"\nMappings by method:")
    for method, count in sorted(method_breakdown.items()):
        print(f"  {method}: {count} mappings")
    
    # Verify Row 23 is included
    row_23_mappings = [m for m in mappings if m['Dest_Row_Number'] == '23']
    if row_23_mappings:
        row_23 = row_23_mappings[0]
        print(f"\n✅ Row 23 verified in updated mapping:")
        print(f"   {row_23['Dest_Field_Name']} → {row_23['Source_Field_Name']}")
        print(f"   Q2 Value: {row_23['Source_Q2_Value']}")
    else:
        print(f"\n❌ Row 23 not found in updated mapping")
    
    return mappings


def final_fresh_population_from_updated_mapping(mappings: List[Dict]) -> Dict:
    """Perform final fresh population using updated consolidated mappings."""
    
    print(f"\n" + "="*80)
    print("FINAL FRESH POPULATION FROM UPDATED MAPPING")
    print("="*80)
    
    # Use original destination file (clean slate)
    original_dest_file = "/Users/michaelkim/code/Bernstein/20240725_IPGP.US-IPG Photonics.xlsx"
    source_file = "/Users/michaelkim/code/Bernstein/IPGP-Financial-Data-Workbook-2024-Q2.xlsx"
    
    print(f"Original destination file: {original_dest_file}")
    print(f"Source file: {source_file}")
    print(f"Using updated consolidated mapping with {len(mappings)} mappings")
    
    # Load workbooks
    source_wb = openpyxl.load_workbook(source_file, data_only=True)
    dest_wb = openpyxl.load_workbook(original_dest_file, data_only=False)
    dest_sheet = dest_wb['Reported']
    
    population_results = []
    values_populated = 0
    source_tracking_added = 0
    errors = []
    sheet_stats = {}
    method_stats = {}
    
    print(f"\nPopulating {len(mappings)} updated consolidated mappings...")
    
    for i, mapping in enumerate(mappings, 1):
        dest_row = int(mapping['Dest_Row_Number'])
        source_sheet_name = mapping['Source_Sheet_Name']
        source_row = mapping['Source_Row_Number']
        dest_field_name = mapping['Dest_Field_Name']
        source_field_name = mapping['Source_Field_Name']
        match_method = mapping['Match_Method']
        
        # Highlight Row 23 processing
        if dest_row == 23:
            print(f"\n[{i}/{len(mappings)}] ⭐ DEST Row {dest_row}: {dest_field_name} (NEWLY ADDED)")
        else:
            print(f"\n[{i}/{len(mappings)}] DEST Row {dest_row}: {dest_field_name}")
        
        print(f"  Method: {match_method}")
        
        try:
            # Get source sheet and Q2 value
            if source_sheet_name in source_wb.sheetnames:
                source_sheet = source_wb[source_sheet_name]
                
                # Handle source row
                if not source_row or source_row.strip() == '':
                    print(f"  ❌ No source row specified")
                    errors.append(f"Row {dest_row}: No source row specified")
                    continue
                
                # Handle composite source rows (like "30+31+32+33")
                if '+' in str(source_row):
                    # Composite field - sum multiple rows
                    composite_rows = [int(r.strip()) for r in str(source_row).split('+')]
                    composite_q2_value = 0
                    
                    print(f"  Composite field from rows: {composite_rows}")
                    for comp_row in composite_rows:
                        comp_value = source_sheet.cell(comp_row, 93).value or 0
                        composite_q2_value += comp_value
                        print(f"    Row {comp_row}: {comp_value}")
                    
                    source_q2_value = composite_q2_value
                    source_location = f"IPGP-Financial-Data-Workbook-2024-Q2.xlsx|{source_sheet_name}|{source_row}|93"
                    print(f"  Composite Q2 total: {source_q2_value}")
                else:
                    # Single source row
                    source_row_num = int(source_row)
                    source_q2_value = source_sheet.cell(source_row_num, 93).value
                    source_location = f"IPGP-Financial-Data-Workbook-2024-Q2.xlsx|{source_sheet_name}|{source_row_num}|93"
                
                current_dest_value = dest_sheet.cell(dest_row, 71).value
                
                print(f"  From {source_sheet_name} Row {source_row}: {source_field_name}")
                print(f"  Q2 value: {source_q2_value}")
                
                if source_q2_value is not None:
                    # Populate Column BS (71) with Q2 value
                    dest_sheet.cell(dest_row, 71).value = source_q2_value
                    values_populated += 1
                    
                    # Add source tracking to Column BT (72)
                    dest_sheet.cell(dest_row, 72).value = source_location
                    source_tracking_added += 1
                    
                    # Track stats
                    if source_sheet_name not in sheet_stats:
                        sheet_stats[source_sheet_name] = 0
                    sheet_stats[source_sheet_name] += 1
                    
                    if match_method not in method_stats:
                        method_stats[match_method] = 0
                    method_stats[match_method] += 1
                    
                    if dest_row == 23:
                        print(f"  ⭐ NEWLY POPULATED BS: {source_q2_value}")
                        print(f"  ⭐ NEWLY TRACKED BT: {source_location}")
                    else:
                        print(f"  ✅ POPULATED BS: {source_q2_value}")
                        print(f"  ✅ TRACKED BT: {source_location}")
                    
                    population_results.append({
                        'Dest_Row': dest_row,
                        'Dest_Field_Name': dest_field_name,
                        'Source_Sheet': source_sheet_name,
                        'Source_Row': source_row,
                        'Source_Field_Name': source_field_name,
                        'Source_Q2_Value': source_q2_value,
                        'Source_Location': source_location,
                        'Match_Method': match_method,
                        'Previous_Value': current_dest_value,
                        'Status': 'POPULATED'
                    })
                else:
                    print(f"  ❌ No Q2 data available")
                    errors.append(f"Row {dest_row}: No Q2 data in source")
            else:
                print(f"  ❌ Source sheet not found: {source_sheet_name}")
                errors.append(f"Row {dest_row}: Source sheet '{source_sheet_name}' not found")
                
        except Exception as e:
            print(f"  ❌ Error processing row: {e}")
            errors.append(f"Row {dest_row}: {str(e)}")
    
    # Save final fresh populated file
    output_file = "/Users/michaelkim/code/Bernstein/FINAL_FRESH_POPULATED_WITH_UPDATED_MAPPING_IPGP.xlsx"
    print(f"\nSaving final fresh populated file to: {output_file}")
    dest_wb.save(output_file)
    
    source_wb.close()
    dest_wb.close()
    
    return {
        'population_results': population_results,
        'values_populated': values_populated,
        'source_tracking_added': source_tracking_added,
        'total_mappings': len(mappings),
        'sheet_stats': sheet_stats,
        'method_stats': method_stats,
        'errors': errors,
        'output_file': output_file
    }


def main():
    """Main entry point for final fresh population with updated mapping."""
    
    print("="*80)
    print("FINAL FRESH POPULATION WITH UPDATED MAPPING")
    print("="*80)
    print("Using updated consolidated mapping file (134 mappings including Row 23)")
    print("Starting fresh from original destination file")
    
    try:
        # Load updated consolidated mappings
        updated_mappings = load_updated_consolidated_mappings()
        
        if not updated_mappings:
            print("ERROR: No updated mappings loaded")
            return
        
        # Perform final fresh population
        population_summary = final_fresh_population_from_updated_mapping(updated_mappings)
        
        # Generate audit trail
        audit_file = "/Users/michaelkim/code/Bernstein/FINAL_FRESH_POPULATION_AUDIT_TRAIL.csv"
        with open(audit_file, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = [
                'Dest_Row', 'Dest_Field_Name', 'Source_Sheet', 'Source_Row',
                'Source_Field_Name', 'Source_Q2_Value', 'Source_Location',
                'Match_Method', 'Previous_Value', 'Status'
            ]
            
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(population_summary['population_results'])
        
        # Final summary
        print(f"\n" + "="*80)
        print("FINAL FRESH POPULATION RESULTS")
        print("="*80)
        print(f"Total mappings processed: {population_summary['total_mappings']}")
        print(f"Values successfully populated (Column BS): {population_summary['values_populated']}")
        print(f"Source locations added (Column BT): {population_summary['source_tracking_added']}")
        print(f"Success rate: {population_summary['values_populated']}/{population_summary['total_mappings']} = {population_summary['values_populated']/population_summary['total_mappings']*100:.1f}%")
        
        print(f"\nPopulation by source sheet:")
        for sheet, count in sorted(population_summary['sheet_stats'].items()):
            print(f"  {sheet}: {count} values")
        
        print(f"\nPopulation by method:")
        for method, count in sorted(population_summary['method_stats'].items()):
            print(f"  {method}: {count} values")
        
        if population_summary['errors']:
            print(f"\nErrors encountered: {len(population_summary['errors'])}")
            for error in population_summary['errors'][:3]:  # Show first 3
                print(f"  {error}")
        
        print(f"\nFinal files:")
        print(f"  📁 Final populated Excel: {population_summary['output_file']}")
        print(f"  📁 Final audit trail: {audit_file}")
        print(f"  📁 Updated consolidated mappings: CONSOLIDATED_FIELD_MAPPINGS.csv (134 mappings)")
        
        # Highlight Row 23 success
        row_23_results = [r for r in population_summary['population_results'] if r['Dest_Row'] == 23]
        if row_23_results:
            row_23 = row_23_results[0]
            print(f"\n⭐ ROW 23 SUCCESS:")
            print(f"   {row_23['Dest_Field_Name']} = {row_23['Source_Q2_Value']:,}")
            print(f"   Source: {row_23['Source_Location']}")
        
        print(f"\n🎉 FINAL FRESH POPULATION COMPLETE!")
        print(f"All {population_summary['values_populated']} verified mappings applied from scratch")
        print(f"with updated consolidated mapping file!")
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
