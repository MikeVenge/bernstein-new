#!/usr/bin/env python3
"""
Field Matching Analysis Between Key Metrics and Reported Tabs

This script matches fields between the source (Key Metrics) and target (Reported) tabs
based on field names and enhanced scoping, then shows the actual values side by side
from the latest available quarter in both files.

Author: AI Assistant
Date: October 2025
"""

import pandas as pd
import openpyxl
import csv
from pathlib import Path
import re
from typing import Dict, List, Tuple, Optional
from difflib import SequenceMatcher


def load_mapping_files() -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Load the two comprehensive mapping CSV files."""
    print("Loading mapping files...")
    
    key_metrics_file = "/Users/michaelkim/code/Bernstein/key_metrics_comprehensive_mapping.csv"
    reported_file = "/Users/michaelkim/code/Bernstein/reported_tab_comprehensive_mapping.csv"
    
    if not Path(key_metrics_file).exists():
        raise FileNotFoundError(f"Key Metrics mapping file not found: {key_metrics_file}")
    if not Path(reported_file).exists():
        raise FileNotFoundError(f"Reported mapping file not found: {reported_file}")
    
    key_metrics_df = pd.read_csv(key_metrics_file)
    reported_df = pd.read_csv(reported_file)
    
    print(f"Key Metrics mapping: {len(key_metrics_df)} fields")
    print(f"Reported mapping: {len(reported_df)} fields")
    
    return key_metrics_df, reported_df


def get_latest_quarter_data(source_file: str, target_file: str) -> Tuple[Dict, Dict]:
    """Get the latest quarter data from both source and target files."""
    print("Loading actual data files to get latest quarter values...")
    
    # Load source file (Key Metrics)
    source_wb = openpyxl.load_workbook(source_file, data_only=True)
    key_metrics_sheet = source_wb['Key Metrics']
    
    # Find Q2 2024 column in Key Metrics (column 93 = 2024-06-30)
    source_latest_col = 93  # Q2 2024 column
    print(f"Using source column {source_latest_col} for Q2 2024")
    
    # Get source data
    source_data = {}
    for row_idx in range(1, key_metrics_sheet.max_row + 1):
        first_col = key_metrics_sheet.cell(row_idx, 1).value
        if first_col:
            value = key_metrics_sheet.cell(row_idx, source_latest_col).value
            source_data[row_idx] = {
                'field_name': str(first_col).strip(),
                'value': value
            }
    
    source_wb.close()
    
    # Load target file (Reported)
    target_wb = openpyxl.load_workbook(target_file, data_only=True)
    reported_sheet = target_wb['Reported']
    
    # Check multiple columns for the best available data in Reported tab
    # Column 70 = Q1 2024, Column 71 = Q2 2024 (but might be empty)
    target_cols_to_try = [71, 70, 69]  # Q2 2024, Q1 2024, Q4 2023
    target_latest_col = None
    
    # Test which column has the most data
    for test_col in target_cols_to_try:
        data_count = 0
        for test_row in range(10, 50):  # Sample some rows
            if reported_sheet.cell(test_row, test_col).value is not None:
                data_count += 1
        print(f"Column {test_col} has {data_count} non-empty values in sample")
        if data_count > 5:  # If we find a column with decent data
            target_latest_col = test_col
            break
    
    if not target_latest_col:
        target_latest_col = 70  # Fallback to Q1 2024
    
    print(f"Using target column {target_latest_col} for latest available quarter")
    
    # Get target data
    target_data = {}
    for row_idx in range(1, reported_sheet.max_row + 1):
        first_col = reported_sheet.cell(row_idx, 1).value
        if first_col:
            value = reported_sheet.cell(row_idx, target_latest_col).value
            target_data[row_idx] = {
                'field_name': str(first_col).strip(),
                'value': value
            }
    
    target_wb.close()
    
    print(f"Loaded data for {len(source_data)} source rows and {len(target_data)} target rows")
    return source_data, target_data


def calculate_similarity(str1: str, str2: str) -> float:
    """Calculate similarity between two strings."""
    if not str1 or not str2:
        return 0.0
    
    # Clean strings for comparison
    clean1 = re.sub(r'[^\w\s]', '', str1.lower()).strip()
    clean2 = re.sub(r'[^\w\s]', '', str2.lower()).strip()
    
    return SequenceMatcher(None, clean1, clean2).ratio()


def match_fields_by_name(key_metrics_df: pd.DataFrame, reported_df: pd.DataFrame) -> List[Dict]:
    """Match fields between the two datasets based on field names and enhanced scoping."""
    print("Matching fields based on names and enhanced scoping...")
    
    matches = []
    
    for _, km_row in key_metrics_df.iterrows():
        km_original = km_row['Original_Field_Name']
        km_cleaned = km_row['Cleaned_Field_Name']
        km_enhanced = km_row['Enhanced_Scoped_Name']
        
        best_match = None
        best_score = 0.0
        
        for _, rep_row in reported_df.iterrows():
            rep_original = rep_row['Original_Field_Name']
            rep_cleaned = rep_row['Cleaned_Field_Name']
            rep_enhanced = rep_row['Enhanced_Scoped_Name']
            
            # Calculate different similarity scores
            scores = [
                calculate_similarity(km_original, rep_original) * 1.0,  # Original names
                calculate_similarity(km_cleaned, rep_cleaned) * 0.9,    # Cleaned names
                calculate_similarity(km_enhanced, rep_enhanced) * 0.8,  # Enhanced scoping
            ]
            
            # Special bonus for exact geographic/product matches
            if any(geo in km_original.lower() and geo in rep_original.lower() 
                   for geo in ['china', 'japan', 'germany', 'north america', 'europe', 'asia']):
                scores.append(0.3)
            
            if any(prod in km_original.lower() and prod in rep_original.lower() 
                   for prod in ['materials processing', 'communications', 'medical', 'advanced']):
                scores.append(0.3)
            
            # Special bonus for common financial terms
            if any(term in km_original.lower() and term in rep_original.lower() 
                   for term in ['total', 'revenue', 'income', 'sales', 'assets']):
                scores.append(0.2)
            
            total_score = sum(scores)
            
            if total_score > best_score:
                best_score = total_score
                best_match = {
                    'reported_row': rep_row,
                    'similarity_score': total_score
                }
        
        # Only include matches above a threshold
        if best_match and best_score > 0.5:
            match_info = {
                'key_metrics_row': km_row,
                'reported_row': best_match['reported_row'],
                'similarity_score': best_match['similarity_score'],
                'match_quality': 'Excellent' if best_score > 0.8 else 'Good' if best_score > 0.65 else 'Fair'
            }
            matches.append(match_info)
    
    print(f"Found {len(matches)} field matches")
    
    # Sort by similarity score (best matches first)
    matches.sort(key=lambda x: x['similarity_score'], reverse=True)
    
    return matches


def enrich_matches_with_values(matches: List[Dict], source_data: Dict, target_data: Dict) -> List[Dict]:
    """Enrich the matches with actual values from the data files."""
    print("Enriching matches with actual data values...")
    
    enriched_matches = []
    
    for match in matches:
        km_row = match['key_metrics_row']
        rep_row = match['reported_row']
        
        # Get source value
        source_value = None
        km_row_num = km_row['Row_Number']
        if km_row_num in source_data:
            source_value = source_data[km_row_num]['value']
        
        # Get target value
        target_value = None
        rep_row_num = rep_row['Row_Number']
        if rep_row_num in target_data:
            target_value = target_data[rep_row_num]['value']
        
        # Create enriched match
        enriched_match = {
            # Key Metrics info
            'km_row_number': km_row['Row_Number'],
            'km_original_name': km_row['Original_Field_Name'],
            'km_cleaned_name': km_row['Cleaned_Field_Name'],
            'km_enhanced_scope': km_row['Enhanced_Scoped_Name'],
            'km_section': km_row.get('Section_Context', ''),
            'km_value': source_value,
            
            # Reported info
            'rep_row_number': rep_row['Row_Number'],
            'rep_original_name': rep_row['Original_Field_Name'],
            'rep_cleaned_name': rep_row['Cleaned_Field_Name'],
            'rep_enhanced_scope': rep_row['Enhanced_Scoped_Name'],
            'rep_major_section': rep_row.get('Major_Section_Context', ''),
            'rep_section': rep_row.get('Section_Context', ''),
            'rep_value': target_value,
            
            # Match info
            'similarity_score': match['similarity_score'],
            'match_quality': match['match_quality'],
            'values_match': check_values_match(source_value, target_value),
            'value_difference': calculate_value_difference(source_value, target_value)
        }
        
        enriched_matches.append(enriched_match)
    
    return enriched_matches


def check_values_match(val1, val2) -> str:
    """Check if two values match (accounting for different formats)."""
    if val1 is None and val2 is None:
        return 'Both Empty'
    if val1 is None or val2 is None:
        return 'One Empty'
    
    try:
        # Try to convert to numbers for comparison
        num1 = float(str(val1).replace(',', '')) if val1 else 0
        num2 = float(str(val2).replace(',', '')) if val2 else 0
        
        if abs(num1 - num2) < 0.01:  # Account for rounding
            return 'Match'
        elif abs(num1 - num2) / max(abs(num1), abs(num2), 1) < 0.05:  # Within 5%
            return 'Close Match'
        else:
            return 'Different'
    except:
        # String comparison
        str1 = str(val1).strip().lower()
        str2 = str(val2).strip().lower()
        if str1 == str2:
            return 'Match'
        else:
            return 'Different'


def calculate_value_difference(val1, val2) -> str:
    """Calculate the difference between two values."""
    if val1 is None or val2 is None:
        return 'N/A'
    
    try:
        num1 = float(str(val1).replace(',', '')) if val1 else 0
        num2 = float(str(val2).replace(',', '')) if val2 else 0
        diff = num1 - num2
        return f"{diff:,.0f}" if abs(diff) > 0.01 else "0"
    except:
        return 'N/A'


def save_field_matching_results(enriched_matches: List[Dict], output_file: str):
    """Save the field matching results to CSV."""
    print(f"Saving field matching results to {output_file}...")
    
    fieldnames = [
        # Key Metrics columns
        'KM_Row_Number',
        'KM_Original_Name',
        'KM_Cleaned_Name',
        'KM_Enhanced_Scope',
        'KM_Section',
        'KM_Latest_Value',
        
        # Reported columns
        'REP_Row_Number',
        'REP_Original_Name', 
        'REP_Cleaned_Name',
        'REP_Enhanced_Scope',
        'REP_Major_Section',
        'REP_Section',
        'REP_Latest_Value',
        
        # Match analysis
        'Similarity_Score',
        'Match_Quality',
        'Values_Match',
        'Value_Difference'
    ]
    
    with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        
        for match in enriched_matches:
            writer.writerow({
                'KM_Row_Number': match['km_row_number'],
                'KM_Original_Name': match['km_original_name'],
                'KM_Cleaned_Name': match['km_cleaned_name'],
                'KM_Enhanced_Scope': match['km_enhanced_scope'],
                'KM_Section': match['km_section'],
                'KM_Latest_Value': match['km_value'] if match['km_value'] is not None else '',
                
                'REP_Row_Number': match['rep_row_number'],
                'REP_Original_Name': match['rep_original_name'],
                'REP_Cleaned_Name': match['rep_cleaned_name'],
                'REP_Enhanced_Scope': match['rep_enhanced_scope'],
                'REP_Major_Section': match['rep_major_section'],
                'REP_Section': match['rep_section'],
                'REP_Latest_Value': match['rep_value'] if match['rep_value'] is not None else '',
                
                'Similarity_Score': f"{match['similarity_score']:.3f}",
                'Match_Quality': match['match_quality'],
                'Values_Match': match['values_match'],
                'Value_Difference': match['value_difference']
            })


def main():
    """Main entry point."""
    print("="*80)
    print("FIELD MATCHING ANALYSIS - KEY METRICS vs REPORTED")
    print("="*80)
    
    source_file = "/Users/michaelkim/code/Bernstein/IPGP-Financial-Data-Workbook-2024-Q2.xlsx"
    target_file = "/Users/michaelkim/code/Bernstein/20240725_IPGP.US-IPG Photonics.xlsx"
    output_file = "/Users/michaelkim/code/Bernstein/field_matching_analysis_results.csv"
    
    try:
        # Step 1: Load mapping files
        print("\n" + "="*50)
        print("STEP 1: LOADING MAPPING FILES")
        print("="*50)
        key_metrics_df, reported_df = load_mapping_files()
        
        # Step 2: Get latest quarter data from actual files
        print("\n" + "="*50)
        print("STEP 2: LOADING LATEST QUARTER DATA")
        print("="*50)
        source_data, target_data = get_latest_quarter_data(source_file, target_file)
        
        # Step 3: Match fields based on names and scoping
        print("\n" + "="*50)
        print("STEP 3: MATCHING FIELDS")
        print("="*50)
        matches = match_fields_by_name(key_metrics_df, reported_df)
        
        # Step 4: Enrich with actual values
        print("\n" + "="*50)
        print("STEP 4: ENRICHING WITH VALUES")
        print("="*50)
        enriched_matches = enrich_matches_with_values(matches, source_data, target_data)
        
        # Step 5: Save results
        print("\n" + "="*50)
        print("STEP 5: SAVING RESULTS")
        print("="*50)
        save_field_matching_results(enriched_matches, output_file)
        
        # Summary statistics
        print("\n" + "="*50)
        print("RESULTS SUMMARY")
        print("="*50)
        print(f"Total field matches found: {len(enriched_matches)}")
        
        # Quality breakdown
        quality_counts = {}
        value_match_counts = {}
        for match in enriched_matches:
            quality = match['match_quality']
            value_match = match['values_match']
            quality_counts[quality] = quality_counts.get(quality, 0) + 1
            value_match_counts[value_match] = value_match_counts.get(value_match, 0) + 1
        
        print("\nMatch quality breakdown:")
        for quality, count in sorted(quality_counts.items()):
            print(f"  {quality}: {count} matches")
        
        print("\nValue comparison breakdown:")
        for status, count in sorted(value_match_counts.items()):
            print(f"  {status}: {count} matches")
        
        # Show top matches
        print("\nTop 10 matches:")
        for i, match in enumerate(enriched_matches[:10]):
            print(f"  {i+1}. {match['km_original_name']} → {match['rep_original_name']}")
            print(f"     Score: {match['similarity_score']:.3f}, Values: {match['values_match']}")
        
        print(f"\n" + "="*80)
        print("FIELD MATCHING ANALYSIS COMPLETE!")
        print("="*80)
        print(f"Results saved to: {output_file}")
        print("Review the CSV file to see side-by-side field matches with actual values.")
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
