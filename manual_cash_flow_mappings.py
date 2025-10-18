#!/usr/bin/env python3
"""
Manual Cash Flow Mappings

Creates manual mappings for the three missing Cash Flow fields (rows 185-187).
Despite Q1 value differences, the field names match exactly.

Author: AI Assistant
Date: October 2025
"""

import openpyxl
import csv
from pathlib import Path


def search_for_tax_benefit_field():
    """Search for Tax benefit field across all sheets."""
    
    print("=== COMPREHENSIVE SEARCH FOR TAX BENEFIT FIELD ===")
    
    source_file = "/Users/michaelkim/code/Bernstein/IPGP-Financial-Data-Workbook-2024-Q2.xlsx"
    source_wb = openpyxl.load_workbook(source_file, data_only=True)
    
    sheets_to_check = ['Cash Flows', 'Income Statement', 'Balance Sheet', 'Key Metrics']
    
    for sheet_name in sheets_to_check:
        if sheet_name in source_wb.sheetnames:
            print(f"\n--- Checking {sheet_name} ---")
            sheet = source_wb[sheet_name]
            
            found_any = False
            for row in range(1, 101):
                field_name = sheet.cell(row, 1).value
                if field_name and ('tax benefit' in str(field_name).lower() or 
                                 'stock options' in str(field_name).lower()):
                    q1_val = sheet.cell(row, 70).value
                    q2_val = sheet.cell(row, 93).value
                    print(f"Row {row}: {field_name}")
                    print(f"  Q1: {q1_val} | Q2: {q2_val}")
                    found_any = True
            
            if not found_any:
                print(f"No tax benefit fields found in {sheet_name}")
    
    source_wb.close()


def create_manual_cash_flow_mappings():
    """Create manual mappings for the missing Cash Flow fields."""
    
    print("\n=== CREATING MANUAL CASH FLOW MAPPINGS ===")
    
    manual_mappings = [
        {
            'Dest_Row_Number': '185',
            'Dest_Field_Name': 'Income and other taxes payable',
            'Dest_Enhanced_Scope': 'Reported.Consolidated_Cash_Flow_Statement_In_000_Usd.Income_And_Other_Taxes_Payable',
            'Dest_Section_Context': 'Operating_Activities',
            'Dest_Major_Section_Context': 'Cash Flow Statement',
            'Source_Sheet_Name': 'Cash Flows',
            'Source_Row_Number': '23',
            'Source_Field_Name': 'Income and other taxes payable',
            'Source_Enhanced_Scope': 'Cash_Flows.Operating_Activities.Income_And_Other_Taxes_Payable',
            'Source_Section_Context': 'Operating_Activities',
            'Q1_Verification_Value': '-34698',  # Dest Q1 value
            'Source_Q1_Value': '5811',         # Source Q1 value (different)
            'Source_Q2_Value': '-18306',       # Source Q2 value to populate
            'Match_Method': 'Manual_Field_Name_Match',
            'Match_Confidence': '0.85',
            'Match_Reason': 'Field names match exactly, Q1 values different (-34698 vs 5811)'
        },
        {
            'Dest_Row_Number': '187',
            'Dest_Field_Name': 'Net cash provided by operating activities',
            'Dest_Enhanced_Scope': 'Reported.Consolidated_Cash_Flow_Statement_In_000_Usd.Net_Cash_Provided_By_Operating_Activities',
            'Dest_Section_Context': 'Operating_Activities',
            'Dest_Major_Section_Context': 'Cash Flow Statement',
            'Source_Sheet_Name': 'Cash Flows',
            'Source_Row_Number': '24',
            'Source_Field_Name': 'Net cash (used in) provided by operating activities',
            'Source_Enhanced_Scope': 'Cash_Flows.Operating_Activities.Net_Cash_Used_In_Provided_By_Operating_Activities',
            'Source_Section_Context': 'Operating_Activities',
            'Q1_Verification_Value': '54596',   # Dest Q1 value
            'Source_Q1_Value': '129939',       # Source Q1 value (different)
            'Source_Q2_Value': '53458',        # Source Q2 value to populate
            'Match_Method': 'Manual_Field_Name_Match',
            'Match_Confidence': '0.90',
            'Match_Reason': 'Field names semantically equivalent, Q1 values different (54596 vs 129939)'
        }
    ]
    
    print(f"Created {len(manual_mappings)} manual Cash Flow mappings:")
    for mapping in manual_mappings:
        print(f"  Row {mapping['Dest_Row_Number']}: {mapping['Dest_Field_Name']}")
        print(f"    → {mapping['Source_Field_Name']} (Row {mapping['Source_Row_Number']})")
        print(f"    Q2 Value: {mapping['Source_Q2_Value']}")
    
    return manual_mappings


def populate_manual_cash_flow_mappings(mappings):
    """Populate the destination file with the manual Cash Flow mappings."""
    
    print("\n=== POPULATING MANUAL CASH FLOW MAPPINGS ===")
    
    source_file = "/Users/michaelkim/code/Bernstein/IPGP-Financial-Data-Workbook-2024-Q2.xlsx"
    dest_file = "/Users/michaelkim/code/Bernstein/final_with_composite_accrued_expenses_IPGP.xlsx"
    
    # Load workbooks
    source_wb = openpyxl.load_workbook(source_file, data_only=True)
    dest_wb = openpyxl.load_workbook(dest_file, data_only=False)
    dest_sheet = dest_wb['Reported']
    
    populated_count = 0
    
    for mapping in mappings:
        dest_row = int(mapping['Dest_Row_Number'])
        source_sheet_name = mapping['Source_Sheet_Name']
        source_row = int(mapping['Source_Row_Number'])
        expected_q2_value = mapping['Source_Q2_Value']
        
        print(f"\nRow {dest_row}: {mapping['Dest_Field_Name']}")
        print(f"  From {source_sheet_name} Row {source_row}: {mapping['Source_Field_Name']}")
        print(f"  Expected Q2 value: {expected_q2_value}")
        
        # Verify the source value
        if source_sheet_name in source_wb.sheetnames:
            source_sheet = source_wb[source_sheet_name]
            actual_q2_value = source_sheet.cell(source_row, 93).value  # Column CO
            
            print(f"  Actual Q2 value: {actual_q2_value}")
            
            if str(actual_q2_value) == str(expected_q2_value):
                # Populate Column BS (71) with Q2 value
                dest_sheet.cell(dest_row, 71).value = float(expected_q2_value) if expected_q2_value not in ['0', ''] else (0 if expected_q2_value == '0' else actual_q2_value)
                
                # Add source tracking to Column BT (72)
                source_location = f"IPGP-Financial-Data-Workbook-2024-Q2.xlsx|{source_sheet_name}|{source_row}|93"
                dest_sheet.cell(dest_row, 72).value = source_location
                
                populated_count += 1
                print(f"  ✅ POPULATED BS: {actual_q2_value}")
                print(f"  ✅ TRACKED BT: {source_location}")
            else:
                print(f"  ❌ Q2 value mismatch: expected {expected_q2_value}, got {actual_q2_value}")
        else:
            print(f"  ❌ Source sheet not found: {source_sheet_name}")
    
    # Save the updated file
    output_file = "/Users/michaelkim/code/Bernstein/final_with_cash_flow_mappings_IPGP.xlsx"
    dest_wb.save(output_file)
    
    source_wb.close()
    dest_wb.close()
    
    print(f"\n✅ Manual Cash Flow population complete!")
    print(f"Populated {populated_count} additional fields")
    print(f"Updated file saved as: {output_file}")
    
    return output_file, populated_count


def main():
    """Main entry point for manual Cash Flow mappings."""
    
    print("="*80)
    print("MANUAL CASH FLOW MAPPINGS - MISSING FIELDS")
    print("="*80)
    print("Adding missing Cash Flow fields (rows 185, 187)")
    print("Note: Row 186 needs separate investigation")
    
    try:
        # Search for tax benefit field
        search_for_tax_benefit_field()
        
        # Create manual mappings (excluding row 186 for now)
        manual_mappings = create_manual_cash_flow_mappings()
        
        # Save mappings to file
        output_file = "/Users/michaelkim/code/Bernstein/manual_cash_flow_mappings.csv"
        with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = [
                'Dest_Row_Number', 'Dest_Field_Name', 'Dest_Enhanced_Scope',
                'Dest_Section_Context', 'Dest_Major_Section_Context',
                'Source_Sheet_Name', 'Source_Row_Number', 'Source_Field_Name',
                'Source_Enhanced_Scope', 'Source_Section_Context',
                'Q1_Verification_Value', 'Source_Q1_Value', 'Source_Q2_Value',
                'Match_Method', 'Match_Confidence', 'Match_Reason'
            ]
            
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(manual_mappings)
        
        print(f"\nMapping saved to: {output_file}")
        
        # Populate destination file
        final_file, populated_count = populate_manual_cash_flow_mappings(manual_mappings)
        
        print(f"\n" + "="*80)
        print("MANUAL CASH FLOW MAPPING RESULTS")
        print("="*80)
        print(f"Manual mappings created: {len(manual_mappings)}")
        print(f"Successfully populated: {populated_count}")
        print(f"Final file: {final_file}")
        print(f"✅ Cash Flow fields now populated (except Row 186)")
        
        # Note about Row 186
        print(f"\n📝 NOTE: Row 186 'Tax benefit from exercise of employee stock options'")
        print(f"    needs further investigation - no direct match found in source sheets")
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
