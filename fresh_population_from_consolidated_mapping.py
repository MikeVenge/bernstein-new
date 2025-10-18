#!/usr/bin/env python3
"""
Fresh Population from Consolidated Mapping

Performs a complete fresh population of the destination file using the 
consolidated mapping file. This starts from the original destination file
and applies all verified mappings in one clean process.

Author: AI Assistant
Date: October 2025
"""

import openpyxl
import pandas as pd
import csv
from pathlib import Path
from typing import Dict, List, Optional


def load_consolidated_mappings() -> List[Dict]:
    """Load all mappings from the consolidated mapping file."""
    
    print("=== LOADING CONSOLIDATED FIELD MAPPINGS ===")
    
    consolidated_file = "/Users/michaelkim/code/Bernstein/CONSOLIDATED_FIELD_MAPPINGS.csv"
    
    if not Path(consolidated_file).exists():
        print(f"ERROR: Consolidated mapping file not found: {consolidated_file}")
        return []
    
    mappings = []
    with open(consolidated_file, 'r', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        mappings = list(reader)
    
    print(f"Loaded {len(mappings)} consolidated mappings")
    
    # Show breakdown by source sheet
    sheet_breakdown = {}
    method_breakdown = {}
    
    for mapping in mappings:
        sheet = mapping['Source_Sheet_Name']
        method = mapping['Match_Method']
        
        sheet_breakdown[sheet] = sheet_breakdown.get(sheet, 0) + 1
        method_breakdown[method] = method_breakdown.get(method, 0) + 1
    
    print(f"\nMappings by source sheet:")
    for sheet, count in sorted(sheet_breakdown.items()):
        print(f"  {sheet}: {count} mappings")
    
    print(f"\nMappings by method:")
    for method, count in sorted(method_breakdown.items()):
        print(f"  {method}: {count} mappings")
    
    return mappings


def fresh_population_from_consolidated(mappings: List[Dict]) -> Dict:
    """Perform fresh population using consolidated mappings."""
    
    print(f"\n" + "="*80)
    print("FRESH POPULATION FROM CONSOLIDATED MAPPINGS")
    print("="*80)
    
    # Use original destination file (before any population)
    original_dest_file = "/Users/michaelkim/code/Bernstein/20240725_IPGP.US-IPG Photonics.xlsx"
    source_file = "/Users/michaelkim/code/Bernstein/IPGP-Financial-Data-Workbook-2024-Q2.xlsx"
    
    print(f"Original destination file: {original_dest_file}")
    print(f"Source file: {source_file}")
    
    # Load workbooks
    source_wb = openpyxl.load_workbook(source_file, data_only=True)
    dest_wb = openpyxl.load_workbook(original_dest_file, data_only=False)
    dest_sheet = dest_wb['Reported']
    
    population_results = []
    values_populated = 0
    source_tracking_added = 0
    errors = []
    sheet_stats = {}
    method_stats = {}
    
    print(f"\nPopulating {len(mappings)} consolidated mappings...")
    
    for i, mapping in enumerate(mappings, 1):
        dest_row = int(mapping['Dest_Row_Number'])
        source_sheet_name = mapping['Source_Sheet_Name']
        source_row = mapping['Source_Row_Number']
        dest_field_name = mapping['Dest_Field_Name']
        source_field_name = mapping['Source_Field_Name']
        match_method = mapping['Match_Method']
        
        print(f"\n[{i}/{len(mappings)}] DEST Row {dest_row}: {dest_field_name}")
        print(f"  Method: {match_method}")
        
        try:
            # Get source sheet and Q2 value
            if source_sheet_name in source_wb.sheetnames:
                source_sheet = source_wb[source_sheet_name]
                
                # Handle source row
                if not source_row or source_row.strip() == '':
                    print(f"  ❌ No source row specified")
                    errors.append(f"Row {dest_row}: No source row specified")
                    continue
                
                # Handle composite source rows (like "30+31+32+33")
                if '+' in str(source_row):
                    # Composite field - sum multiple rows
                    composite_rows = [int(r.strip()) for r in str(source_row).split('+')]
                    composite_q2_value = 0
                    
                    print(f"  Composite field from rows: {composite_rows}")
                    for comp_row in composite_rows:
                        comp_value = source_sheet.cell(comp_row, 93).value or 0
                        composite_q2_value += comp_value
                        print(f"    Row {comp_row}: {comp_value}")
                    
                    source_q2_value = composite_q2_value
                    source_location = f"IPGP-Financial-Data-Workbook-2024-Q2.xlsx|{source_sheet_name}|{source_row}|93"
                    print(f"  Composite Q2 total: {source_q2_value}")
                else:
                    # Single source row
                    source_row_num = int(source_row)
                    source_q2_value = source_sheet.cell(source_row_num, 93).value
                    source_location = f"IPGP-Financial-Data-Workbook-2024-Q2.xlsx|{source_sheet_name}|{source_row_num}|93"
                
                current_dest_value = dest_sheet.cell(dest_row, 71).value
                
                print(f"  From {source_sheet_name} Row {source_row}: {source_field_name}")
                print(f"  Q2 value: {source_q2_value}")
                
                if source_q2_value is not None:
                    # Populate Column BS (71) with Q2 value
                    dest_sheet.cell(dest_row, 71).value = source_q2_value
                    values_populated += 1
                    
                    # Add source tracking to Column BT (72)
                    dest_sheet.cell(dest_row, 72).value = source_location
                    source_tracking_added += 1
                    
                    # Track stats
                    if source_sheet_name not in sheet_stats:
                        sheet_stats[source_sheet_name] = 0
                    sheet_stats[source_sheet_name] += 1
                    
                    if match_method not in method_stats:
                        method_stats[match_method] = 0
                    method_stats[match_method] += 1
                    
                    print(f"  ✅ POPULATED BS: {source_q2_value}")
                    print(f"  ✅ TRACKED BT: {source_location}")
                    
                    population_results.append({
                        'Dest_Row': dest_row,
                        'Dest_Field_Name': dest_field_name,
                        'Source_Sheet': source_sheet_name,
                        'Source_Row': source_row,
                        'Source_Field_Name': source_field_name,
                        'Source_Q2_Value': source_q2_value,
                        'Source_Location': source_location,
                        'Match_Method': match_method,
                        'Previous_Value': current_dest_value,
                        'Status': 'POPULATED'
                    })
                else:
                    print(f"  ❌ No Q2 data available")
                    errors.append(f"Row {dest_row}: No Q2 data in source")
                    population_results.append({
                        'Dest_Row': dest_row,
                        'Dest_Field_Name': dest_field_name,
                        'Source_Sheet': source_sheet_name,
                        'Source_Row': source_row,
                        'Source_Field_Name': source_field_name,
                        'Source_Q2_Value': '',
                        'Source_Location': '',
                        'Match_Method': match_method,
                        'Previous_Value': current_dest_value,
                        'Status': 'NO_Q2_DATA'
                    })
            else:
                print(f"  ❌ Source sheet not found: {source_sheet_name}")
                errors.append(f"Row {dest_row}: Source sheet '{source_sheet_name}' not found")
                
        except Exception as e:
            print(f"  ❌ Error processing row: {e}")
            errors.append(f"Row {dest_row}: {str(e)}")
    
    # Save fresh populated file
    output_file = "/Users/michaelkim/code/Bernstein/FRESH_POPULATED_FROM_CONSOLIDATED_IPGP.xlsx"
    print(f"\nSaving fresh populated file to: {output_file}")
    dest_wb.save(output_file)
    
    source_wb.close()
    dest_wb.close()
    
    return {
        'population_results': population_results,
        'values_populated': values_populated,
        'source_tracking_added': source_tracking_added,
        'total_mappings': len(mappings),
        'sheet_stats': sheet_stats,
        'method_stats': method_stats,
        'errors': errors,
        'output_file': output_file
    }


def generate_fresh_population_audit_trail(population_summary: Dict):
    """Generate audit trail for the fresh population."""
    
    audit_file = "/Users/michaelkim/code/Bernstein/FRESH_POPULATION_AUDIT_TRAIL.csv"
    
    with open(audit_file, 'w', newline='', encoding='utf-8') as csvfile:
        fieldnames = [
            'Dest_Row', 'Dest_Field_Name', 'Source_Sheet', 'Source_Row',
            'Source_Field_Name', 'Source_Q2_Value', 'Source_Location',
            'Match_Method', 'Previous_Value', 'Status'
        ]
        
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(population_summary['population_results'])
    
    print(f"Fresh population audit trail saved to: {audit_file}")
    return audit_file


def main():
    """Main entry point for fresh population from consolidated mapping."""
    
    print("="*80)
    print("FRESH POPULATION FROM CONSOLIDATED MAPPING")
    print("="*80)
    print("Starting fresh from original destination file")
    print("Using consolidated mapping file with all verified mappings")
    
    try:
        # Load consolidated mappings
        consolidated_mappings = load_consolidated_mappings()
        
        if not consolidated_mappings:
            print("ERROR: No consolidated mappings loaded")
            return
        
        # Perform fresh population
        population_summary = fresh_population_from_consolidated(consolidated_mappings)
        
        # Generate audit trail
        audit_file = generate_fresh_population_audit_trail(population_summary)
        
        # Final summary
        print(f"\n" + "="*80)
        print("FRESH POPULATION RESULTS")
        print("="*80)
        print(f"Total mappings processed: {population_summary['total_mappings']}")
        print(f"Values successfully populated (Column BS): {population_summary['values_populated']}")
        print(f"Source locations added (Column BT): {population_summary['source_tracking_added']}")
        print(f"Success rate: {population_summary['values_populated']}/{population_summary['total_mappings']} = {population_summary['values_populated']/population_summary['total_mappings']*100:.1f}%")
        
        print(f"\nPopulation by source sheet:")
        for sheet, count in sorted(population_summary['sheet_stats'].items()):
            print(f"  {sheet}: {count} values")
        
        print(f"\nPopulation by method:")
        for method, count in sorted(population_summary['method_stats'].items()):
            print(f"  {method}: {count} values")
        
        if population_summary['errors']:
            print(f"\nErrors encountered: {len(population_summary['errors'])}")
            for error in population_summary['errors'][:5]:  # Show first 5
                print(f"  {error}")
        
        print(f"\nFinal files:")
        print(f"  📁 Populated Excel: {population_summary['output_file']}")
        print(f"  📁 Audit trail: {audit_file}")
        print(f"  📁 Source mappings: CONSOLIDATED_FIELD_MAPPINGS.csv")
        
        print(f"\n🎉 FRESH POPULATION COMPLETE!")
        print(f"All {population_summary['values_populated']} verified mappings applied from scratch")
        print(f"with full source traceability and audit trail!")
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
