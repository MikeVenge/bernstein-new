#!/usr/bin/env python3
"""
Final Improved Key Metrics Field Mapping with Perfect Context Differentiation

This script properly distinguishes between:
- "North America" under "Revenue by region:" → Revenue_Statement.Geographic_Breakdown.North_America
- "North America" under "Revenue by region (% of total):" → Revenue_Statement.Geographic_Breakdown_Percent.North_America

Author: AI Assistant
Date: October 2025
"""

import openpyxl
import csv
from pathlib import Path
import re
from typing import Dict, List, Tuple, Optional


def analyze_key_metrics_final(source_file: str) -> List[Dict]:
    """
    Final analysis with perfect context differentiation.
    """
    print("Loading workbook for final improved analysis...")
    wb = openpyxl.load_workbook(source_file, data_only=True)
    
    if 'Key Metrics' not in wb.sheetnames:
        print(f"ERROR: Key Metrics sheet not found. Available sheets: {wb.sheetnames}")
        wb.close()
        return []
    
    sheet = wb['Key Metrics']
    
    # Get column headers - focus on the specific columns we need
    column_headers = []
    
    # Add specific columns we're interested in
    important_cols = [92, 93]  # CN (Q1 2024), CO (Q2 2024)
    for col_idx in important_cols:
        header_val = sheet.cell(4, col_idx).value
        if header_val:
            column_headers.append({
                'col_idx': col_idx,
                'header': str(header_val),
                'clean_header': clean_date_header(str(header_val))
            })
    
    # Also get other columns for completeness
    for col_idx in range(2, min(sheet.max_column + 1, 20)):
        if col_idx not in important_cols:
            header_val = sheet.cell(4, col_idx).value
            if header_val:
                column_headers.append({
                    'col_idx': col_idx,
                    'header': str(header_val),
                    'clean_header': clean_date_header(str(header_val))
                })
    
    field_mappings = []
    current_section_context = None
    is_percentage_section = False
    
    for row_idx in range(5, sheet.max_row + 1):  # Start from row 5 to skip headers
        first_col = sheet.cell(row_idx, 1).value
        if not first_col:
            continue
            
        first_col_str = str(first_col).strip()
        
        # Detect section headers (lines ending with colon)
        if first_col_str.endswith(':'):
            current_section_context = first_col_str.rstrip(':')
            is_percentage_section = ('% of total' in first_col_str.lower() or 
                                   'percent' in first_col_str.lower())
            print(f"Section: {current_section_context} (Percentage: {is_percentage_section})")
            continue
        
        # Process data fields
        if (len(first_col_str) > 1 and 
            not first_col_str.lower().startswith('ipg photonics') and
            first_col_str.lower() not in ['quarter ended', 'cumulative quarter ended']):
            
            # Get values
            row_values = {}
            for col_info in column_headers:
                col_idx = col_info['col_idx']
                value = sheet.cell(row_idx, col_idx).value
                row_values[col_info['clean_header']] = value
            
            # Create enhanced scoped name based on current context
            field_name = clean_field_name(first_col_str)
            enhanced_scoped_name = create_context_aware_scoped_name(
                first_col_str, current_section_context, is_percentage_section
            )
            
            field_mapping = {
                'row_number': row_idx,
                'original_field_name': first_col_str,
                'cleaned_field_name': field_name,
                'section_context': current_section_context or '',
                'is_percentage_section': is_percentage_section,
                'enhanced_scoped_name': enhanced_scoped_name,
                'values': row_values,
                'has_data': any(v is not None and str(v).strip() for v in row_values.values())
            }
            
            field_mappings.append(field_mapping)
            
            # Print examples
            if any(keyword in first_col_str.lower() for keyword in ['north america', 'germany', 'china', 'total']):
                print(f"  → {first_col_str} → {enhanced_scoped_name}")
    
    wb.close()
    return field_mappings


def create_context_aware_scoped_name(field_name: str, section_context: str, is_percentage: bool) -> str:
    """
    Create perfectly context-aware scoped names.
    """
    field_lower = field_name.lower()
    section_lower = section_context.lower() if section_context else ''
    
    # Geographic regions
    geographic_mapping = {
        'north america': 'North_America',
        'united states': 'United_States',
        'germany': 'Germany',
        'other europe': 'Other_Europe', 
        'china': 'China',
        'japan': 'Japan',
        'other asia': 'Other_Asia',
        'other asian countries': 'Other_Asia',
        'korea': 'Korea',
        'rest of world': 'Rest_Of_World'
    }
    
    for geo_key, geo_name in geographic_mapping.items():
        if geo_key in field_lower:
            if 'revenue by region' in section_lower:
                if is_percentage:
                    return f"Revenue_Statement.Geographic_Breakdown_Percent.{geo_name}"
                else:
                    return f"Revenue_Statement.Geographic_Breakdown.{geo_name}"
            elif 'assets by region' in section_lower:
                if is_percentage:
                    return f"Balance_Sheet.Geographic_Assets_Percent.{geo_name}"
                else:
                    return f"Balance_Sheet.Geographic_Assets.{geo_name}"
            elif 'employees by region' in section_lower:
                return f"Operations.Employee_Distribution.{geo_name}"
            else:
                return f"Key_Metrics.Geographic_Data.{geo_name}"
    
    # Applications
    application_mapping = {
        'materials processing': 'Materials_Processing',
        'other applications': 'Other_Applications',
        'advanced applications': 'Advanced_Applications',
        'communications': 'Communications',
        'medical': 'Medical'
    }
    
    for app_key, app_name in application_mapping.items():
        if app_key in field_lower:
            if 'revenue by application' in section_lower:
                if is_percentage:
                    return f"Revenue_Statement.Application_Breakdown_Percent.{app_name}"
                else:
                    return f"Revenue_Statement.Application_Breakdown.{app_name}"
            else:
                return f"Key_Metrics.Application_Data.{app_name}"
    
    # Products
    if any(term in field_lower for term in ['high-power', 'pulsed', 'qcw', 'systems', 'laser']):
        clean_name = clean_field_name(field_name)
        if 'revenue by product' in section_lower:
            if is_percentage:
                return f"Revenue_Statement.Product_Breakdown_Percent.{clean_name}"
            else:
                return f"Revenue_Statement.Product_Breakdown.{clean_name}"
        else:
            return f"Key_Metrics.Product_Data.{clean_name}"
    
    # Totals
    if 'total' in field_lower:
        clean_name = clean_field_name(field_name)
        if 'revenue' in section_lower:
            if is_percentage:
                return f"Revenue_Statement.Totals_Percent.{clean_name}"
            else:
                return f"Revenue_Statement.Totals.{clean_name}"
        elif 'backlog' in section_lower:
            return f"Operations.Backlog_Metrics.{clean_name}"
        elif 'assets' in section_lower:
            return f"Balance_Sheet.Asset_Totals.{clean_name}"
        elif 'employee' in section_lower:
            return f"Operations.Employee_Totals.{clean_name}"
        else:
            return f"Key_Metrics.Totals.{clean_name}"
    
    # Customer and other metrics
    if any(term in field_lower for term in ['customer', 'backlog', 'orders']):
        clean_name = clean_field_name(field_name)
        return f"Operations.Customer_Metrics.{clean_name}"
    
    # Employee metrics
    if any(term in field_lower for term in ['employee', 'research', 'manufacturing', 'sales', 'service']):
        clean_name = clean_field_name(field_name)
        return f"Operations.Employee_Metrics.{clean_name}"
    
    # Default fallback
    clean_name = clean_field_name(field_name)
    if section_context:
        section_clean = clean_field_name(section_context)
        return f"Key_Metrics.{section_clean}.{clean_name}"
    else:
        return f"Key_Metrics.Other.{clean_name}"


def clean_field_name(name: str) -> str:
    """Clean and standardize field names."""
    if not name:
        return ""
    
    cleaned = re.sub(r'\s+', '_', name.strip())
    cleaned = re.sub(r'[^\w\s-]', '', cleaned)
    cleaned = re.sub(r'[-_]+', '_', cleaned)
    cleaned = cleaned.strip('_')
    
    if cleaned:
        words = cleaned.split('_')
        cleaned = '_'.join(word.capitalize() for word in words)
    
    return cleaned


def clean_date_header(header) -> str:
    """Clean date headers."""
    # Handle datetime objects
    if hasattr(header, 'strftime'):
        # Convert to Q1, Q2, Q3, Q4 format
        year = header.year
        month = header.month
        quarter = (month - 1) // 3 + 1
        return f"{year}_Q{quarter}"
    
    header_str = str(header).strip()
    
    # Handle string dates like "2024-03-31"
    if re.match(r'\d{4}-\d{2}-\d{2}', header_str):
        year = header_str[:4]
        month = header_str[5:7]
        quarter = (int(month) - 1) // 3 + 1
        return f"{year}_Q{quarter}"
    
    return header_str.replace('-', '_').replace(' ', '_')


def save_final_mapping(field_mappings: List[Dict], output_file: str):
    """Save the final improved mappings."""
    with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
        fieldnames = [
            'Row_Number',
            'Original_Field_Name',
            'Cleaned_Field_Name',
            'Section_Context',
            'Is_Percentage_Section',
            'Enhanced_Scoped_Name',
            'Has_Data',
            'Q1_2024_Value',
            'Q2_2024_Value'
        ]
        
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        
        for mapping in field_mappings:
            # Get Q1 and Q2 2024 values - look for the actual column data
            q1_val = mapping['values'].get('2024_Q1', '')
            q2_val = mapping['values'].get('2024_Q2', '')
            
            # Also try direct column access if date parsing didn't work
            if not q1_val:
                for col_name, col_val in mapping['values'].items():
                    if '2024-03-31' in str(col_name) or 'CN_92' in str(col_name):
                        q1_val = col_val
                        break
            
            if not q2_val:
                for col_name, col_val in mapping['values'].items():
                    if '2024-06-30' in str(col_name) or 'CO_93' in str(col_name):
                        q2_val = col_val
                        break
            
            writer.writerow({
                'Row_Number': mapping['row_number'],
                'Original_Field_Name': mapping['original_field_name'],
                'Cleaned_Field_Name': mapping['cleaned_field_name'],
                'Section_Context': mapping['section_context'],
                'Is_Percentage_Section': mapping['is_percentage_section'],
                'Enhanced_Scoped_Name': mapping['enhanced_scoped_name'],
                'Has_Data': 'Yes' if mapping['has_data'] else 'No',
                'Q1_2024_Value': q1_val if q1_val is not None else '',
                'Q2_2024_Value': q2_val if q2_val is not None else ''
            })


def main():
    """Main entry point."""
    source_file = "/Users/michaelkim/code/Bernstein/IPGP-Financial-Data-Workbook-2024-Q2.xlsx"
    output_file = "/Users/michaelkim/code/Bernstein/final_improved_key_metrics_mapping.csv"
    
    print("="*80)
    print("FINAL IMPROVED KEY METRICS MAPPING")
    print("="*80)
    print("Perfect context differentiation for fields with same names")
    print(f"Source file: {source_file}")
    print(f"Output file: {output_file}")
    
    if not Path(source_file).exists():
        print(f"ERROR: Source file not found: {source_file}")
        return
    
    try:
        field_mappings = analyze_key_metrics_final(source_file)
        
        print(f"\n" + "="*50)
        print("CONTEXT DIFFERENTIATION EXAMPLES")
        print("="*50)
        
        # Show how same field names are differentiated
        field_contexts = {}
        for mapping in field_mappings:
            field_name = mapping['original_field_name']
            if field_name in ['North America', 'Germany', 'China', 'Japan']:
                if field_name not in field_contexts:
                    field_contexts[field_name] = []
                field_contexts[field_name].append({
                    'row': mapping['row_number'],
                    'section': mapping['section_context'],
                    'percentage': mapping['is_percentage_section'],
                    'scoped_name': mapping['enhanced_scoped_name']
                })
        
        for field_name, contexts in field_contexts.items():
            print(f"\n{field_name}:")
            for ctx in contexts:
                pct_indicator = " (Percentage)" if ctx['percentage'] else " (Absolute)"
                print(f"  Row {ctx['row']}: {ctx['section']}{pct_indicator}")
                print(f"    → {ctx['scoped_name']}")
        
        save_final_mapping(field_mappings, output_file)
        
        print(f"\n" + "="*80)
        print("SUCCESS! PERFECT CONTEXT DIFFERENTIATION ACHIEVED")
        print("="*80)
        print(f"Final mapping saved to: {output_file}")
        print(f"Total fields mapped: {len(field_mappings)}")
        print("\nNow fields with identical names are properly distinguished by their context!")
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
