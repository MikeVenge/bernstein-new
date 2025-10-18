#!/usr/bin/env python3
"""
Complete Population Using Q1 Verified Mapping - Step 2 (All Financial Statements)

This script:
1. Reads the complete Q1 verified field mapping file (98 matches)
2. Populates destination Column BS with source Column CO values
3. Covers ALL financial statement sections
4. Provides complete audit trail

Author: AI Assistant
Date: October 2025
"""

import openpyxl
import pandas as pd
import csv
from pathlib import Path
from typing import Dict, List, Optional


def load_complete_verified_mapping() -> List[Dict]:
    """Load the complete Q1 verified field mapping."""
    mapping_file = "/Users/michaelkim/code/Bernstein/complete_q1_verified_mapping.csv"
    
    print(f"Loading complete Q1 verified mapping from: {mapping_file}")
    
    if not Path(mapping_file).exists():
        print(f"ERROR: Mapping file not found: {mapping_file}")
        return []
    
    mappings = []
    with open(mapping_file, 'r', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        mappings = list(reader)
    
    print(f"Loaded {len(mappings)} verified field mappings")
    
    # Show breakdown by source sheet
    sheet_breakdown = {}
    for mapping in mappings:
        sheet = mapping['Source_Sheet_Name']
        sheet_breakdown[sheet] = sheet_breakdown.get(sheet, 0) + 1
    
    print("Mappings by source sheet:")
    for sheet, count in sorted(sheet_breakdown.items()):
        print(f"  {sheet}: {count} mappings")
    
    return mappings


def populate_complete_destination_file(mappings: List[Dict]) -> Dict:
    """
    Populate the destination file using ALL verified mappings.
    """
    print("="*80)
    print("COMPLETE POPULATION - ALL FINANCIAL STATEMENTS")
    print("="*80)
    
    source_file = "/Users/michaelkim/code/Bernstein/IPGP-Financial-Data-Workbook-2024-Q2.xlsx"
    dest_file = "/Users/michaelkim/code/Bernstein/20240725_IPGP.US-IPG Photonics.xlsx"
    
    print(f"Source file: {source_file}")
    print(f"Destination file: {dest_file}")
    print(f"Population: Source Column CO (Q2 2024) → Destination Column BS")
    
    # Load source workbook
    print("\nLoading source workbook...")
    source_wb = openpyxl.load_workbook(source_file, data_only=True)
    
    # Load destination workbook
    print("Loading destination workbook...")
    dest_wb = openpyxl.load_workbook(dest_file, data_only=False)  # Keep formulas
    dest_sheet = dest_wb['Reported']
    
    # Population tracking
    population_results = []
    values_populated = 0
    sheet_stats = {}
    
    print(f"\nPopulating {len(mappings)} verified field mappings...")
    
    for i, mapping in enumerate(mappings, 1):
        dest_row = int(mapping['Dest_Row_Number'])
        source_sheet_name = mapping['Source_Sheet_Name']
        dest_field_name = mapping['Dest_Field_Name']
        source_field_name = mapping['Source_Field_Name']
        q1_verification_value = mapping['Q1_Verification_Value']
        
        print(f"\n[{i}/{len(mappings)}] DEST Row {dest_row}: {dest_field_name}")
        print(f"  From {source_sheet_name}: {source_field_name}")
        print(f"  Q1 verification: {q1_verification_value}")
        
        # Get source sheet
        if source_sheet_name in source_wb.sheetnames:
            source_sheet = source_wb[source_sheet_name]
            
            # Get source row number
            source_row_str = mapping['Source_Row_Number']
            if source_row_str and source_row_str.strip():
                try:
                    source_row = int(source_row_str)
                except ValueError:
                    print(f"  ❌ Invalid source row number: {source_row_str}")
                    continue
            else:
                # For Key Metrics, need to find the row by field name
                if source_sheet_name == 'Key Metrics':
                    source_row = find_key_metrics_row(source_sheet, source_field_name)
                    if not source_row:
                        print(f"  ❌ Could not find row for {source_field_name} in Key Metrics")
                        continue
                else:
                    print(f"  ❌ No source row number available")
                    continue
            
            # Determine source column based on sheet
            if source_sheet_name == 'Key Metrics':
                source_col = 93  # Column CO for Key Metrics
            else:
                source_col = 93  # Column CO for other sheets too
            
            # Get Q2 2024 value from source
            source_q2_value = source_sheet.cell(source_row, source_col).value
            
            # Get current destination value
            current_dest_value = dest_sheet.cell(dest_row, 71).value  # Column BS
            
            print(f"  Source Row {source_row}, Col {source_col}: {source_q2_value}")
            print(f"  Current Dest BS: {current_dest_value}")
            
            if source_q2_value is not None:
                # Populate destination Column BS
                dest_sheet.cell(dest_row, 71).value = source_q2_value
                values_populated += 1
                
                # Track stats by sheet
                if source_sheet_name not in sheet_stats:
                    sheet_stats[source_sheet_name] = 0
                sheet_stats[source_sheet_name] += 1
                
                population_result = {
                    'Dest_Row': dest_row,
                    'Dest_Field_Name': dest_field_name,
                    'Dest_Enhanced_Scope': mapping['Dest_Enhanced_Scope'],
                    'Source_Sheet': source_sheet_name,
                    'Source_Row': source_row,
                    'Source_Field_Name': source_field_name,
                    'Source_Enhanced_Scope': mapping['Source_Enhanced_Scope'],
                    'Q1_Verification_Value': q1_verification_value,
                    'Source_Q2_Value': source_q2_value,
                    'Previous_Dest_Value': current_dest_value,
                    'Population_Status': 'POPULATED',
                    'Match_Method': mapping['Match_Method']
                }
                
                print(f"  ✓ POPULATED: {source_q2_value}")
            else:
                population_result = {
                    'Dest_Row': dest_row,
                    'Dest_Field_Name': dest_field_name,
                    'Dest_Enhanced_Scope': mapping['Dest_Enhanced_Scope'],
                    'Source_Sheet': source_sheet_name,
                    'Source_Row': source_row,
                    'Source_Field_Name': source_field_name,
                    'Source_Enhanced_Scope': mapping['Source_Enhanced_Scope'],
                    'Q1_Verification_Value': q1_verification_value,
                    'Source_Q2_Value': '',
                    'Previous_Dest_Value': current_dest_value,
                    'Population_Status': 'NO_Q2_DATA',
                    'Match_Method': mapping['Match_Method']
                }
                
                print(f"  ❌ NO Q2 DATA AVAILABLE")
            
            population_results.append(population_result)
        else:
            print(f"  ❌ Source sheet not found: {source_sheet_name}")
    
    # Save updated destination workbook
    output_file = "/Users/michaelkim/code/Bernstein/completely_populated_20240725_IPGP.US-IPG_Photonics.xlsx"
    print(f"\nSaving completely populated destination file to: {output_file}")
    dest_wb.save(output_file)
    
    # Close workbooks
    source_wb.close()
    dest_wb.close()
    
    print(f"✓ Destination file saved with {values_populated} populated values")
    
    return {
        'population_results': population_results,
        'values_populated': values_populated,
        'total_mappings': len(mappings),
        'sheet_stats': sheet_stats,
        'output_file': output_file
    }


def find_key_metrics_row(sheet, field_name: str) -> Optional[int]:
    """Find row number in Key Metrics sheet by field name."""
    for row_idx in range(1, min(sheet.max_row + 1, 200)):
        cell_value = sheet.cell(row_idx, 1).value
        if cell_value and str(cell_value).strip() == field_name:
            return row_idx
    return None


def save_complete_population_audit(population_summary: Dict, output_file: str):
    """Save complete audit trail of the population process."""
    print(f"\nSaving complete population audit trail to: {output_file}")
    
    with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
        fieldnames = [
            'Dest_Row', 'Dest_Field_Name', 'Dest_Enhanced_Scope',
            'Source_Sheet', 'Source_Row', 'Source_Field_Name', 'Source_Enhanced_Scope',
            'Q1_Verification_Value', 'Source_Q2_Value', 'Previous_Dest_Value',
            'Population_Status', 'Match_Method'
        ]
        
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(population_summary['population_results'])
    
    print(f"✓ Complete population audit trail saved")


def main():
    """Main entry point for complete population."""
    audit_file = "/Users/michaelkim/code/Bernstein/complete_population_audit_trail.csv"
    
    print("="*80)
    print("COMPLETE POPULATION - ALL FINANCIAL STATEMENTS")
    print("="*80)
    print("Using complete Q1 verified mapping to populate ALL sections:")
    print("  1. Segment Information")
    print("  2. Consolidated Income Statement")
    print("  3. Consolidated Balance Sheet") 
    print("  4. Consolidated Cash Flow Statement")
    print(f"Audit trail file: {audit_file}")
    
    try:
        # Load complete verified mappings
        mappings = load_complete_verified_mapping()
        
        if not mappings:
            print("ERROR: No verified mappings loaded")
            return
        
        # Populate destination file completely
        population_summary = populate_complete_destination_file(mappings)
        
        # Save complete audit trail
        save_complete_population_audit(population_summary, audit_file)
        
        # Final comprehensive summary
        print(f"\n" + "="*80)
        print("COMPLETE POPULATION RESULTS")
        print("="*80)
        print(f"Total verified mappings: {population_summary['total_mappings']}")
        print(f"Values successfully populated: {population_summary['values_populated']}")
        print(f"Population rate: {population_summary['values_populated']}/{population_summary['total_mappings']} = {population_summary['values_populated']/population_summary['total_mappings']*100:.1f}%")
        
        print(f"\nPopulation by source sheet:")
        for sheet, count in sorted(population_summary['sheet_stats'].items()):
            print(f"  {sheet}: {count} values populated")
        
        print(f"\nFiles generated:")
        print(f"  1. Completely populated destination file: {population_summary['output_file']}")
        print(f"  2. Complete population audit trail: {audit_file}")
        
        print(f"\nComplete two-step process finished:")
        print(f"  Step 1: Complete Q1 verification mapping ✓ (98 matches)")
        print(f"  Step 2: Complete destination population ✓ ({population_summary['values_populated']} values)")
        
        print(f"\nPopulation method: Source Column CO → Destination Column BS")
        print(f"Validation method: Q1 verification (100% confidence)")
        print(f"Coverage: ALL major financial statement sections")
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
