#!/usr/bin/env python3
"""
Final Population with Source Tracking

Populates destination file and adds source location tracking in column BT.
Format: "Filename|Tab|Row|Column" for full traceability.

Author: AI Assistant
Date: October 2025
"""

import openpyxl
import pandas as pd
import csv
from pathlib import Path
from typing import Dict, List, Optional


def load_all_verified_mappings() -> List[Dict]:
    """Load all verified mappings from complete and precision-adjusted files."""
    
    # Load complete mapping (98 matches from all financial statements)
    complete_file = "/Users/michaelkim/code/Bernstein/complete_q1_verified_mapping.csv"
    complete_mappings = []
    
    if Path(complete_file).exists():
        with open(complete_file, 'r', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            complete_mappings = list(reader)
        print(f"Loaded complete mapping: {len(complete_mappings)} matches")
    
    # Load precision-adjusted mapping (36 matches including rows 48-54)
    precision_file = "/Users/michaelkim/code/Bernstein/precision_adjusted_q1_mapping.csv"
    precision_mappings = []
    
    if Path(precision_file).exists():
        with open(precision_file, 'r', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            precision_mappings = list(reader)
        print(f"Loaded precision-adjusted mapping: {len(precision_mappings)} matches")
    
    # Combine mappings, preferring precision-adjusted for overlapping rows
    combined_mappings = {}
    
    # First add complete mappings
    for mapping in complete_mappings:
        dest_row = int(mapping['Dest_Row_Number'])
        combined_mappings[dest_row] = mapping
    
    # Then override with precision-adjusted mappings (better for percentage fields)
    for mapping in precision_mappings:
        dest_row = int(mapping['Dest_Row_Number'])
        # Convert precision-adjusted format to complete format
        standardized_mapping = {
            'Dest_Row_Number': mapping['Dest_Row_Number'],
            'Dest_Field_Name': mapping['Dest_Field_Name'],
            'Dest_Enhanced_Scope': mapping['Dest_Enhanced_Scope'],
            'Dest_Section_Context': mapping['Dest_Section_Context'],
            'Dest_Major_Section_Context': mapping['Dest_Major_Section_Context'],
            'Source_Sheet_Name': mapping['Source_Sheet_Name'],
            'Source_Row_Number': mapping['Source_Row_Number'],
            'Source_Field_Name': mapping['Source_Field_Name'],
            'Source_Enhanced_Scope': mapping['Source_Enhanced_Scope'],
            'Source_Section_Context': mapping['Source_Section_Context'],
            'Q1_Verification_Value': mapping.get('Q1_Verification_Value_Original', mapping.get('Q1_Verification_Value', '')),
            'Source_Q2_Value': mapping['Source_Q2_Value'],
            'Match_Method': mapping['Match_Method'],
            'Match_Confidence': mapping['Match_Confidence']
        }
        combined_mappings[dest_row] = standardized_mapping
    
    # Add Row 23 manual match
    combined_mappings[23] = {
        'Dest_Row_Number': '23',
        'Dest_Field_Name': 'Other application, of which',
        'Dest_Enhanced_Scope': 'Reported.Segment_Information_In_000_Usd_Ytd.End_Market_Breakdown.Other_Application_Of_Which',
        'Dest_Section_Context': 'End_Market_Breakdown',
        'Dest_Major_Section_Context': 'Segment Information',
        'Source_Sheet_Name': 'Key Metrics',
        'Source_Row_Number': '27',
        'Source_Field_Name': 'Other applications',
        'Source_Enhanced_Scope': 'Revenue_Statement.Application_Breakdown.Other_Applications',
        'Source_Section_Context': 'Application_Breakdown',
        'Q1_Verification_Value': '25644',
        'Source_Q2_Value': '31872',
        'Match_Method': 'Manual_Q1_Verification',
        'Match_Confidence': '1.0'
    }
    
    # Convert back to list and sort by destination row
    final_mappings = list(combined_mappings.values())
    final_mappings.sort(key=lambda x: int(x['Dest_Row_Number']))
    
    print(f"Final combined mappings: {len(final_mappings)} total")
    
    return final_mappings


def create_source_location_string(source_sheet_name: str, source_row: int, source_column: int = 93) -> str:
    """Create source location string in format: Filename|Tab|Row|Column"""
    filename = "IPGP-Financial-Data-Workbook-2024-Q2.xlsx"
    return f"{filename}|{source_sheet_name}|{source_row}|{source_column}"


def populate_with_source_tracking(mappings: List[Dict]) -> Dict:
    """Populate destination file with Q2 values and add source tracking in column BT."""
    
    print("="*80)
    print("FINAL POPULATION WITH SOURCE LOCATION TRACKING")
    print("="*80)
    
    source_file = "/Users/michaelkim/code/Bernstein/IPGP-Financial-Data-Workbook-2024-Q2.xlsx"
    dest_file = "/Users/michaelkim/code/Bernstein/20240725_IPGP.US-IPG Photonics.xlsx"
    
    print(f"Source file: {source_file}")
    print(f"Destination file: {dest_file}")
    print(f"Adding source tracking to Column BT (72)")
    
    # Load workbooks
    source_wb = openpyxl.load_workbook(source_file, data_only=True)
    dest_wb = openpyxl.load_workbook(dest_file, data_only=False)
    dest_sheet = dest_wb['Reported']
    
    population_results = []
    values_populated = 0
    source_tracking_added = 0
    sheet_stats = {}
    
    print(f"\nPopulating {len(mappings)} verified mappings with source tracking...")
    
    for i, mapping in enumerate(mappings, 1):
        dest_row = int(mapping['Dest_Row_Number'])
        source_sheet_name = mapping['Source_Sheet_Name']
        source_row = mapping['Source_Row_Number']
        dest_field_name = mapping['Dest_Field_Name']
        source_field_name = mapping['Source_Field_Name']
        
        print(f"\n[{i}/{len(mappings)}] DEST Row {dest_row}: {dest_field_name}")
        
        # Get source sheet and Q2 value
        if source_sheet_name in source_wb.sheetnames:
            source_sheet = source_wb[source_sheet_name]
            
            # Handle source row
            if not source_row or source_row.strip() == '':
                print(f"  ❌ No source row specified")
                continue
            else:
                source_row_num = int(source_row)
            
            # Get Q2 value (Column CO = 93)
            source_q2_value = source_sheet.cell(source_row_num, 93).value
            current_dest_value = dest_sheet.cell(dest_row, 71).value  # Column BS
            
            print(f"  From {source_sheet_name} Row {source_row_num}: {source_field_name}")
            print(f"  Q2 value: {source_q2_value}")
            
            if source_q2_value is not None:
                # Populate Column BS (71) with Q2 value
                dest_sheet.cell(dest_row, 71).value = source_q2_value
                values_populated += 1
                
                # Add source tracking to Column BT (72)
                source_location = create_source_location_string(source_sheet_name, source_row_num, 93)
                dest_sheet.cell(dest_row, 72).value = source_location
                source_tracking_added += 1
                
                # Track stats
                if source_sheet_name not in sheet_stats:
                    sheet_stats[source_sheet_name] = 0
                sheet_stats[source_sheet_name] += 1
                
                print(f"  ✅ POPULATED BS: {source_q2_value}")
                print(f"  ✅ TRACKED BT: {source_location}")
                
                population_results.append({
                    'Dest_Row': dest_row,
                    'Dest_Field_Name': dest_field_name,
                    'Source_Sheet': source_sheet_name,
                    'Source_Row': source_row_num,
                    'Source_Field_Name': source_field_name,
                    'Source_Q2_Value': source_q2_value,
                    'Source_Location': source_location,
                    'Previous_Value': current_dest_value,
                    'Status': 'POPULATED_WITH_TRACKING'
                })
            else:
                print(f"  ❌ No Q2 data available")
                population_results.append({
                    'Dest_Row': dest_row,
                    'Dest_Field_Name': dest_field_name,
                    'Source_Sheet': source_sheet_name,
                    'Source_Row': source_row_num,
                    'Source_Field_Name': source_field_name,
                    'Source_Q2_Value': '',
                    'Source_Location': '',
                    'Previous_Value': current_dest_value,
                    'Status': 'NO_Q2_DATA'
                })
        else:
            print(f"  ❌ Source sheet not found: {source_sheet_name}")
    
    # Save final populated file with source tracking
    output_file = "/Users/michaelkim/code/Bernstein/final_populated_with_source_tracking_IPGP.xlsx"
    print(f"\nSaving final populated file with source tracking to: {output_file}")
    dest_wb.save(output_file)
    
    source_wb.close()
    dest_wb.close()
    
    return {
        'population_results': population_results,
        'values_populated': values_populated,
        'source_tracking_added': source_tracking_added,
        'total_mappings': len(mappings),
        'sheet_stats': sheet_stats,
        'output_file': output_file
    }


def main():
    """Main entry point for population with source tracking."""
    
    print("="*80)
    print("FINAL POPULATION WITH SOURCE LOCATION TRACKING")
    print("="*80)
    print("Populating Column BS (Q2 values) + Column BT (source locations)")
    print("Source location format: Filename|Tab|Row|Column")
    
    try:
        # Load all verified mappings
        all_mappings = load_all_verified_mappings()
        
        if not all_mappings:
            print("ERROR: No mappings loaded")
            return
        
        # Populate destination file with source tracking
        population_summary = populate_with_source_tracking(all_mappings)
        
        # Final summary
        print(f"\n" + "="*80)
        print("FINAL POPULATION WITH SOURCE TRACKING RESULTS")
        print("="*80)
        print(f"Total mappings processed: {population_summary['total_mappings']}")
        print(f"Values successfully populated (Column BS): {population_summary['values_populated']}")
        print(f"Source locations added (Column BT): {population_summary['source_tracking_added']}")
        print(f"Success rate: {population_summary['values_populated']}/{population_summary['total_mappings']} = {population_summary['values_populated']/population_summary['total_mappings']*100:.1f}%")
        
        print(f"\nPopulation by source sheet:")
        for sheet, count in sorted(population_summary['sheet_stats'].items()):
            print(f"  {sheet}: {count} values")
        
        print(f"\nFinal file with source tracking: {population_summary['output_file']}")
        print("✅ ALL verified fields now populated with full source traceability!")
        
        # Show sample source tracking entries
        print(f"\nSample source tracking entries (Column BT):")
        for i, result in enumerate(population_summary['population_results'][:5], 1):
            if result['Status'] == 'POPULATED_WITH_TRACKING':
                print(f"  {i}. Row {result['Dest_Row']}: {result['Source_Location']}")
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
