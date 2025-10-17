#!/usr/bin/env python3
"""
Utility to populate column BS in the 'Reported' tab of the target workbook
with data from the source workbook.
"""

import openpyxl
from openpyxl import load_workbook
from pathlib import Path

def populate_data():
    """
    Extract data from IPGP-Financial-Data-Workbook-2024-Q2.xlsx 
    and populate column BS in the 'Reported' tab of 20240725_IPGP.US-IPG Photonics.xlsx
    """
    
    # Define file paths
    source_file = Path("IPGP-Financial-Data-Workbook-2024-Q2.xlsx")
    target_file = Path("20240725_IPGP.US-IPG Photonics.xlsx")
    
    # Check if files exist
    if not source_file.exists():
        print(f"Error: Source file '{source_file}' not found")
        return
    
    if not target_file.exists():
        print(f"Error: Target file '{target_file}' not found")
        return
    
    print(f"Loading source workbook: {source_file}")
    source_wb = load_workbook(source_file, data_only=True)
    
    print(f"Loading target workbook: {target_file}")
    target_wb = load_workbook(target_file)
    
    # Print available sheets in both workbooks
    print(f"\nSource workbook sheets: {source_wb.sheetnames}")
    print(f"Target workbook sheets: {target_wb.sheetnames}")
    
    # Check if 'Reported' tab exists in target workbook
    if 'Reported' not in target_wb.sheetnames:
        print(f"\nError: 'Reported' tab not found in target workbook")
        print(f"Available tabs: {target_wb.sheetnames}")
        return
    
    # Get the Reported sheet
    reported_sheet = target_wb['Reported']
    
    # Display data from each source sheet
    print("\n" + "="*80)
    print("EXAMINING SOURCE SHEETS")
    print("="*80)
    
    for sheet_name in source_wb.sheetnames:
        source_sheet = source_wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} ---")
        print(f"Dimensions: {source_sheet.max_row} rows x {source_sheet.max_column} columns")
        
        # Display first few rows
        print(f"First 5 rows (first 5 columns):")
        for row_idx in range(1, min(6, source_sheet.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(6, source_sheet.max_column + 1)):
                cell_value = source_sheet.cell(row_idx, col_idx).value
                row_data.append(str(cell_value)[:30] if cell_value else "")
            print(f"  Row {row_idx}: {' | '.join(row_data)}")
    
    print("\n" + "="*80)
    print("EXAMINING TARGET 'REPORTED' SHEET")
    print("="*80)
    print(f"Dimensions: {reported_sheet.max_row} rows x {reported_sheet.max_column} columns")
    
    # Display current data in columns BO, BP, BQ, BR, BS with context
    print("\nCurrent data in columns BO-BS (first 20 rows with row labels):")
    bo_col_idx = 67  # BO is the 67th column
    bs_col_idx = 71  # BS is the 71st column
    
    # Print header
    print(f"{'Row':>4} | {'Column A':40s} | {'BO':15s} | {'BP':15s} | {'BQ':15s} | {'BR':15s} | {'BS':15s}")
    print("-" * 130)
    
    for row_idx in range(1, min(21, reported_sheet.max_row + 1)):
        row_label = reported_sheet.cell(row_idx, 1).value  # Column A
        bo_val = reported_sheet.cell(row_idx, 67).value
        bp_val = reported_sheet.cell(row_idx, 68).value
        bq_val = reported_sheet.cell(row_idx, 69).value
        br_val = reported_sheet.cell(row_idx, 70).value
        bs_val = reported_sheet.cell(row_idx, 71).value
        
        print(f"{row_idx:4d} | {str(row_label)[:40]:40s} | {str(bo_val)[:15]:15s} | {str(bp_val)[:15]:15s} | {str(bq_val)[:15]:15s} | {str(br_val)[:15]:15s} | {str(bs_val)[:15]:15s}")
    
    # Find the 2024 Q2 column in source sheets
    print("\n" + "="*80)
    print("SEARCHING FOR 2024 Q2 DATA IN SOURCE")
    print("="*80)
    
    for sheet_name in ['Income Statement', 'Balance Sheet', 'Cash Flows']:
        if sheet_name in source_wb.sheetnames:
            source_sheet = source_wb[sheet_name]
            print(f"\n--- {sheet_name} ---")
            
            # Check row 4 for dates (where quarter dates are)
            print("Checking dates in row 4:")
            for col_idx in range(2, min(20, source_sheet.max_column + 1)):
                date_val = source_sheet.cell(4, col_idx).value
                if date_val:
                    date_str = str(date_val)
                    if '2024' in date_str and ('06-30' in date_str or '2024-06-30' in date_str or 'Q2' in date_str):
                        print(f"  Column {col_idx}: {date_val} <-- 2024 Q2 FOUND!")
                    elif '2024' in date_str:
                        print(f"  Column {col_idx}: {date_val}")
            
            # Show last few columns to find latest data
            print(f"\nLast 5 date columns:")
            for col_idx in range(max(2, source_sheet.max_column - 4), source_sheet.max_column + 1):
                date_val = source_sheet.cell(4, col_idx).value
                print(f"  Column {col_idx}: {date_val}")
    
    # Now populate column BS with 2024 Q2 data from column 93
    # Using previous quarter (Q1) verification
    print("\n" + "="*80)
    print("POPULATING COLUMN BS WITH 2024 Q2 DATA (WITH Q1 VERIFICATION)")
    print("="*80)
    
    source_col_2024q1 = 92  # Column with 2024-03-31 data (Q1)
    source_col_2024q2 = 93  # Column with 2024-06-30 data (Q2)
    target_col_br = 70      # Column BR (previous quarter in target)
    target_col_bs = 71      # Column BS (current quarter to populate)
    
    # Build hierarchical context for field names
    def build_scoped_field_name(sheet_name, row_idx, label, source_sheet):
        """Build a fully scoped field name with context."""
        
        # Look backwards to find the most recent section header
        section_header = None
        
        for check_row in range(row_idx - 1, max(0, row_idx - 10), -1):
            check_label = source_sheet.cell(check_row, 1).value
            if check_label and isinstance(check_label, str):
                check_label = check_label.strip()
                
                # Look for clear section headers
                if any(pattern in check_label.lower() for pattern in [
                    'revenue by region', 'segment breakdown', 'product breakdown',
                    'end market breakdown', 'revenue by application', 'revenue by product type',
                    'assets', 'current assets', 'non-current assets',
                    'liabilities', 'current liabilities', 'equity',
                    'operating activities', 'investing activities', 'financing activities',
                    'cash flows from', 'supplemental'
                ]):
                    section_header = check_label.rstrip(':')
                    break
                
                # Also check for lines ending with colon (section headers)
                elif check_label.endswith(':') and len(check_label) > 5:
                    section_header = check_label.rstrip(':')
                    break
        
        # Build the scoped name
        if section_header:
            scoped_name = f"{sheet_name}.{section_header}.{label}"
        else:
            scoped_name = f"{sheet_name}.{label}"
        
        return scoped_name
    
    # Build a comprehensive lookup: (sheet_name, row_label, q1_value) -> q2_value
    # We'll use Q1 value as part of the key for exact matching
    source_data_by_value = {}  # Maps (sheet, q1_value) -> [(scoped_label, q2_value, row)]
    source_data_by_label = {}  # Maps (sheet, label) -> {q1, q2, row, scoped_name}
    
    for sheet_name in ['Income Statement', 'Balance Sheet', 'Cash Flows', 'Key Metrics']:
        if sheet_name in source_wb.sheetnames:
            source_sheet = source_wb[sheet_name]
            print(f"\nMapping data from '{sheet_name}'...")
            
            for row_idx in range(1, source_sheet.max_row + 1):
                row_label = source_sheet.cell(row_idx, 1).value
                if row_label and isinstance(row_label, str) and row_label.strip():
                    label = row_label.strip()
                    
                    # Build scoped field name with context
                    scoped_name = build_scoped_field_name(sheet_name, row_idx, label, source_sheet)
                    
                    # Get both Q1 and Q2 values
                    value_q1 = source_sheet.cell(row_idx, source_col_2024q1).value
                    value_q2 = source_sheet.cell(row_idx, source_col_2024q2).value
                    
                    # Store by label
                    label_key = (sheet_name, label)
                    source_data_by_label[label_key] = {
                        'q1': value_q1,
                        'q2': value_q2,
                        'source_row': row_idx,
                        'scoped_name': scoped_name
                    }
                    
                    # Also store by Q1 value for reverse lookup
                    if value_q1 is not None:
                        try:
                            q1_float = float(value_q1)
                            value_key = (sheet_name, q1_float)
                            if value_key not in source_data_by_value:
                                source_data_by_value[value_key] = []
                            source_data_by_value[value_key].append({
                                'label': label,
                                'scoped_name': scoped_name,
                                'q2': value_q2,
                                'row': row_idx
                            })
                        except (ValueError, TypeError):
                            pass
    
    print(f"\nTotal items mapped from source: {len(source_data_by_label)}")
    print(f"Total Q1 values indexed: {len(source_data_by_value)}")
    
    # Now populate the Reported sheet with verification
    print(f"\nPopulating column BS in 'Reported' sheet (with Q1 verification)...")
    print("="*100)
    
    values_populated = 0
    values_not_found = []
    
    # First, set the year in row 1 (2024)
    reported_sheet.cell(1, target_col_bs).value = 2024
    print(f"Row   1: Year set to 2024")
    
    # Set "Reported" in row 2
    reported_sheet.cell(2, target_col_bs).value = "Reported"
    print(f"Row   2: Type set to 'Reported'")
    
    # Process data rows
    for target_row_idx in range(5, reported_sheet.max_row + 1):
        target_row_label = reported_sheet.cell(target_row_idx, 1).value
        
        if not (target_row_label and isinstance(target_row_label, str) and target_row_label.strip()):
            continue
        
        label = target_row_label.strip()
        
        # Get the Q1 value from column BR (previous quarter)
        q1_value_target = reported_sheet.cell(target_row_idx, target_col_br).value
        
        # Skip if Q1 is a formula or None
        if q1_value_target is None or (isinstance(q1_value_target, str) and q1_value_target.startswith('=')):
            continue
        
        # Try to convert Q1 to float for matching
        try:
            q1_target_float = float(q1_value_target)
        except (ValueError, TypeError):
            continue
        
        # Try to find matching data using Q1 value first (most reliable)
        found_match = False
        best_matches = []
        
        # Search across all sheets for Q1 value match
        for sheet_name in ['Income Statement', 'Balance Sheet', 'Cash Flows', 'Key Metrics']:
            value_key = (sheet_name, q1_target_float)
            
            # Try exact match first
            if value_key in source_data_by_value:
                for match in source_data_by_value[value_key]:
                    best_matches.append({
                        'sheet': sheet_name,
                        'label': match['label'],
                        'q2': match['q2'],
                        'row': match['row'],
                        'q1': q1_target_float,
                        'match_type': 'exact_value'
                    })
            
            # Try with sign flip (some values might be negated)
            value_key_neg = (sheet_name, -q1_target_float)
            if value_key_neg in source_data_by_value:
                for match in source_data_by_value[value_key_neg]:
                    # Negate Q2 value as well
                    q2_negated = -match['q2'] if match['q2'] is not None else None
                    best_matches.append({
                        'sheet': sheet_name,
                        'label': match['label'],
                        'q2': q2_negated,
                        'row': match['row'],
                        'q1': q1_target_float,
                        'match_type': 'sign_flip'
                    })
            
            # Also try with small tolerance
            tolerance = 0.5
            for (s, v), matches in source_data_by_value.items():
                if s == sheet_name and abs(v - q1_target_float) < tolerance:
                    for match in matches:
                        best_matches.append({
                            'sheet': sheet_name,
                            'label': match['label'],
                            'q2': match['q2'],
                            'row': match['row'],
                            'q1': v,
                            'match_type': 'fuzzy_value',
                            'diff': abs(v - q1_target_float)
                        })
        
        # If we found matches, pick the best one
        if best_matches:
            # Prefer exact value matches, then label matches, then smallest difference
            best_match = None
            
            # First try exact value + label match
            for match in best_matches:
                if match['match_type'] == 'exact_value' and match['label'] == label:
                    best_match = match
                    break
            
            # Then try just exact value match
            if not best_match:
                for match in best_matches:
                    if match['match_type'] == 'exact_value':
                        best_match = match
                        break
            
            # Then try fuzzy with label match
            if not best_match:
                for match in best_matches:
                    if match.get('diff', 999) < 0.1 and match['label'] == label:
                        best_match = match
                        break
            
            # Finally just pick the closest
            if not best_match and best_matches:
                best_matches.sort(key=lambda x: x.get('diff', 0))
                best_match = best_matches[0]
            
            if best_match and best_match['q2'] is not None:
                reported_sheet.cell(target_row_idx, target_col_bs).value = best_match['q2']
                values_populated += 1
                match_info = f"[{best_match['match_type']}]"
                if 'diff' in best_match:
                    match_info += f" diff={best_match['diff']:.2f}"
                print(f"Row {target_row_idx:3d}: {label[:35]:35s} | Q1: {q1_target_float:>12.2f} | Q2: {best_match['q2']:>12} | {best_match['sheet'][:20]:20s} | {match_info}")
                found_match = True
        
        if not found_match:
            values_not_found.append({
                'row': target_row_idx,
                'label': label,
                'q1_target': q1_value_target
            })
    
    # Print summary report
    print("\n" + "="*80)
    print("SUMMARY REPORT")
    print("="*80)
    print(f"Values successfully populated: {values_populated}")
    print(f"Values not found: {len(values_not_found)}")
    
    # Report values not found
    if values_not_found:
        print("\n" + "="*80)
        print("DATA FIELDS NOT FOUND IN SOURCE:")
        print("="*80)
        for item in values_not_found:
            print(f"Row {item['row']:3d}: {item['label'][:60]:60s} | Q1 Target: {item['q1_target']}")
            
            # Search for similar Q1 values in source (for debugging)
            try:
                q1_float = float(item['q1_target'])
                print(f"  Searching for Q1 values close to {q1_float}...")
                
                close_matches = []
                for (sheet, val), matches in source_data_by_value.items():
                    diff = abs(val - q1_float)
                    if diff < 100:  # Within 100 of target
                        for m in matches:
                            close_matches.append({
                                'sheet': sheet,
                                'label': m['label'],
                                'q1': val,
                                'q2': m['q2'],
                                'diff': diff
                            })
                
                # Sort by difference and show top 5
                close_matches.sort(key=lambda x: x['diff'])
                for cm in close_matches[:5]:
                    print(f"    {cm['sheet'][:20]:20s} | {cm['label'][:40]:40s} | Q1: {cm['q1']:>10.2f} (diff: {cm['diff']:>8.2f}) | Q2: {cm['q2']}")
            except:
                pass
    
    # Generate verification report CSV
    print("\n" + "="*80)
    print("GENERATING VERIFICATION REPORT")
    print("="*80)
    
    verification_file = target_file.parent / "verification_report.csv"
    
    with open(verification_file, 'w', encoding='utf-8') as f:
        # Write header with scoped field name
        f.write("Row,Target Field Name,Source Sheet,Source Field Name (Scoped),Q1 Target,Q1 Source,Q2 Target (Populated),Match Status\n")
        
        # Track all processed rows
        processed_rows = {}
        
        # First, add all the successfully matched rows
        for target_row_idx in range(5, reported_sheet.max_row + 1):
            target_row_label = reported_sheet.cell(target_row_idx, 1).value
            
            if not (target_row_label and isinstance(target_row_label, str) and target_row_label.strip()):
                continue
            
            label = target_row_label.strip()
            q1_value_target = reported_sheet.cell(target_row_idx, target_col_br).value
            q2_value_target = reported_sheet.cell(target_row_idx, target_col_bs).value
            
            # Skip formula rows
            if q1_value_target is None or (isinstance(q1_value_target, str) and q1_value_target.startswith('=')):
                continue
            
            # Try to convert Q1 to float
            try:
                q1_target_float = float(q1_value_target)
            except (ValueError, TypeError):
                continue
            
            # Search for match
            found_match = False
            match_sheet = ""
            match_label = ""
            scoped_match_label = ""
            q1_source = ""
            q2_source = ""
            match_status = "NOT FOUND"
            
            # Try exact match first
            for sheet_name in ['Income Statement', 'Balance Sheet', 'Cash Flows', 'Key Metrics']:
                value_key = (sheet_name, q1_target_float)
                
                if value_key in source_data_by_value:
                    for match in source_data_by_value[value_key]:
                        match_sheet = sheet_name
                        match_label = match['label']
                        scoped_match_label = match['scoped_name']
                        q1_source = q1_target_float  # The Q1 value IS the key
                        q2_source = match['q2']
                        match_status = "MATCHED"
                        found_match = True
                        break
                
                if found_match:
                    break
            
            # If not found, check with sign flip
            if not found_match:
                for sheet_name in ['Income Statement', 'Balance Sheet', 'Cash Flows', 'Key Metrics']:
                    value_key_neg = (sheet_name, -q1_target_float)
                    
                    if value_key_neg in source_data_by_value:
                        for match in source_data_by_value[value_key_neg]:
                            match_sheet = sheet_name
                            match_label = match['label']
                            scoped_match_label = match['scoped_name']
                            q1_source = -q1_target_float  # Negated Q1 value
                            q2_source = -match['q2'] if match['q2'] is not None else ''
                            match_status = "MATCHED (SIGN FLIP)"
                            found_match = True
                            break
                    
                    if found_match:
                        break
            
            # Escape commas and quotes in labels for CSV
            label_clean = label.replace('"', '""')
            scoped_match_label_clean = scoped_match_label.replace('"', '""') if scoped_match_label else ""
            
            # Format values for CSV
            q1_src_str = str(q1_source) if q1_source != "" else ""
            q2_src_str = str(q2_source) if q2_source != "" else ""
            q2_tgt_str = str(q2_value_target) if q2_value_target is not None else ""
            
            # Write to CSV with scoped field name
            f.write(f'{target_row_idx},"{label_clean}","{match_sheet}","{scoped_match_label_clean}",{q1_value_target},{q1_src_str},{q2_tgt_str},"{match_status}"\n')
            
            processed_rows[target_row_idx] = True
    
    print(f"Verification report saved to: {verification_file}")
    
    # Save the updated workbook
    output_file = target_file.parent / f"updated_{target_file.name}"
    print(f"\nSaving updated workbook to: {output_file}")
    target_wb.save(output_file)
    
    print(f"\n" + "="*80)
    print(f"COMPLETE!")
    print(f"="*80)
    print(f"Updated workbook saved as: {output_file}")
    print(f"Verification report saved as: {verification_file}")
    print(f"\nStatistics:")
    print(f"  - Successfully populated: {values_populated} values")
    print(f"  - Not found in source: {len(values_not_found)} fields")
    
    # Close workbooks
    source_wb.close()
    target_wb.close()

if __name__ == "__main__":
    try:
        populate_data()
    except Exception as e:
        print(f"\nError occurred: {str(e)}")
        import traceback
        traceback.print_exc()

