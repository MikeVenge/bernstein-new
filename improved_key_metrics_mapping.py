#!/usr/bin/env python3
"""
Improved Key Metrics Field Mapping with Enhanced Context Detection

This script properly handles hierarchical contexts like:
- "North America" under "Revenue by region:"
- "North America" under "Revenue by region (% of total):"

Author: AI Assistant
Date: October 2025
"""

import openpyxl
import csv
from pathlib import Path
import re
from typing import Dict, List, Tuple, Optional


def analyze_key_metrics_with_improved_context(source_file: str) -> List[Dict]:
    """
    Analyze Key Metrics with improved hierarchical context detection.
    """
    print("Loading workbook and analyzing Key Metrics sheet with improved context...")
    wb = openpyxl.load_workbook(source_file, data_only=True)
    
    if 'Key Metrics' not in wb.sheetnames:
        print(f"ERROR: Key Metrics sheet not found. Available sheets: {wb.sheetnames}")
        wb.close()
        return []
    
    sheet = wb['Key Metrics']
    print(f"Key Metrics sheet has {sheet.max_row} rows and {sheet.max_column} columns")
    
    # Get column headers
    header_row = 4
    column_headers = []
    for col_idx in range(2, min(sheet.max_column + 1, 20)):
        header_val = sheet.cell(header_row, col_idx).value
        if header_val:
            column_headers.append({
                'col_idx': col_idx,
                'header': str(header_val),
                'clean_header': clean_date_header(str(header_val))
            })
    
    print(f"Found {len(column_headers)} data columns")
    
    # Process all rows with improved context tracking
    field_mappings = []
    
    # Enhanced context tracking
    current_major_section = None
    current_section = None
    current_subsection = None
    
    # Track section hierarchy more precisely
    section_stack = []
    
    for row_idx in range(1, sheet.max_row + 1):
        first_col = sheet.cell(row_idx, 1).value
        if not first_col:
            continue
            
        first_col_str = str(first_col).strip()
        
        # Skip header rows
        if (row_idx <= 4 or 
            first_col_str.lower().startswith('ipg photonics') or
            first_col_str.lower() in ['key metrics', 'quarter ended', 'cumulative quarter ended']):
            continue
        
        # Enhanced row classification
        row_type, context_info = classify_row_with_context(first_col_str, row_idx)
        
        # Update context hierarchy
        if row_type == 'major_section':
            current_major_section = clean_field_name(first_col_str)
            current_section = None
            current_subsection = None
            section_stack = [current_major_section]
            print(f"Major Section: {current_major_section}")
            
        elif row_type == 'section_header':
            current_section = clean_field_name(first_col_str.rstrip(':'))
            current_subsection = None
            section_stack = [current_major_section, current_section] if current_major_section else [current_section]
            print(f"  Section: {current_section}")
            
        elif row_type == 'subsection_header':
            current_subsection = clean_field_name(first_col_str.rstrip(':'))
            section_stack = [s for s in [current_major_section, current_section, current_subsection] if s]
            print(f"    Subsection: {current_subsection}")
            
        elif row_type == 'data_field':
            # Get all values for this row
            row_values = {}
            for col_info in column_headers:
                col_idx = col_info['col_idx']
                value = sheet.cell(row_idx, col_idx).value
                row_values[col_info['clean_header']] = value
            
            # Create enhanced field mapping with full context
            field_name = clean_field_name(first_col_str)
            
            # Build hierarchical scoped name
            scoped_name = build_hierarchical_scoped_name(
                section_stack, field_name, context_info
            )
            
            # Apply enhanced scoping rules
            enhanced_scoped_name = apply_enhanced_scoping_with_context(
                scoped_name, field_name, section_stack, context_info
            )
            
            field_mapping = {
                'row_number': row_idx,
                'original_field_name': first_col_str,
                'cleaned_field_name': field_name,
                'major_section_context': current_major_section or '',
                'section_context': current_section or '',
                'subsection_context': current_subsection or '',
                'full_context_path': ' > '.join(section_stack),
                'basic_scoped_name': scoped_name,
                'enhanced_scoped_name': enhanced_scoped_name,
                'context_info': context_info,
                'row_type': row_type,
                'values': row_values,
                'has_data': any(v is not None and str(v).strip() for v in row_values.values())
            }
            
            field_mappings.append(field_mapping)
            
            # Print key fields for monitoring
            if any(keyword in first_col_str.lower() for keyword in ['north america', 'germany', 'china', 'total']):
                print(f"      Field: {first_col_str} → {enhanced_scoped_name}")
                print(f"        Context Path: {' > '.join(section_stack)}")
    
    wb.close()
    return field_mappings


def classify_row_with_context(first_col: str, row_idx: int) -> Tuple[str, Dict]:
    """
    Enhanced row classification that captures context information.
    Returns tuple of (row_type, context_info)
    """
    first_col_lower = first_col.lower()
    context_info = {}
    
    # Major sections (clear statement-level headers)
    if any(pattern in first_col_lower for pattern in [
        'key metrics', 'financial statements', 'segment information'
    ]):
        return 'major_section', context_info
    
    # Section headers - must end with colon and match patterns
    if first_col.endswith(':'):
        # Determine section type
        if 'revenue by region' in first_col_lower:
            if '% of total' in first_col_lower or 'percent' in first_col_lower:
                context_info['value_type'] = 'percentage'
                context_info['section_type'] = 'revenue_by_region_pct'
            else:
                context_info['value_type'] = 'absolute'
                context_info['section_type'] = 'revenue_by_region'
            return 'section_header', context_info
            
        elif 'revenue by application' in first_col_lower:
            if '% of total' in first_col_lower or 'percent' in first_col_lower:
                context_info['value_type'] = 'percentage'
                context_info['section_type'] = 'revenue_by_application_pct'
            else:
                context_info['value_type'] = 'absolute'
                context_info['section_type'] = 'revenue_by_application'
            return 'section_header', context_info
            
        elif 'revenue by product' in first_col_lower:
            if '% of total' in first_col_lower or 'percent' in first_col_lower:
                context_info['value_type'] = 'percentage'
                context_info['section_type'] = 'revenue_by_product_pct'
            else:
                context_info['value_type'] = 'absolute'
                context_info['section_type'] = 'revenue_by_product'
            return 'section_header', context_info
            
        elif any(pattern in first_col_lower for pattern in [
            'long', 'lived assets', 'employees by', 'backlog', 'customer'
        ]):
            context_info['section_type'] = 'other_metrics'
            return 'section_header', context_info
        else:
            return 'subsection_header', context_info
    
    # Data fields
    if (len(first_col) > 1 and 
        not first_col_lower.startswith('ipg photonics') and
        first_col_lower not in ['quarter ended', 'cumulative quarter ended']):
        
        # Detect field characteristics
        if any(geo in first_col_lower for geo in [
            'north america', 'germany', 'china', 'japan', 'europe', 'asia', 'korea'
        ]):
            context_info['field_type'] = 'geographic'
        elif any(app in first_col_lower for app in [
            'materials processing', 'communications', 'medical', 'advanced'
        ]):
            context_info['field_type'] = 'application'
        elif any(prod in first_col_lower for prod in [
            'high-power', 'pulsed', 'qcw', 'systems', 'laser'
        ]):
            context_info['field_type'] = 'product'
        elif 'total' in first_col_lower:
            context_info['field_type'] = 'total'
        else:
            context_info['field_type'] = 'other'
            
        return 'data_field', context_info
    
    return 'other', context_info


def build_hierarchical_scoped_name(section_stack: List[str], field_name: str, context_info: Dict) -> str:
    """
    Build hierarchical scoped name using the full context stack.
    """
    components = ['Key_Metrics']
    
    # Add all levels from section stack
    components.extend(section_stack)
    
    # Add field name
    if field_name:
        components.append(field_name)
    
    return '.'.join(components)


def apply_enhanced_scoping_with_context(scoped_name: str, field_name: str, 
                                       section_stack: List[str], context_info: Dict) -> str:
    """
    Apply enhanced scoping rules using context information.
    """
    field_lower = field_name.lower()
    
    # Use context information to create more precise scoping
    section_type = context_info.get('section_type', '')
    value_type = context_info.get('value_type', '')
    field_type = context_info.get('field_type', '')
    
    # Geographic fields with context awareness
    geographic_mapping = {
        'north_america': 'North_America',
        'united_states': 'United_States',
        'germany': 'Germany', 
        'other_europe': 'Other_Europe',
        'china': 'China',
        'japan': 'Japan',
        'other_asia': 'Other_Asia',
        'korea': 'Korea',
        'rest_of_world': 'Rest_Of_World'
    }
    
    for geo_key, geo_name in geographic_mapping.items():
        if (geo_key.replace('_', ' ') in field_lower or 
            geo_key in field_lower or
            field_lower == geo_key.split('_')[-1]):
            
            if 'revenue_by_region' in section_type:
                if value_type == 'percentage':
                    return f"Revenue_Statement.Geographic_Breakdown_Percent.{geo_name}"
                else:
                    return f"Revenue_Statement.Geographic_Breakdown.{geo_name}"
            else:
                return f"Key_Metrics.Geographic_Data.{geo_name}"
    
    # Application fields with context
    application_mapping = {
        'materials_processing': 'Materials_Processing',
        'other_applications': 'Other_Applications',
        'advanced_applications': 'Advanced_Applications',
        'communications': 'Communications',
        'medical': 'Medical'
    }
    
    for app_key, app_name in application_mapping.items():
        if app_key.replace('_', ' ') in field_lower:
            if 'revenue_by_application' in section_type:
                if value_type == 'percentage':
                    return f"Revenue_Statement.Application_Breakdown_Percent.{app_name}"
                else:
                    return f"Revenue_Statement.Application_Breakdown.{app_name}"
            else:
                return f"Key_Metrics.Application_Data.{app_name}"
    
    # Product fields with context
    if field_type == 'product' or any(term in field_lower for term in ['high-power', 'pulsed', 'qcw', 'systems']):
        if 'revenue_by_product' in section_type:
            if value_type == 'percentage':
                return f"Revenue_Statement.Product_Breakdown_Percent.{field_name}"
            else:
                return f"Revenue_Statement.Product_Breakdown.{field_name}"
        else:
            return f"Key_Metrics.Product_Data.{field_name}"
    
    # Total fields with context
    if 'total' in field_lower:
        if 'revenue' in section_type:
            if value_type == 'percentage':
                return f"Revenue_Statement.Totals_Percent.{field_name}"
            else:
                return f"Revenue_Statement.Totals.{field_name}"
        else:
            return f"Key_Metrics.Totals.{field_name}"
    
    # Default: return enhanced version of original with context
    return scoped_name


def clean_field_name(name: str) -> str:
    """Clean and standardize field names."""
    if not name:
        return ""
    
    # Remove extra whitespace and special characters
    cleaned = re.sub(r'\s+', '_', name.strip())
    cleaned = re.sub(r'[^\w\s-]', '', cleaned)
    cleaned = re.sub(r'[-_]+', '_', cleaned)
    cleaned = cleaned.strip('_')
    
    # Capitalize appropriately
    if cleaned:
        words = cleaned.split('_')
        cleaned = '_'.join(word.capitalize() for word in words)
    
    return cleaned


def clean_date_header(header: str) -> str:
    """Clean date headers for column names."""
    if 'datetime' in str(type(header)).lower():
        return header.strftime('%Y_Q%m')
    
    header_str = str(header).strip()
    if re.match(r'\d{4}-\d{2}-\d{2}', header_str):
        year = header_str[:4]
        month = header_str[5:7]
        quarter = f"Q{(int(month) - 1) // 3 + 1}"
        return f"{year}_{quarter}"
    
    return header_str.replace('-', '_').replace(' ', '_')


def save_improved_mapping(field_mappings: List[Dict], output_file: str):
    """Save the improved field mappings to CSV."""
    with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
        fieldnames = [
            'Row_Number',
            'Original_Field_Name',
            'Cleaned_Field_Name',
            'Major_Section_Context',
            'Section_Context',
            'Subsection_Context',
            'Full_Context_Path',
            'Basic_Scoped_Name',
            'Enhanced_Scoped_Name',
            'Context_Info',
            'Row_Type',
            'Has_Data',
            '2024_Q1_Value',  # Add specific quarter for testing
            '2024_Q2_Value'
        ]
        
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        
        for mapping in field_mappings:
            # Get Q1 and Q2 2024 values if available
            q1_2024 = mapping['values'].get('2024_Q1', '')
            q2_2024 = mapping['values'].get('2024_Q2', '')
            
            writer.writerow({
                'Row_Number': mapping['row_number'],
                'Original_Field_Name': mapping['original_field_name'],
                'Cleaned_Field_Name': mapping['cleaned_field_name'],
                'Major_Section_Context': mapping['major_section_context'],
                'Section_Context': mapping['section_context'],
                'Subsection_Context': mapping['subsection_context'],
                'Full_Context_Path': mapping['full_context_path'],
                'Basic_Scoped_Name': mapping['basic_scoped_name'],
                'Enhanced_Scoped_Name': mapping['enhanced_scoped_name'],
                'Context_Info': str(mapping['context_info']),
                'Row_Type': mapping['row_type'],
                'Has_Data': 'Yes' if mapping['has_data'] else 'No',
                '2024_Q1_Value': q1_2024 if q1_2024 is not None else '',
                '2024_Q2_Value': q2_2024 if q2_2024 is not None else ''
            })


def main():
    """Main entry point."""
    source_file = "/Users/michaelkim/code/Bernstein/IPGP-Financial-Data-Workbook-2024-Q2.xlsx"
    output_file = "/Users/michaelkim/code/Bernstein/improved_key_metrics_mapping.csv"
    
    print("="*80)
    print("IMPROVED KEY METRICS FIELD MAPPING WITH ENHANCED CONTEXT")
    print("="*80)
    print(f"Source file: {source_file}")
    print(f"Output file: {output_file}")
    
    if not Path(source_file).exists():
        print(f"ERROR: Source file not found: {source_file}")
        return
    
    try:
        # Analyze with improved context
        field_mappings = analyze_key_metrics_with_improved_context(source_file)
        
        if not field_mappings:
            print("ERROR: No field mappings generated")
            return
        
        print(f"\n" + "="*50)
        print("RESULTS SUMMARY")
        print("="*50)
        print(f"Total fields mapped: {len(field_mappings)}")
        
        # Show context differentiation
        context_examples = {}
        for mapping in field_mappings:
            field_name = mapping['original_field_name']
            if field_name.lower() in ['north america', 'germany', 'china']:
                context_path = mapping['full_context_path']
                enhanced_scope = mapping['enhanced_scoped_name']
                if field_name not in context_examples:
                    context_examples[field_name] = []
                context_examples[field_name].append({
                    'context': context_path,
                    'scope': enhanced_scope,
                    'row': mapping['row_number']
                })
        
        print("\nContext Differentiation Examples:")
        for field_name, contexts in context_examples.items():
            print(f"\n{field_name}:")
            for ctx in contexts:
                print(f"  Row {ctx['row']}: {ctx['context']} → {ctx['scope']}")
        
        # Save results
        save_improved_mapping(field_mappings, output_file)
        
        print(f"\n" + "="*80)
        print("PROCESSING COMPLETE!")
        print("="*80)
        print(f"Improved field mapping saved to: {output_file}")
        print("Now fields with same names but different contexts are properly distinguished!")
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
