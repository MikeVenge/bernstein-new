#!/usr/bin/env python3
"""
Comprehensive Field Scoping Processor for IPGP Financial Data

This script creates fully contextualized field names with proper hierarchical scoping
for ALL fields in the financial statements, not just those with obvious data patterns.

Strategy: Capture every meaningful row and create full hierarchical scoping.

Author: AI Assistant
Date: October 2025
"""

import openpyxl
import csv
from pathlib import Path
import re
from typing import Dict, List, Tuple, Optional


def process_all_financial_sheets(source_file: str, output_file: str):
    """
    Process all financial statement sheets and create comprehensive scoped field names.
    """
    print("="*80)
    print("COMPREHENSIVE FIELD SCOPING PROCESSOR")
    print("="*80)
    print(f"Source file: {source_file}")
    print(f"Output file: {output_file}")
    
    wb = openpyxl.load_workbook(source_file, data_only=True)
    relevant_sheets = ['Income Statement', 'Balance Sheet', 'Cash Flows', 'Key Metrics']
    available_sheets = [name for name in relevant_sheets if name in wb.sheetnames]
    
    print(f"\nProcessing sheets: {available_sheets}")
    
    all_scoped_fields = []
    
    for sheet_name in available_sheets:
        sheet = wb[sheet_name]
        print(f"\n" + "="*50)
        print(f"PROCESSING SHEET: {sheet_name}")
        print("="*50)
        
        # Build context hierarchy for this sheet
        scoped_fields = process_sheet_with_full_context(sheet, sheet_name)
        all_scoped_fields.extend(scoped_fields)
        
        print(f"Captured {len(scoped_fields)} fields from {sheet_name}")
    
    # Save to CSV
    save_comprehensive_results(all_scoped_fields, output_file)
    
    wb.close()
    
    print(f"\n" + "="*80)
    print("PROCESSING COMPLETE!")
    print("="*80)
    print(f"Total fields processed: {len(all_scoped_fields)}")
    print(f"Results saved to: {output_file}")


def process_sheet_with_full_context(sheet, sheet_name: str) -> List[Dict]:
    """
    Process a single sheet with full hierarchical context tracking.
    """
    scoped_fields = []
    
    # Context stack
    current_statement = clean_field_name(sheet_name)
    current_section = None
    current_subsection = None
    current_subsubsection = None
    
    # Track indentation levels and patterns
    section_patterns = [
        r'revenue by', r'segment', r'breakdown', r'application', r'product',
        r'region', r'assets', r'liabilities', r'equity', r'cash flows',
        r'operating activities', r'investing activities', r'financing activities'
    ]
    
    for row_idx in range(1, min(sheet.max_row + 1, 200)):
        # Get row data
        row_values = []
        for col_idx in range(1, min(21, sheet.max_column + 1)):
            val = sheet.cell(row_idx, col_idx).value
            row_values.append(val)
        
        first_col = str(row_values[0]).strip() if row_values[0] is not None else ""
        
        if not first_col:
            continue
            
        # Skip header rows
        if (first_col.lower().startswith('ipg photonics') or 
            first_col.lower() in ['quarter ended', 'cumulative quarter ended'] or
            row_idx <= 4):  # Skip first few header rows
            continue
        
        # Classify and update context
        row_type = classify_row_comprehensive(first_col, row_values)
        
        if row_type == 'major_section':
            current_section = clean_field_name(first_col.rstrip(':'))
            current_subsection = None
            current_subsubsection = None
            print(f"  Section: {current_section}")
            
        elif row_type == 'subsection':
            current_subsection = clean_field_name(first_col.rstrip(':'))
            current_subsubsection = None
            print(f"    Subsection: {current_subsection}")
            
        elif row_type == 'subsubsection':
            current_subsubsection = clean_field_name(first_col.rstrip(':'))
            
        elif row_type == 'data_field':
            # Create comprehensive scoped name
            field_name = clean_field_name(first_col)
            scoped_name = build_comprehensive_scoped_name(
                current_statement,
                current_section,
                current_subsection,
                current_subsubsection,
                field_name
            )
            
            # Apply enhanced scoping rules
            enhanced_scoped_name = apply_comprehensive_scoping_rules(
                scoped_name, field_name, current_statement, current_section
            )
            
            # Get sample values
            sample_values = []
            for val in row_values[1:6]:
                if val is not None and str(val).strip():
                    sample_values.append(str(val).strip()[:20])
            
            scoped_field = {
                'sheet': sheet_name,
                'row': row_idx,
                'original_name': first_col,
                'field_name': field_name,
                'basic_scoped_name': scoped_name,
                'enhanced_scoped_name': enhanced_scoped_name,
                'statement': current_statement,
                'section': current_section,
                'subsection': current_subsection,
                'subsubsection': current_subsubsection,
                'row_type': row_type,
                'sample_values': sample_values[:3]
            }
            
            scoped_fields.append(scoped_field)
            
            # Print progress for key fields
            if any(keyword in first_col.lower() for keyword in ['total', 'revenue', 'income', 'cash']):
                print(f"    Field: {first_col} → {enhanced_scoped_name}")
    
    return scoped_fields


def classify_row_comprehensive(first_col: str, row_values: List) -> str:
    """
    Comprehensive row classification with more liberal rules.
    """
    first_col_lower = first_col.lower()
    
    # Major sections (clear section headers)
    major_section_patterns = [
        r'revenue by', r'segment breakdown', r'product breakdown',
        r'application breakdown', r'end market', r'by region',
        r'assets', r'current assets', r'non.?current assets',
        r'liabilities', r'current liabilities', r'equity',
        r'cash flows? from', r'operating activities',
        r'investing activities', r'financing activities'
    ]
    
    for pattern in major_section_patterns:
        if re.search(pattern, first_col_lower):
            return 'major_section'
    
    # Subsections (lines ending with colon)
    if first_col.endswith(':') and len(first_col) > 5:
        return 'subsection'
    
    # Sub-subsections (indented or specific patterns)
    if (first_col.startswith('  ') or 
        any(pattern in first_col_lower for pattern in ['total', 'subtotal'])):
        return 'subsubsection'
    
    # Data fields - be very liberal
    # Any meaningful text that's not clearly a header
    if (len(first_col) > 1 and 
        not first_col_lower.startswith('ipg photonics') and
        first_col_lower not in ['quarter ended', 'cumulative quarter ended', ''] and
        not first_col.startswith('(') and
        not re.match(r'^\d{4}-\d{2}-\d{2}', first_col)):  # Not a date
        return 'data_field'
    
    return 'other'


def clean_field_name(name: str) -> str:
    """Enhanced field name cleaning."""
    if not name:
        return ""
    
    # Remove extra whitespace and special characters
    cleaned = re.sub(r'\s+', '_', name.strip())
    cleaned = re.sub(r'[^\w\s-]', '', cleaned)
    cleaned = re.sub(r'[-_]+', '_', cleaned)
    cleaned = cleaned.strip('_')
    
    # Handle special cases
    if cleaned.lower() == 'total':
        return 'Total'
    if cleaned.lower() == 'other':
        return 'Other'
    
    # Capitalize appropriately
    if cleaned:
        words = cleaned.split('_')
        cleaned = '_'.join(word.capitalize() for word in words)
    
    return cleaned


def build_comprehensive_scoped_name(statement: str, section: str, subsection: str, 
                                   subsubsection: str, field: str) -> str:
    """
    Build comprehensive hierarchical scoped name.
    """
    components = [comp for comp in [statement, section, subsection, subsubsection, field] if comp]
    
    if not components:
        return "Unknown_Field"
    
    return '.'.join(components)


def apply_comprehensive_scoping_rules(scoped_name: str, field_name: str, 
                                    statement: str, section: str) -> str:
    """
    Apply comprehensive financial statement scoping rules.
    """
    field_lower = field_name.lower()
    
    # Geographic regions - highest priority
    geographic_terms = {
        'north_america': 'North_America',
        'united_states': 'United_States', 
        'germany': 'Germany',
        'china': 'China',
        'japan': 'Japan',
        'korea': 'Korea',
        'other_europe': 'Other_Europe',
        'other_asia': 'Other_Asia',
        'rest_of_world': 'Rest_Of_World'
    }
    
    for geo_key, geo_name in geographic_terms.items():
        if geo_key.replace('_', ' ') in field_lower or geo_key in field_lower:
            return f"Revenue_Statement.Geographic_Breakdown.{geo_name}"
    
    # Product categories
    product_terms = {
        'materials_processing': 'Materials_Processing',
        'other_applications': 'Other_Applications',
        'advanced_applications': 'Advanced_Applications',
        'communications': 'Communications',
        'medical': 'Medical',
        'high_power': 'High_Power_CW_Lasers',
        'pulsed': 'Pulsed_Lasers',
        'qcw': 'QCW_Lasers',
        'systems': 'Systems'
    }
    
    for prod_key, prod_name in product_terms.items():
        if prod_key.replace('_', ' ') in field_lower or prod_key in field_lower:
            return f"Revenue_Statement.Product_Breakdown.{prod_name}"
    
    # Financial statement specific scoping
    if statement and 'cash' in statement.lower():
        if section and 'operating' in section.lower():
            return f"CashFlow_Statement.Operating_Activities.{field_name}"
        elif section and 'investing' in section.lower():
            return f"CashFlow_Statement.Investing_Activities.{field_name}"
        elif section and 'financing' in section.lower():
            return f"CashFlow_Statement.Financing_Activities.{field_name}"
        else:
            return f"CashFlow_Statement.{field_name}"
    
    elif statement and 'balance' in statement.lower():
        if 'asset' in field_lower:
            return f"Balance_Sheet.Assets.{field_name}"
        elif any(term in field_lower for term in ['liability', 'payable', 'debt']):
            return f"Balance_Sheet.Liabilities.{field_name}"
        elif any(term in field_lower for term in ['equity', 'retained', 'capital']):
            return f"Balance_Sheet.Equity.{field_name}"
        else:
            return f"Balance_Sheet.{field_name}"
    
    elif statement and 'income' in statement.lower():
        if any(term in field_lower for term in ['revenue', 'sales', 'net_sales']):
            return f"Income_Statement.Revenue.{field_name}"
        elif any(term in field_lower for term in ['cost', 'expense']):
            return f"Income_Statement.Expenses.{field_name}"
        elif 'total' in field_lower:
            return f"Income_Statement.Totals.{field_name}"
        else:
            return f"Income_Statement.{field_name}"
    
    # Default: return enhanced version of original
    return scoped_name


def save_comprehensive_results(scoped_fields: List[Dict], output_file: str):
    """
    Save comprehensive results to CSV.
    """
    with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
        fieldnames = [
            'Sheet_Name',
            'Row_Number', 
            'Original_Field_Name',
            'Cleaned_Field_Name',
            'Basic_Scoped_Name',
            'Enhanced_Scoped_Name',
            'Statement_Context',
            'Section_Context',
            'Subsection_Context',
            'Subsubsection_Context',
            'Row_Type',
            'Sample_Values'
        ]
        
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        
        for field in scoped_fields:
            writer.writerow({
                'Sheet_Name': field['sheet'],
                'Row_Number': field['row'],
                'Original_Field_Name': field['original_name'],
                'Cleaned_Field_Name': field['field_name'],
                'Basic_Scoped_Name': field['basic_scoped_name'],
                'Enhanced_Scoped_Name': field['enhanced_scoped_name'],
                'Statement_Context': field['statement'] or '',
                'Section_Context': field['section'] or '',
                'Subsection_Context': field['subsection'] or '',
                'Subsubsection_Context': field['subsubsection'] or '',
                'Row_Type': field['row_type'],
                'Sample_Values': ' | '.join(field['sample_values'])
            })


def main():
    """Main entry point."""
    source_file = "/Users/michaelkim/code/Bernstein/IPGP-Financial-Data-Workbook-2024-Q2.xlsx"
    output_file = "/Users/michaelkim/code/Bernstein/comprehensive_field_scoping_results.csv"
    
    if not Path(source_file).exists():
        print(f"ERROR: Source file not found: {source_file}")
        return
    
    try:
        process_all_financial_sheets(source_file, output_file)
        
        print(f"\n" + "="*80)
        print("SUCCESS!")
        print("="*80)
        print(f"Comprehensive field scoping complete.")
        print(f"Review the results in: {output_file}")
        print("\nThis CSV contains fully scoped field names like:")
        print("- Revenue_Statement.Geographic_Breakdown.China")
        print("- CashFlow_Statement.Operating_Activities.Total")
        print("- Income_Statement.Revenue.Net_Sales")
        print("- Balance_Sheet.Assets.Total_Current_Assets")
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
