#!/usr/bin/env python3
"""
Parameterized Field Mapper

A flexible, parameterized field mapping and population tool that can be used
with different source files, destination files, mapping files, and target columns.

Usage:
    python parameterized_field_mapper.py --source SOURCE_FILE --destination DEST_FILE 
                                        --mapping MAPPING_FILE --column COLUMN_NUMBER
                                        [--output OUTPUT_FILE] [--audit AUDIT_FILE]

Author: AI Assistant
Date: October 2025
"""

import openpyxl
import pandas as pd
import csv
import argparse
from pathlib import Path
from typing import Dict, List, Optional, Tuple


class ParameterizedFieldMapper:
    """Parameterized field mapping and population engine."""
    
    def __init__(self, source_file: str, destination_file: str, mapping_file: str, 
                 target_column: int, output_file: Optional[str] = None, 
                 audit_file: Optional[str] = None):
        """
        Initialize the parameterized field mapper.
        
        Args:
            source_file: Path to source Excel file
            destination_file: Path to destination Excel file
            mapping_file: Path to CSV mapping file
            target_column: Column number to populate in destination (1-based)
            output_file: Optional output file path (default: auto-generated)
            audit_file: Optional audit trail file path (default: auto-generated)
        """
        self.source_file = Path(source_file)
        self.destination_file = Path(destination_file)
        self.mapping_file = Path(mapping_file)
        self.target_column = target_column
        self.source_tracking_column = target_column + 1  # Next column for source tracking
        
        # Auto-generate output paths if not provided
        if output_file:
            self.output_file = Path(output_file)
        else:
            self.output_file = self.destination_file.parent / f"populated_{self.destination_file.name}"
        
        if audit_file:
            self.audit_file = Path(audit_file)
        else:
            self.audit_file = self.destination_file.parent / f"audit_trail_{self.destination_file.stem}.csv"
        
        # Statistics
        self.stats = {
            'mappings_processed': 0,
            'values_populated': 0,
            'source_tracking_added': 0,
            'errors': [],
            'sheet_stats': {},
            'method_stats': {}
        }
    
    def load_mapping_file(self) -> List[Dict]:
        """Load and validate the mapping file."""
        
        print(f"=== LOADING MAPPING FILE ===")
        print(f"Mapping file: {self.mapping_file}")
        
        if not self.mapping_file.exists():
            raise FileNotFoundError(f"Mapping file not found: {self.mapping_file}")
        
        mappings = []
        with open(self.mapping_file, 'r', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            mappings = list(reader)
        
        print(f"Loaded {len(mappings)} mappings")
        
        # Validate required fields
        required_fields = ['Dest_Row_Number', 'Source_Sheet_Name', 'Source_Row_Number', 'Source_Q2_Value']
        missing_fields = []
        
        for field in required_fields:
            if field not in mappings[0].keys():
                missing_fields.append(field)
        
        if missing_fields:
            raise ValueError(f"Missing required fields in mapping file: {missing_fields}")
        
        # Show breakdown
        sheet_breakdown = {}
        for mapping in mappings:
            sheet = mapping['Source_Sheet_Name']
            sheet_breakdown[sheet] = sheet_breakdown.get(sheet, 0) + 1
        
        print(f"Mappings by source sheet:")
        for sheet, count in sorted(sheet_breakdown.items()):
            print(f"  {sheet}: {count} mappings")
        
        return mappings
    
    def validate_files(self) -> Tuple[bool, List[str]]:
        """Validate that all required files exist and are accessible."""
        
        print(f"\n=== VALIDATING INPUT FILES ===")
        errors = []
        
        # Check source file
        if not self.source_file.exists():
            errors.append(f"Source file not found: {self.source_file}")
        else:
            print(f"✅ Source file: {self.source_file}")
        
        # Check destination file
        if not self.destination_file.exists():
            errors.append(f"Destination file not found: {self.destination_file}")
        else:
            print(f"✅ Destination file: {self.destination_file}")
        
        # Check mapping file
        if not self.mapping_file.exists():
            errors.append(f"Mapping file not found: {self.mapping_file}")
        else:
            print(f"✅ Mapping file: {self.mapping_file}")
        
        # Validate column number
        if self.target_column < 1:
            errors.append(f"Invalid target column: {self.target_column} (must be >= 1)")
        else:
            print(f"✅ Target column: {self.target_column} (tracking column: {self.source_tracking_column})")
        
        return len(errors) == 0, errors
    
    def create_source_location_string(self, source_sheet_name: str, source_row: str, 
                                    source_column: int = 93) -> str:
        """Create source location string for tracking."""
        return f"{self.source_file.name}|{source_sheet_name}|{source_row}|{source_column}"
    
    def process_single_mapping(self, mapping: Dict, source_wb: openpyxl.Workbook, 
                             dest_sheet: openpyxl.worksheet.worksheet.Worksheet) -> Dict:
        """Process a single field mapping."""
        
        dest_row = int(mapping['Dest_Row_Number'])
        source_sheet_name = mapping['Source_Sheet_Name']
        source_row = mapping['Source_Row_Number']
        dest_field_name = mapping.get('Dest_Field_Name', f'Row {dest_row}')
        source_field_name = mapping.get('Source_Field_Name', 'Unknown')
        match_method = mapping.get('Match_Method', 'Unknown')
        expected_q2_value = mapping['Source_Q2_Value']
        
        result = {
            'Dest_Row': dest_row,
            'Dest_Field_Name': dest_field_name,
            'Source_Sheet': source_sheet_name,
            'Source_Row': source_row,
            'Source_Field_Name': source_field_name,
            'Expected_Q2_Value': expected_q2_value,
            'Actual_Q2_Value': None,
            'Source_Location': '',
            'Match_Method': match_method,
            'Status': 'ERROR'
        }
        
        try:
            # Validate source sheet exists
            if source_sheet_name not in source_wb.sheetnames:
                result['Status'] = 'SOURCE_SHEET_NOT_FOUND'
                self.stats['errors'].append(f"Row {dest_row}: Source sheet '{source_sheet_name}' not found")
                return result
            
            source_sheet = source_wb[source_sheet_name]
            
            # Handle source row
            if not source_row or source_row.strip() == '':
                result['Status'] = 'NO_SOURCE_ROW'
                self.stats['errors'].append(f"Row {dest_row}: No source row specified")
                return result
            
            # Handle composite source rows (like "30+31+32+33")
            if '+' in str(source_row):
                # Composite field - sum multiple rows
                composite_rows = [int(r.strip()) for r in str(source_row).split('+')]
                composite_q2_value = 0
                
                for comp_row in composite_rows:
                    comp_value = source_sheet.cell(comp_row, 93).value or 0
                    composite_q2_value += comp_value
                
                actual_q2_value = composite_q2_value
                source_location = self.create_source_location_string(source_sheet_name, source_row)
                result['Status'] = 'COMPOSITE_POPULATED'
            else:
                # Single source row
                source_row_num = int(source_row)
                actual_q2_value = source_sheet.cell(source_row_num, 93).value
                source_location = self.create_source_location_string(source_sheet_name, source_row)
                result['Status'] = 'POPULATED'
            
            result['Actual_Q2_Value'] = actual_q2_value
            result['Source_Location'] = source_location
            
            # Populate destination if we have data
            if actual_q2_value is not None:
                # Populate target column with Q2 value
                dest_sheet.cell(dest_row, self.target_column).value = actual_q2_value
                self.stats['values_populated'] += 1
                
                # Add source tracking to next column
                dest_sheet.cell(dest_row, self.source_tracking_column).value = source_location
                self.stats['source_tracking_added'] += 1
                
                # Update stats
                if source_sheet_name not in self.stats['sheet_stats']:
                    self.stats['sheet_stats'][source_sheet_name] = 0
                self.stats['sheet_stats'][source_sheet_name] += 1
                
                if match_method not in self.stats['method_stats']:
                    self.stats['method_stats'][match_method] = 0
                self.stats['method_stats'][match_method] += 1
                
            else:
                result['Status'] = 'NO_SOURCE_DATA'
                self.stats['errors'].append(f"Row {dest_row}: No source data available")
            
        except Exception as e:
            result['Status'] = 'PROCESSING_ERROR'
            result['Error'] = str(e)
            self.stats['errors'].append(f"Row {dest_row}: {str(e)}")
        
        return result
    
    def populate_fields(self, mappings: List[Dict]) -> List[Dict]:
        """Populate destination fields using the mappings."""
        
        print(f"\n=== POPULATING FIELDS ===")
        print(f"Source file: {self.source_file}")
        print(f"Destination file: {self.destination_file}")
        print(f"Target column: {self.target_column}")
        print(f"Source tracking column: {self.source_tracking_column}")
        
        # Load workbooks
        source_wb = openpyxl.load_workbook(self.source_file, data_only=True)
        dest_wb = openpyxl.load_workbook(self.destination_file, data_only=False)
        dest_sheet = dest_wb['Reported']  # Assume 'Reported' sheet for now
        
        population_results = []
        
        print(f"\nProcessing {len(mappings)} mappings...")
        
        for i, mapping in enumerate(mappings, 1):
            result = self.process_single_mapping(mapping, source_wb, dest_sheet)
            population_results.append(result)
            
            # Progress reporting
            if i % 10 == 0 or i <= 5 or i == len(mappings):
                status_icon = "✅" if result['Status'] in ['POPULATED', 'COMPOSITE_POPULATED'] else "❌"
                print(f"[{i}/{len(mappings)}] {status_icon} Row {result['Dest_Row']}: {result['Dest_Field_Name']}")
                if result['Status'] in ['POPULATED', 'COMPOSITE_POPULATED']:
                    print(f"    Value: {result['Actual_Q2_Value']} | Method: {result['Match_Method']}")
        
        # Save populated file
        print(f"\nSaving populated file to: {self.output_file}")
        dest_wb.save(self.output_file)
        
        source_wb.close()
        dest_wb.close()
        
        self.stats['mappings_processed'] = len(mappings)
        
        return population_results
    
    def generate_audit_trail(self, population_results: List[Dict]):
        """Generate audit trail CSV file."""
        
        print(f"\nGenerating audit trail: {self.audit_file}")
        
        with open(self.audit_file, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = [
                'Dest_Row', 'Dest_Field_Name', 'Source_Sheet', 'Source_Row',
                'Source_Field_Name', 'Expected_Q2_Value', 'Actual_Q2_Value',
                'Source_Location', 'Match_Method', 'Status'
            ]
            
            # Add Error field if there are any errors
            if any('Error' in result for result in population_results):
                fieldnames.append('Error')
            
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            
            for result in population_results:
                writer.writerow(result)
    
    def print_summary(self):
        """Print comprehensive summary of the mapping operation."""
        
        print(f"\n" + "="*80)
        print("PARAMETERIZED FIELD MAPPING RESULTS")
        print("="*80)
        
        print(f"INPUT PARAMETERS:")
        print(f"  Source file: {self.source_file}")
        print(f"  Destination file: {self.destination_file}")
        print(f"  Mapping file: {self.mapping_file}")
        print(f"  Target column: {self.target_column}")
        print(f"  Source tracking column: {self.source_tracking_column}")
        
        print(f"\nEXECUTION RESULTS:")
        print(f"  Total mappings processed: {self.stats['mappings_processed']}")
        print(f"  Values successfully populated: {self.stats['values_populated']}")
        print(f"  Source tracking entries added: {self.stats['source_tracking_added']}")
        
        if self.stats['mappings_processed'] > 0:
            success_rate = self.stats['values_populated'] / self.stats['mappings_processed'] * 100
            print(f"  Success rate: {self.stats['values_populated']}/{self.stats['mappings_processed']} = {success_rate:.1f}%")
        
        if self.stats['sheet_stats']:
            print(f"\nPopulation by source sheet:")
            for sheet, count in sorted(self.stats['sheet_stats'].items()):
                print(f"  {sheet}: {count} values")
        
        if self.stats['method_stats']:
            print(f"\nPopulation by method:")
            for method, count in sorted(self.stats['method_stats'].items()):
                print(f"  {method}: {count} values")
        
        if self.stats['errors']:
            print(f"\nErrors encountered: {len(self.stats['errors'])}")
            for error in self.stats['errors'][:5]:  # Show first 5
                print(f"  {error}")
            if len(self.stats['errors']) > 5:
                print(f"  ... and {len(self.stats['errors']) - 5} more errors")
        
        print(f"\nOUTPUT FILES:")
        print(f"  📁 Populated Excel: {self.output_file}")
        print(f"  📁 Audit trail: {self.audit_file}")
    
    def run(self) -> bool:
        """Run the complete field mapping and population process."""
        
        print("="*80)
        print("PARAMETERIZED FIELD MAPPER")
        print("="*80)
        
        try:
            # Validate files
            valid, validation_errors = self.validate_files()
            if not valid:
                print("❌ VALIDATION FAILED:")
                for error in validation_errors:
                    print(f"  {error}")
                return False
            
            # Load mappings
            mappings = self.load_mapping_file()
            if not mappings:
                print("❌ No mappings loaded")
                return False
            
            # Populate fields
            population_results = self.populate_fields(mappings)
            
            # Generate audit trail
            self.generate_audit_trail(population_results)
            
            # Print summary
            self.print_summary()
            
            print(f"\n🎉 PARAMETERIZED FIELD MAPPING COMPLETE!")
            
            return True
            
        except Exception as e:
            print(f"❌ ERROR: {e}")
            import traceback
            traceback.print_exc()
            return False


def parse_arguments():
    """Parse command line arguments."""
    
    parser = argparse.ArgumentParser(
        description='Parameterized Field Mapper - Map and populate fields between Excel files',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Basic usage
  python parameterized_field_mapper.py --source source.xlsx --destination dest.xlsx 
                                       --mapping mappings.csv --column 71
  
  # With custom output files
  python parameterized_field_mapper.py --source source.xlsx --destination dest.xlsx 
                                       --mapping mappings.csv --column 71
                                       --output populated_dest.xlsx --audit audit.csv
        """
    )
    
    parser.add_argument('--source', '-s', required=True,
                       help='Path to source Excel file')
    parser.add_argument('--destination', '-d', required=True,
                       help='Path to destination Excel file')
    parser.add_argument('--mapping', '-m', required=True,
                       help='Path to CSV mapping file')
    parser.add_argument('--column', '-c', type=int, required=True,
                       help='Target column number to populate (1-based)')
    parser.add_argument('--output', '-o', 
                       help='Output file path (default: populated_[destination])')
    parser.add_argument('--audit', '-a',
                       help='Audit trail file path (default: audit_trail_[destination].csv)')
    
    return parser.parse_args()


def main():
    """Main entry point for parameterized field mapper."""
    
    # Parse command line arguments
    args = parse_arguments()
    
    # Create and run mapper
    mapper = ParameterizedFieldMapper(
        source_file=args.source,
        destination_file=args.destination,
        mapping_file=args.mapping,
        target_column=args.column,
        output_file=args.output,
        audit_file=args.audit
    )
    
    success = mapper.run()
    
    if success:
        print(f"\n✅ SUCCESS: Field mapping completed successfully!")
    else:
        print(f"\n❌ FAILURE: Field mapping failed!")
        exit(1)


if __name__ == "__main__":
    main()
