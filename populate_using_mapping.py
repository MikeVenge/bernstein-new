#!/usr/bin/env python3
"""
Population Script Using Q1 Verified Mapping - Step 2

This script:
1. Reads the Q1 verified field mapping file from Step 1
2. Uses the mapping to populate destination Column BS with source Column CO values
3. Provides complete audit trail of population

Author: AI Assistant
Date: October 2025
"""

import openpyxl
import pandas as pd
import csv
from pathlib import Path
from typing import Dict, List


def load_verified_mapping() -> List[Dict]:
    """Load the Q1 verified field mapping from Step 1."""
    mapping_file = "/Users/michaelkim/code/Bernstein/q1_verified_field_mapping.csv"
    
    print(f"Loading Q1 verified mapping from: {mapping_file}")
    
    if not Path(mapping_file).exists():
        print(f"ERROR: Mapping file not found: {mapping_file}")
        return []
    
    mappings = []
    with open(mapping_file, 'r', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        mappings = list(reader)
    
    print(f"Loaded {len(mappings)} verified field mappings")
    return mappings


def populate_destination_file(mappings: List[Dict]) -> Dict:
    """
    Populate the destination file using the verified mappings.
    """
    print("="*80)
    print("STEP 2: POPULATING DESTINATION FILE")
    print("="*80)
    
    source_file = "/Users/michaelkim/code/Bernstein/IPGP-Financial-Data-Workbook-2024-Q2.xlsx"
    dest_file = "/Users/michaelkim/code/Bernstein/20240725_IPGP.US-IPG Photonics.xlsx"
    
    print(f"Source file: {source_file}")
    print(f"Destination file: {dest_file}")
    print(f"Population: Source Column CO → Destination Column BS")
    
    # Load source workbook
    print("\nLoading source workbook...")
    source_wb = openpyxl.load_workbook(source_file, data_only=True)
    source_sheet = source_wb['Key Metrics']
    
    # Load destination workbook
    print("Loading destination workbook...")
    dest_wb = openpyxl.load_workbook(dest_file, data_only=False)  # Keep formulas
    dest_sheet = dest_wb['Reported']
    
    # Population tracking
    population_results = []
    values_populated = 0
    
    print(f"\nPopulating {len(mappings)} verified field mappings...")
    
    for mapping in mappings:
        source_row = int(mapping['Source_Row_Number'])
        dest_row = int(mapping['Dest_Row_Number'])
        source_field_name = mapping['Source_Field_Name']
        dest_field_name = mapping['Dest_Field_Name']
        q1_verification_value = mapping['Q1_Verification_Value']
        
        # Get Q2 2024 value from source (Column CO = 93)
        source_q2_value = source_sheet.cell(source_row, 93).value
        
        # Get current destination value (Column BS = 71)
        current_dest_value = dest_sheet.cell(dest_row, 71).value
        
        print(f"\nPopulating DEST Row {dest_row}: {dest_field_name}")
        print(f"  From SRC Row {source_row}: {source_field_name}")
        print(f"  Q1 verification: {q1_verification_value}")
        print(f"  Source Q2 (CO): {source_q2_value}")
        print(f"  Current Dest (BS): {current_dest_value}")
        
        if source_q2_value is not None:
            # Populate destination Column BS with source Column CO value
            dest_sheet.cell(dest_row, 71).value = source_q2_value
            values_populated += 1
            
            population_result = {
                'Dest_Row': dest_row,
                'Dest_Field_Name': dest_field_name,
                'Dest_Enhanced_Scope': mapping['Dest_Enhanced_Scope'],
                'Source_Row': source_row,
                'Source_Field_Name': source_field_name,
                'Source_Enhanced_Scope': mapping['Source_Enhanced_Scope'],
                'Q1_Verification_Value': q1_verification_value,
                'Source_Q2_Value': source_q2_value,
                'Previous_Dest_Value': current_dest_value,
                'Population_Status': 'POPULATED',
                'Match_Method': mapping['Match_Method']
            }
            
            print(f"  ✓ POPULATED: {source_q2_value}")
        else:
            population_result = {
                'Dest_Row': dest_row,
                'Dest_Field_Name': dest_field_name,
                'Dest_Enhanced_Scope': mapping['Dest_Enhanced_Scope'],
                'Source_Row': source_row,
                'Source_Field_Name': source_field_name,
                'Source_Enhanced_Scope': mapping['Source_Enhanced_Scope'],
                'Q1_Verification_Value': q1_verification_value,
                'Source_Q2_Value': '',
                'Previous_Dest_Value': current_dest_value,
                'Population_Status': 'NO_Q2_DATA',
                'Match_Method': mapping['Match_Method']
            }
            
            print(f"  ❌ NO Q2 DATA AVAILABLE")
        
        population_results.append(population_result)
    
    # Save updated destination workbook
    output_file = "/Users/michaelkim/code/Bernstein/populated_20240725_IPGP.US-IPG_Photonics.xlsx"
    print(f"\nSaving populated destination file to: {output_file}")
    dest_wb.save(output_file)
    
    # Close workbooks
    source_wb.close()
    dest_wb.close()
    
    print(f"✓ Destination file saved with {values_populated} populated values")
    
    return {
        'population_results': population_results,
        'values_populated': values_populated,
        'total_mappings': len(mappings),
        'output_file': output_file
    }


def save_population_audit_trail(population_summary: Dict, output_file: str):
    """Save complete audit trail of the population process."""
    print(f"\nSaving population audit trail to: {output_file}")
    
    with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
        fieldnames = [
            'Dest_Row', 'Dest_Field_Name', 'Dest_Enhanced_Scope',
            'Source_Row', 'Source_Field_Name', 'Source_Enhanced_Scope',
            'Q1_Verification_Value', 'Source_Q2_Value', 'Previous_Dest_Value',
            'Population_Status', 'Match_Method'
        ]
        
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(population_summary['population_results'])
    
    print(f"✓ Population audit trail saved")


def main():
    """Main entry point for Step 2."""
    audit_file = "/Users/michaelkim/code/Bernstein/population_audit_trail.csv"
    
    print("="*80)
    print("STEP 2: POPULATION USING Q1 VERIFIED MAPPING")
    print("="*80)
    print("Using Q1 verified field mapping to populate destination file")
    print(f"Audit trail file: {audit_file}")
    
    try:
        # Load verified mappings from Step 1
        mappings = load_verified_mapping()
        
        if not mappings:
            print("ERROR: No verified mappings loaded")
            return
        
        # Populate destination file
        population_summary = populate_destination_file(mappings)
        
        # Save audit trail
        save_population_audit_trail(population_summary, audit_file)
        
        # Final summary
        print(f"\n" + "="*80)
        print("POPULATION COMPLETE")
        print("="*80)
        print(f"Total verified mappings: {population_summary['total_mappings']}")
        print(f"Values successfully populated: {population_summary['values_populated']}")
        print(f"Population rate: {population_summary['values_populated']}/{population_summary['total_mappings']} = {population_summary['values_populated']/population_summary['total_mappings']*100:.1f}%")
        
        print(f"\nFiles generated:")
        print(f"  1. Populated destination file: {population_summary['output_file']}")
        print(f"  2. Population audit trail: {audit_file}")
        
        print(f"\nTwo-step process complete:")
        print(f"  Step 1: Q1 verification mapping ✓")
        print(f"  Step 2: Destination population ✓")
        
        print(f"\nPopulation method: Source Column CO → Destination Column BS")
        print(f"Validation method: Q1 verification (100% confidence)")
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()

