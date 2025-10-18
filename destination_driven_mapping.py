#!/usr/bin/env python3
"""
Destination-Driven Field Mapping Strategy

Uses the destination file (Reported tab) as the master field list, then finds
matches from source files to populate those destination fields.

This approach recognizes that:
1. Destination file contains ALL fields that need to be populated
2. Source files are incomplete - each contains only a subset
3. Multiple source files may be needed to complete the destination
4. Destination file defines the requirements

Author: AI Assistant
Date: October 2025
"""

import openpyxl
import csv
import pandas as pd
from pathlib import Path
import re
from typing import Dict, List, Tuple, Optional
from difflib import SequenceMatcher


def get_destination_master_fields(target_file: str) -> List[Dict]:
    """
    Get the master field list from the destination file (Reported tab).
    This becomes our superset of fields that need to be populated.
    """
    print("="*80)
    print("DESTINATION-DRIVEN MAPPING STRATEGY")
    print("="*80)
    print("Using destination file as master field list...")
    
    wb = openpyxl.load_workbook(target_file, data_only=True)
    
    if 'Reported' not in wb.sheetnames:
        print(f"ERROR: Reported tab not found. Available sheets: {wb.sheetnames}")
        wb.close()
        return []
    
    sheet = wb['Reported']
    print(f"Reported tab has {sheet.max_row} rows and {sheet.max_column} columns")
    
    # Get destination column headers (quarters/periods)
    destination_columns = {}
    header_row = 5
    for col_idx in range(2, min(sheet.max_column + 1, 80)):
        header_val = sheet.cell(header_row, col_idx).value
        if header_val and str(header_val).strip():
            destination_columns[col_idx] = {
                'header': str(header_val).strip(),
                'clean_header': clean_quarter_header(str(header_val))
            }
    
    print(f"Found {len(destination_columns)} destination columns")
    
    # Process all destination fields
    destination_fields = []
    current_major_section = None
    current_section = None
    current_subsection = None
    
    for row_idx in range(1, sheet.max_row + 1):
        first_col = sheet.cell(row_idx, 1).value
        if not first_col:
            continue
            
        first_col_str = str(first_col).strip()
        
        # Skip header rows
        if (row_idx <= 6 or 
            not first_col_str or
            first_col_str.lower().startswith('ipg photonics') or
            first_col_str.lower() in ['reported', 'quarterly']):
            continue
        
        # Classify and update context
        row_type = classify_destination_row(first_col_str, row_idx)
        
        if row_type == 'major_section':
            current_major_section = clean_field_name(first_col_str)
            current_section = None
            current_subsection = None
            
        elif row_type == 'section_header':
            current_section = clean_field_name(first_col_str)
            current_subsection = None
            
        elif row_type == 'subsection_header':
            current_subsection = clean_field_name(first_col_str.rstrip(':'))
            
        elif row_type == 'data_field':
            # Get current values in destination columns
            destination_values = {}
            for col_idx, col_info in destination_columns.items():
                value = sheet.cell(row_idx, col_idx).value
                destination_values[col_info['clean_header']] = value
            
            # Create destination field record
            field_name = clean_field_name(first_col_str)
            enhanced_scoped_name = build_destination_scoped_name(
                current_major_section, current_section, current_subsection, field_name
            )
            
            destination_field = {
                'row_number': row_idx,
                'original_field_name': first_col_str,
                'cleaned_field_name': field_name,
                'major_section_context': current_major_section or '',
                'section_context': current_section or '',
                'subsection_context': current_subsection or '',
                'enhanced_scoped_name': enhanced_scoped_name,
                'destination_values': destination_values,
                'needs_population': not any(v is not None and str(v).strip() for v in destination_values.values()),
                'row_type': row_type
            }
            
            destination_fields.append(destination_field)
    
    wb.close()
    print(f"Identified {len(destination_fields)} destination fields that need population")
    
    # Count fields that need population
    needs_pop = sum(1 for f in destination_fields if f['needs_population'])
    print(f"Fields needing population: {needs_pop}")
    print(f"Fields already populated: {len(destination_fields) - needs_pop}")
    
    return destination_fields


def find_source_matches_for_destination(destination_fields: List[Dict], 
                                       source_files: List[str]) -> List[Dict]:
    """
    For each destination field, try to find matches in source files.
    """
    print(f"\nSearching for matches across {len(source_files)} source files...")
    
    # Load all source files and their data
    source_data = {}
    for source_file in source_files:
        source_data[source_file] = load_source_file_data(source_file)
    
    # For each destination field, find best source match
    matched_fields = []
    
    for dest_field in destination_fields:
        dest_name = dest_field['original_field_name']
        dest_scoped = dest_field['enhanced_scoped_name']
        
        print(f"\nFinding source for: {dest_name}")
        
        best_match = None
        best_score = 0.0
        best_source_file = None
        
        # Search across all source files
        for source_file, source_fields in source_data.items():
            for source_field in source_fields:
                # Calculate similarity score
                score = calculate_field_similarity(dest_field, source_field)
                
                if score > best_score:
                    best_score = score
                    best_match = source_field
                    best_source_file = source_file
        
        # Create matched field record
        if best_match and best_score > 0.5:  # Minimum threshold
            matched_field = {
                'destination_field': dest_field,
                'source_match': best_match,
                'source_file': best_source_file,
                'similarity_score': best_score,
                'match_quality': get_match_quality(best_score),
                'can_populate': best_match is not None
            }
            
            print(f"  → Found in {Path(best_source_file).name}: {best_match['original_field_name']} (Score: {best_score:.3f})")
        else:
            matched_field = {
                'destination_field': dest_field,
                'source_match': None,
                'source_file': None,
                'similarity_score': 0.0,
                'match_quality': 'No Match',
                'can_populate': False
            }
            print(f"  → No match found")
        
        matched_fields.append(matched_field)
    
    return matched_fields


def load_source_file_data(source_file: str) -> List[Dict]:
    """
    Load field data from a source file.
    """
    source_fields = []
    
    try:
        wb = openpyxl.load_workbook(source_file, data_only=True)
        
        # Try Key Metrics sheet first
        if 'Key Metrics' in wb.sheetnames:
            sheet = wb['Key Metrics']
            source_fields.extend(extract_fields_from_sheet(sheet, 'Key Metrics', source_file))
        
        # Try other common sheet names
        for sheet_name in ['Income Statement', 'Balance Sheet', 'Cash Flows']:
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                source_fields.extend(extract_fields_from_sheet(sheet, sheet_name, source_file))
        
        wb.close()
        
    except Exception as e:
        print(f"Error loading {source_file}: {e}")
    
    return source_fields


def extract_fields_from_sheet(sheet, sheet_name: str, source_file: str) -> List[Dict]:
    """
    Extract field information from a sheet.
    """
    fields = []
    
    # Get data columns (look for Q1 2024 and Q2 2024)
    data_columns = {}
    header_row = 4 if sheet_name == 'Key Metrics' else 5
    
    for col_idx in range(2, min(sheet.max_column + 1, 100)):
        header_val = sheet.cell(header_row, col_idx).value
        if header_val:
            header_str = str(header_val).strip()
            # Look for 2024 Q1 and Q2 specifically
            if '2024' in header_str:
                data_columns[col_idx] = {
                    'header': header_str,
                    'clean_header': clean_date_header(header_str)
                }
    
    # Extract fields
    for row_idx in range(1, min(sheet.max_row + 1, 200)):
        first_col = sheet.cell(row_idx, 1).value
        if not first_col:
            continue
            
        first_col_str = str(first_col).strip()
        
        # Skip headers
        if (row_idx <= header_row or 
            first_col_str.lower().startswith('ipg photonics') or
            len(first_col_str) < 2):
            continue
        
        # Get values from data columns
        field_values = {}
        for col_idx, col_info in data_columns.items():
            value = sheet.cell(row_idx, col_idx).value
            field_values[col_info['clean_header']] = value
        
        # Create field record
        field = {
            'original_field_name': first_col_str,
            'cleaned_field_name': clean_field_name(first_col_str),
            'sheet_name': sheet_name,
            'source_file': source_file,
            'row_number': row_idx,
            'values': field_values,
            'has_data': any(v is not None and str(v).strip() for v in field_values.values())
        }
        
        fields.append(field)
    
    return fields


def calculate_field_similarity(dest_field: Dict, source_field: Dict) -> float:
    """
    Calculate similarity between destination and source fields.
    """
    dest_name = dest_field['original_field_name'].lower()
    source_name = source_field['original_field_name'].lower()
    
    # Basic string similarity
    base_score = SequenceMatcher(None, dest_name, source_name).ratio()
    
    # Bonus for exact matches
    if dest_name == source_name:
        base_score += 0.5
    
    # Bonus for geographic terms
    geo_terms = ['north america', 'germany', 'china', 'japan', 'europe', 'asia']
    if any(term in dest_name and term in source_name for term in geo_terms):
        base_score += 0.3
    
    # Bonus for financial terms
    fin_terms = ['revenue', 'total', 'income', 'sales', 'assets', 'cash']
    if any(term in dest_name and term in source_name for term in fin_terms):
        base_score += 0.2
    
    # Penalty if source has no data
    if not source_field.get('has_data', False):
        base_score *= 0.8
    
    return min(base_score, 1.0)


def get_match_quality(score: float) -> str:
    """Get match quality description."""
    if score >= 0.9:
        return 'Excellent'
    elif score >= 0.7:
        return 'Good'
    elif score >= 0.5:
        return 'Fair'
    else:
        return 'Poor'


def classify_destination_row(first_col: str, row_idx: int) -> str:
    """Classify destination row types."""
    first_col_lower = first_col.lower()
    first_col_upper = first_col.upper()
    
    # Major sections
    if (first_col == first_col_upper and len(first_col) > 10) or any(pattern in first_col_lower for pattern in [
        'financial statements', 'segment information', 'income statement', 'balance sheet', 'cash flow'
    ]):
        return 'major_section'
    
    # Section headers
    section_patterns = [
        r'segment breakdown', r'end market breakdown', r'revenue by', r'by region',
        r'product breakdown', r'application breakdown', r'geographic breakdown',
        r'current assets', r'current liabilities', r'equity',
        r'operating activities', r'investing activities', r'financing activities'
    ]
    
    for pattern in section_patterns:
        if re.search(pattern, first_col_lower):
            return 'section_header'
    
    # Subsection headers
    if first_col.endswith(':') or first_col.endswith(', of which'):
        return 'subsection_header'
    
    # Data fields
    if (len(first_col) > 1 and 
        not first_col_lower.startswith('ipg photonics') and
        first_col_lower not in ['reported', 'quarterly']):
        return 'data_field'
    
    return 'other'


def build_destination_scoped_name(major_section: str, section: str, subsection: str, field: str) -> str:
    """Build scoped name for destination fields."""
    components = ['Destination']
    
    if major_section:
        components.append(major_section)
    if section:
        components.append(section)
    if subsection:
        components.append(subsection)
    if field:
        components.append(field)
    
    return '.'.join(components)


def clean_field_name(name: str) -> str:
    """Clean field names."""
    if not name:
        return ""
    
    cleaned = re.sub(r'\s+', '_', name.strip())
    cleaned = re.sub(r'[^\w\s-]', '', cleaned)
    cleaned = re.sub(r'[-_]+', '_', cleaned)
    cleaned = cleaned.strip('_')
    
    if cleaned:
        words = cleaned.split('_')
        cleaned = '_'.join(word.capitalize() for word in words)
    
    return cleaned


def clean_quarter_header(header: str) -> str:
    """Clean quarter headers."""
    header_str = str(header).strip()
    
    # Handle patterns like "1Q24", "2Q24"
    if re.match(r'\d[QH]\d{2}', header_str):
        quarter = header_str[0]
        year_short = header_str[2:]
        year_full = f"20{year_short}" if int(year_short) < 50 else f"19{year_short}"
        return f"{year_full}_Q{quarter}"
    
    return header_str.replace('-', '_').replace(' ', '_')


def clean_date_header(header) -> str:
    """Clean date headers."""
    if hasattr(header, 'strftime'):
        year = header.year
        month = header.month
        quarter = (month - 1) // 3 + 1
        return f"{year}_Q{quarter}"
    
    header_str = str(header).strip()
    if re.match(r'\d{4}-\d{2}-\d{2}', header_str):
        year = header_str[:4]
        month = header_str[5:7]
        quarter = (int(month) - 1) // 3 + 1
        return f"{year}_Q{quarter}"
    
    return header_str.replace('-', '_').replace(' ', '_')


def save_destination_driven_results(matched_fields: List[Dict], output_file: str):
    """Save destination-driven mapping results."""
    with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
        fieldnames = [
            'Dest_Row_Number',
            'Dest_Field_Name',
            'Dest_Enhanced_Scope',
            'Dest_Major_Section',
            'Dest_Section',
            'Dest_Q1_2024_Value',
            'Dest_Needs_Population',
            'Source_File',
            'Source_Sheet',
            'Source_Row_Number',
            'Source_Field_Name',
            'Source_Q1_2024_Value',
            'Source_Q2_2024_Value',
            'Similarity_Score',
            'Match_Quality',
            'Can_Populate'
        ]
        
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        
        for match in matched_fields:
            dest = match['destination_field']
            source = match['source_match']
            
            # Get source values
            source_q1 = ''
            source_q2 = ''
            source_sheet = ''
            source_row = ''
            source_field_name = ''
            
            if source:
                source_values = source.get('values', {})
                source_q1 = source_values.get('2024_Q1', '')
                source_q2 = source_values.get('2024_Q2', '')
                source_sheet = source.get('sheet_name', '')
                source_row = source.get('row_number', '')
                source_field_name = source.get('original_field_name', '')
            
            # Get destination Q1 value (Column BR = 70)
            dest_values = dest.get('destination_values', {})
            dest_q1 = dest_values.get('1Q24', '') or dest_values.get('2024_Q1', '')
            
            # If not found by name, look for Q1 2024 pattern
            if not dest_q1:
                for col_name, col_val in dest_values.items():
                    if 'Q1' in str(col_name) and '24' in str(col_name):
                        dest_q1 = col_val
                        break
            
            writer.writerow({
                'Dest_Row_Number': dest['row_number'],
                'Dest_Field_Name': dest['original_field_name'],
                'Dest_Enhanced_Scope': dest['enhanced_scoped_name'],
                'Dest_Major_Section': dest['major_section_context'],
                'Dest_Section': dest['section_context'],
                'Dest_Q1_2024_Value': dest_q1 if dest_q1 is not None else '',
                'Dest_Needs_Population': 'Yes' if dest['needs_population'] else 'No',
                'Source_File': Path(match['source_file']).name if match['source_file'] else '',
                'Source_Sheet': source_sheet,
                'Source_Row_Number': source_row,
                'Source_Field_Name': source_field_name,
                'Source_Q1_2024_Value': source_q1 if source_q1 is not None else '',
                'Source_Q2_2024_Value': source_q2 if source_q2 is not None else '',
                'Similarity_Score': f"{match['similarity_score']:.3f}",
                'Match_Quality': match['match_quality'],
                'Can_Populate': 'Yes' if match['can_populate'] else 'No'
            })


def main():
    """Main entry point."""
    target_file = "/Users/michaelkim/code/Bernstein/20240725_IPGP.US-IPG Photonics.xlsx"
    source_files = [
        "/Users/michaelkim/code/Bernstein/IPGP-Financial-Data-Workbook-2024-Q2.xlsx"
        # Add more source files here as needed
    ]
    output_file = "/Users/michaelkim/code/Bernstein/destination_driven_mapping_results.csv"
    
    print("DESTINATION-DRIVEN FIELD MAPPING")
    print("Using destination file as the master field list")
    print(f"Target file: {target_file}")
    print(f"Source files: {[Path(f).name for f in source_files]}")
    print(f"Output file: {output_file}")
    
    try:
        # Step 1: Get master field list from destination
        destination_fields = get_destination_master_fields(target_file)
        
        # Step 2: Find source matches for each destination field
        matched_fields = find_source_matches_for_destination(destination_fields, source_files)
        
        # Step 3: Save results
        save_destination_driven_results(matched_fields, output_file)
        
        # Summary
        can_populate = sum(1 for m in matched_fields if m['can_populate'])
        needs_population = sum(1 for m in matched_fields if m['destination_field']['needs_population'])
        
        print(f"\n" + "="*80)
        print("DESTINATION-DRIVEN MAPPING COMPLETE")
        print("="*80)
        print(f"Total destination fields: {len(matched_fields)}")
        print(f"Fields needing population: {needs_population}")
        print(f"Fields with source matches: {can_populate}")
        print(f"Coverage: {can_populate/needs_population*100:.1f}% of needed fields have sources")
        print(f"\nResults saved to: {output_file}")
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
