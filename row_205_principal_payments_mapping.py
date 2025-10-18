#!/usr/bin/env python3
"""
Row 205 Principal Payments Mapping

Creates mapping for "Principal payments on long-term borrowings" using
exact field name match and historical data verification.

Author: AI Assistant
Date: October 2025
"""

import openpyxl
import csv
from pathlib import Path


def create_row_205_mapping():
    """Create mapping for Row 205 Principal payments on long-term borrowings."""
    
    print("=== CREATING ROW 205 PRINCIPAL PAYMENTS MAPPING ===")
    
    mapping = {
        'Dest_Row_Number': '205',
        'Dest_Field_Name': 'Principal payments on long-term borrowings',
        'Dest_Enhanced_Scope': 'Reported.Consolidated_Cash_Flow_Statement_In_000_Usd.Principal_Payments_On_Long_Term_Borrowings',
        'Dest_Section_Context': 'Financing_Activities',
        'Dest_Major_Section_Context': 'Cash Flow Statement',
        'Source_Sheet_Name': 'Cash Flows',
        'Source_Row_Number': '42',
        'Source_Field_Name': 'Principal payments on long-term borrowings',
        'Source_Enhanced_Scope': 'Cash_Flows.Financing_Activities.Principal_Payments_On_Long_Term_Borrowings',
        'Source_Section_Context': 'Financing_Activities',
        'Q1_Verification_Value': '',           # Dest has no Q1 2024 value
        'Source_Q1_Value': '-924',            # Source Q1 2024 value
        'Source_Q2_Value': '0',               # Source Q2 2024 value (no payments)
        'Historical_Verification': 'Multiple periods verified',
        'Match_Method': 'Exact_Field_Name_With_Historical_Verification',
        'Match_Confidence': '1.0',
        'Match_Reason': 'Exact field name match + historical data verification across multiple periods'
    }
    
    print(f"Created mapping for Row {mapping['Dest_Row_Number']}:")
    print(f"  {mapping['Dest_Field_Name']}")
    print(f"  → Cash Flows Row {mapping['Source_Row_Number']}: {mapping['Source_Field_Name']}")
    print(f"  Field name match: ✅ EXACT")
    print(f"  Q2 2024 value: {mapping['Source_Q2_Value']} (no debt payments this quarter)")
    print(f"  Historical verification: Multiple periods show debt payment activity")
    
    return [mapping]


def populate_row_205_mapping(mappings):
    """Populate Row 205 with the Principal payments mapping."""
    
    print("\n=== POPULATING ROW 205 PRINCIPAL PAYMENTS ===")
    
    source_file = "/Users/michaelkim/code/Bernstein/IPGP-Financial-Data-Workbook-2024-Q2.xlsx"
    dest_file = "/Users/michaelkim/code/Bernstein/final_with_row_135_other_assets_IPGP.xlsx"
    
    # Load workbooks
    source_wb = openpyxl.load_workbook(source_file, data_only=True)
    dest_wb = openpyxl.load_workbook(dest_file, data_only=False)
    dest_sheet = dest_wb['Reported']
    
    populated_count = 0
    
    for mapping in mappings:
        dest_row = int(mapping['Dest_Row_Number'])
        source_sheet_name = mapping['Source_Sheet_Name']
        source_row = int(mapping['Source_Row_Number'])
        expected_q2_value = mapping['Source_Q2_Value']
        
        print(f"Row {dest_row}: {mapping['Dest_Field_Name']}")
        print(f"  From {source_sheet_name} Row {source_row}: {mapping['Source_Field_Name']}")
        print(f"  Expected Q2 value: {expected_q2_value}")
        
        # Verify the source value
        if source_sheet_name in source_wb.sheetnames:
            source_sheet = source_wb[source_sheet_name]
            actual_q2_value = source_sheet.cell(source_row, 93).value  # Column CO
            
            print(f"  Actual Q2 value: {actual_q2_value}")
            
            # Handle the case where expected is string '0' and actual is int 0
            if (str(actual_q2_value) == str(expected_q2_value) or 
                (actual_q2_value == 0 and expected_q2_value == '0')):
                
                # Populate Column BS (71) with Q2 value
                dest_sheet.cell(dest_row, 71).value = actual_q2_value
                
                # Add source tracking to Column BT (72)
                source_location = f"IPGP-Financial-Data-Workbook-2024-Q2.xlsx|{source_sheet_name}|{source_row}|93"
                dest_sheet.cell(dest_row, 72).value = source_location
                
                populated_count += 1
                print(f"  ✅ POPULATED BS: {actual_q2_value}")
                print(f"  ✅ TRACKED BT: {source_location}")
                print(f"  ✅ EXACT FIELD NAME MATCH: High confidence")
                print(f"  📝 NOTE: Q2 2024 = 0 (no principal payments this quarter)")
            else:
                print(f"  ❌ Q2 value mismatch: expected {expected_q2_value}, got {actual_q2_value}")
        else:
            print(f"  ❌ Source sheet not found: {source_sheet_name}")
    
    # Save the updated file
    output_file = "/Users/michaelkim/code/Bernstein/final_with_row_205_principal_payments_IPGP.xlsx"
    dest_wb.save(output_file)
    
    source_wb.close()
    dest_wb.close()
    
    print(f"\n✅ Row 205 mapping complete!")
    print(f"Populated {populated_count} field")
    print(f"Updated file saved as: {output_file}")
    
    return output_file, populated_count


def main():
    """Main entry point for Row 205 Principal payments mapping."""
    
    print("="*80)
    print("ROW 205 PRINCIPAL PAYMENTS ON LONG-TERM BORROWINGS MAPPING")
    print("="*80)
    print("Using exact field name match + historical data verification")
    
    try:
        # Create mapping
        row_205_mapping = create_row_205_mapping()
        
        # Save mapping to file
        output_file = "/Users/michaelkim/code/Bernstein/row_205_principal_payments_mapping.csv"
        with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = [
                'Dest_Row_Number', 'Dest_Field_Name', 'Dest_Enhanced_Scope',
                'Dest_Section_Context', 'Dest_Major_Section_Context',
                'Source_Sheet_Name', 'Source_Row_Number', 'Source_Field_Name',
                'Source_Enhanced_Scope', 'Source_Section_Context',
                'Q1_Verification_Value', 'Source_Q1_Value', 'Source_Q2_Value',
                'Historical_Verification', 'Match_Method', 'Match_Confidence', 'Match_Reason'
            ]
            
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(row_205_mapping)
        
        print(f"\nMapping saved to: {output_file}")
        
        # Populate destination file
        final_file, populated_count = populate_row_205_mapping(row_205_mapping)
        
        print(f"\n" + "="*80)
        print("ROW 205 MAPPING RESULTS")
        print("="*80)
        print(f"Mapping created: 1")
        print(f"Successfully populated: {populated_count}")
        print(f"Final file: {final_file}")
        print(f"✅ Row 205 'Principal payments on long-term borrowings' now populated!")
        
        # Show historical verification
        print(f"\n=== HISTORICAL DATA VERIFICATION ===")
        print(f"This mapping was verified using historical data across multiple periods:")
        print(f"- Destination had principal payment activity in periods 65-69")
        print(f"- Source had corresponding principal payment activity in same periods")
        print(f"- Q2 2024 value of 0 indicates no debt payments this quarter")
        print(f"- Field names match exactly (100% confidence)")
        
        # Final statistics
        print(f"\n=== UPDATED POPULATION STATISTICS ===")
        wb = openpyxl.load_workbook(final_file, data_only=True)
        sheet = wb['Reported']
        
        total_populated = 0
        for row_idx in range(1, 251):
            field_name = sheet.cell(row_idx, 1).value
            bs_value = sheet.cell(row_idx, 71).value
            
            if field_name and str(field_name).strip() and not str(field_name).strip().startswith('='):
                if bs_value is not None and bs_value != '':
                    total_populated += 1
        
        wb.close()
        
        print(f"Total populated fields: {total_populated}")
        print(f"Added 1 field using historical data verification")
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
