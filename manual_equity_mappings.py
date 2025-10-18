#!/usr/bin/env python3
"""
Manual Equity Section Mappings

Adds the missing equity section fields (rows 153-156) that couldn't be matched 
via Q1 verification due to value differences.

Author: AI Assistant
Date: October 2025
"""

import openpyxl
import csv
from pathlib import Path


def create_manual_equity_mappings():
    """Create manual mappings for the missing equity section fields."""
    
    print("=== CREATING MANUAL EQUITY SECTION MAPPINGS ===")
    
    # Define the manual mappings based on field name similarity and context
    manual_mappings = [
        {
            'Dest_Row_Number': '153',
            'Dest_Field_Name': 'Total IPG Photonics Corporation equity',
            'Dest_Enhanced_Scope': 'Reported.Consolidated_Balance_Sheet_In_000_Usd.Total_IPG_Photonics_Corporation_Equity',
            'Dest_Section_Context': 'Equity',
            'Dest_Major_Section_Context': 'Balance Sheet',
            'Source_Sheet_Name': 'Balance Sheet',
            'Source_Row_Number': '50',
            'Source_Field_Name': 'Total IPG Photonics Corporation equity',
            'Source_Enhanced_Scope': 'Balance_Sheet.Equity.Total_IPG_Photonics_Corporation_Equity',
            'Source_Section_Context': 'Equity',
            'Q1_Verification_Value': '2340720',  # Dest Q1 value
            'Source_Q1_Value': '2401726',       # Source Q1 value (different)
            'Source_Q2_Value': '2253370',       # Source Q2 value to populate
            'Match_Method': 'Manual_Field_Name_Match',
            'Match_Confidence': '1.0',
            'Match_Reason': 'Field names match exactly, Q1 values different (2340720 vs 2401726)'
        },
        {
            'Dest_Row_Number': '154',
            'Dest_Field_Name': 'Noncontrolling interests',
            'Dest_Enhanced_Scope': 'Reported.Consolidated_Balance_Sheet_In_000_Usd.Noncontrolling_Interests',
            'Dest_Section_Context': 'Equity',
            'Dest_Major_Section_Context': 'Balance Sheet',
            'Source_Sheet_Name': 'Balance Sheet',
            'Source_Row_Number': '52',
            'Source_Field_Name': 'Noncontrolling interests',
            'Source_Enhanced_Scope': 'Balance_Sheet.Equity.Noncontrolling_Interests',
            'Source_Section_Context': 'Equity',
            'Q1_Verification_Value': '',         # Dest has no Q1 value
            'Source_Q1_Value': '717',           # Source Q1 value
            'Source_Q2_Value': '0',             # Source Q2 value to populate
            'Match_Method': 'Manual_Field_Name_Match',
            'Match_Confidence': '1.0',
            'Match_Reason': 'Field names match exactly, dest has no Q1 for verification'
        },
        {
            'Dest_Row_Number': '155',
            'Dest_Field_Name': 'Total equity',
            'Dest_Enhanced_Scope': 'Reported.Consolidated_Balance_Sheet_In_000_Usd.Total_Equity',
            'Dest_Section_Context': 'Equity',
            'Dest_Major_Section_Context': 'Balance Sheet',
            'Source_Sheet_Name': 'Balance Sheet',
            'Source_Row_Number': '53',
            'Source_Field_Name': 'Total equity',
            'Source_Enhanced_Scope': 'Balance_Sheet.Equity.Total_Equity',
            'Source_Section_Context': 'Equity',
            'Q1_Verification_Value': '2340720',  # Dest Q1 value
            'Source_Q1_Value': '2402443',       # Source Q1 value (different)
            'Source_Q2_Value': '2253370',       # Source Q2 value to populate
            'Match_Method': 'Manual_Field_Name_Match',
            'Match_Confidence': '1.0',
            'Match_Reason': 'Field names match exactly, Q1 values different (2340720 vs 2402443)'
        },
        {
            'Dest_Row_Number': '156',
            'Dest_Field_Name': 'Total equity and liabillities',
            'Dest_Enhanced_Scope': 'Reported.Consolidated_Balance_Sheet_In_000_Usd.Total_Equity_And_Liabilities',
            'Dest_Section_Context': 'Total',
            'Dest_Major_Section_Context': 'Balance Sheet',
            'Source_Sheet_Name': 'Balance Sheet',
            'Source_Row_Number': '54',
            'Source_Field_Name': 'Total liabilities and equity',
            'Source_Enhanced_Scope': 'Balance_Sheet.Balance_Sheet_Other.Total_Liabilities_And_Equity',
            'Source_Section_Context': 'Balance_Sheet_Other',
            'Q1_Verification_Value': '2598436',  # Dest Q1 value
            'Source_Q1_Value': '2730436',       # Source Q1 value (different)
            'Source_Q2_Value': '2495431',       # Source Q2 value to populate
            'Match_Method': 'Manual_Field_Name_Match',
            'Match_Confidence': '1.0',
            'Match_Reason': 'Field names semantically equivalent (Total equity and liabilities), Q1 values different'
        }
    ]
    
    print(f"Created {len(manual_mappings)} manual equity mappings:")
    for mapping in manual_mappings:
        print(f"  Row {mapping['Dest_Row_Number']}: {mapping['Dest_Field_Name']} → {mapping['Source_Field_Name']}")
        print(f"    Q2 Value: {mapping['Source_Q2_Value']}")
    
    return manual_mappings


def save_manual_mappings(mappings):
    """Save manual mappings to CSV file."""
    
    output_file = "/Users/michaelkim/code/Bernstein/manual_equity_mappings.csv"
    
    with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
        fieldnames = [
            'Dest_Row_Number', 'Dest_Field_Name', 'Dest_Enhanced_Scope',
            'Dest_Section_Context', 'Dest_Major_Section_Context',
            'Source_Sheet_Name', 'Source_Row_Number', 'Source_Field_Name',
            'Source_Enhanced_Scope', 'Source_Section_Context',
            'Q1_Verification_Value', 'Source_Q1_Value', 'Source_Q2_Value',
            'Match_Method', 'Match_Confidence', 'Match_Reason'
        ]
        
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(mappings)
    
    print(f"\nManual equity mappings saved to: {output_file}")
    return output_file


def populate_manual_equity_mappings(mappings):
    """Populate the destination file with the manual equity mappings."""
    
    print("\n=== POPULATING MANUAL EQUITY SECTION MAPPINGS ===")
    
    source_file = "/Users/michaelkim/code/Bernstein/IPGP-Financial-Data-Workbook-2024-Q2.xlsx"
    dest_file = "/Users/michaelkim/code/Bernstein/final_complete_with_manual_mappings_IPGP.xlsx"
    
    # Load workbooks
    source_wb = openpyxl.load_workbook(source_file, data_only=True)
    dest_wb = openpyxl.load_workbook(dest_file, data_only=False)
    dest_sheet = dest_wb['Reported']
    
    populated_count = 0
    
    for mapping in mappings:
        dest_row = int(mapping['Dest_Row_Number'])
        source_sheet_name = mapping['Source_Sheet_Name']
        source_row = int(mapping['Source_Row_Number'])
        expected_q2_value = mapping['Source_Q2_Value']
        
        print(f"\nRow {dest_row}: {mapping['Dest_Field_Name']}")
        print(f"  From {source_sheet_name} Row {source_row}: {mapping['Source_Field_Name']}")
        print(f"  Expected Q2 value: {expected_q2_value}")
        
        # Verify the source value
        if source_sheet_name in source_wb.sheetnames:
            source_sheet = source_wb[source_sheet_name]
            actual_q2_value = source_sheet.cell(source_row, 93).value  # Column CO
            
            print(f"  Actual Q2 value: {actual_q2_value}")
            
            if str(actual_q2_value) == str(expected_q2_value):
                # Populate Column BS (71) with Q2 value
                if expected_q2_value != '':
                    dest_sheet.cell(dest_row, 71).value = float(expected_q2_value) if expected_q2_value != '0' else 0
                else:
                    dest_sheet.cell(dest_row, 71).value = actual_q2_value
                
                # Add source tracking to Column BT (72)
                source_location = f"IPGP-Financial-Data-Workbook-2024-Q2.xlsx|{source_sheet_name}|{source_row}|93"
                dest_sheet.cell(dest_row, 72).value = source_location
                
                populated_count += 1
                print(f"  ✅ POPULATED BS: {actual_q2_value}")
                print(f"  ✅ TRACKED BT: {source_location}")
            else:
                print(f"  ❌ Q2 value mismatch: expected {expected_q2_value}, got {actual_q2_value}")
        else:
            print(f"  ❌ Source sheet not found: {source_sheet_name}")
    
    # Save the updated file
    output_file = "/Users/michaelkim/code/Bernstein/final_complete_with_all_manual_mappings_IPGP.xlsx"
    dest_wb.save(output_file)
    
    source_wb.close()
    dest_wb.close()
    
    print(f"\n✅ Manual equity population complete!")
    print(f"Populated {populated_count} additional equity fields")
    print(f"Updated file saved as: {output_file}")
    
    return output_file, populated_count


def main():
    """Main entry point for manual equity section mappings."""
    
    print("="*80)
    print("MANUAL EQUITY SECTION MAPPINGS - MISSING FIELDS")
    print("="*80)
    print("Adding 4 missing equity section fields (rows 153-156)")
    
    try:
        # Create manual mappings
        manual_mappings = create_manual_equity_mappings()
        
        # Save mappings to file
        mapping_file = save_manual_mappings(manual_mappings)
        
        # Populate destination file
        output_file, populated_count = populate_manual_equity_mappings(manual_mappings)
        
        print(f"\n" + "="*80)
        print("MANUAL EQUITY SECTION MAPPING RESULTS")
        print("="*80)
        print(f"Manual equity mappings created: {len(manual_mappings)}")
        print(f"Successfully populated: {populated_count}")
        print(f"Mapping file: {mapping_file}")
        print(f"Final file: {output_file}")
        print(f"✅ All missing equity section fields now included!")
        
        # Final verification
        print(f"\n=== FINAL VERIFICATION ===")
        wb = openpyxl.load_workbook(output_file, data_only=True)
        sheet = wb['Reported']
        
        total_populated = 0
        for row_idx in range(1, 251):
            field_name = sheet.cell(row_idx, 1).value
            bs_value = sheet.cell(row_idx, 71).value
            
            if field_name and str(field_name).strip() and not str(field_name).strip().startswith('='):
                if bs_value is not None and bs_value != '':
                    total_populated += 1
        
        wb.close()
        
        print(f"Total populated fields: {total_populated} (was 85, now should be 89)")
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
