#!/usr/bin/env python3
"""
Comprehensive Q1 Verification - All Financial Statements

Uses all enhanced mappings (Key Metrics, Income Statement, Balance Sheet, Cash Flows)
to find Q1 verification matches for all empty destination fields.

Author: AI Assistant
Date: October 2025
"""

import openpyxl
import pandas as pd
import csv
from pathlib import Path
from typing import Dict, List, Optional


def round_for_comparison(value, precision=6):
    """Round value for comparison, handling None values."""
    if value is None:
        return None
    try:
        return round(float(value), precision)
    except:
        return value


def load_all_source_enhanced_mappings() -> Dict[str, List[Dict]]:
    """Load enhanced mappings from all source sheets."""
    
    source_mappings = {}
    
    # Define source files
    source_files = {
        'Key Metrics': '/Users/michaelkim/code/Bernstein/final_improved_key_metrics_mapping.csv',
        'Income Statement': '/Users/michaelkim/code/Bernstein/income_statement_enhanced_mapping.csv',
        'Balance Sheet': '/Users/michaelkim/code/Bernstein/balance_sheet_enhanced_mapping.csv',
        'Cash Flows': '/Users/michaelkim/code/Bernstein/cash_flows_enhanced_mapping.csv'
    }
    
    print("=== LOADING ALL SOURCE ENHANCED MAPPINGS ===")
    
    for sheet_name, file_path in source_files.items():
        if Path(file_path).exists():
            df = pd.read_csv(file_path)
            fields = []
            
            for _, row in df.iterrows():
                field_info = {
                    'sheet_name': sheet_name,
                    'row_number': row['Row_Number'],
                    'original_field_name': row['Original_Field_Name'],
                    'enhanced_scoped_name': row['Enhanced_Scoped_Name'],
                    'section_context': row['Section_Context'] if pd.notna(row['Section_Context']) else '',
                    'q1_2024_value': row['Q1_2024_Value'] if pd.notna(row['Q1_2024_Value']) else None,
                    'q2_2024_value': row['Q2_2024_Value'] if pd.notna(row['Q2_2024_Value']) else None,
                    'q1_rounded': round_for_comparison(row['Q1_2024_Value'] if pd.notna(row['Q1_2024_Value']) else None)
                }
                fields.append(field_info)
            
            source_mappings[sheet_name] = fields
            print(f"  {sheet_name}: {len(fields)} fields")
        else:
            print(f"  {sheet_name}: FILE NOT FOUND - {file_path}")
    
    total_source_fields = sum(len(fields) for fields in source_mappings.values())
    print(f"Total source fields loaded: {total_source_fields}")
    
    return source_mappings


def load_destination_data() -> Dict[int, Dict]:
    """Load destination field information and Q1 values."""
    
    print("\n=== LOADING DESTINATION DATA ===")
    
    # Load destination enhanced mapping
    dest_file = "/Users/michaelkim/code/Bernstein/reported_tab_comprehensive_mapping.csv"
    dest_scoping = {}
    
    if Path(dest_file).exists():
        dest_df = pd.read_csv(dest_file)
        for _, row in dest_df.iterrows():
            row_num = row['Row_Number']
            dest_scoping[row_num] = {
                'original_field_name': row['Original_Field_Name'],
                'enhanced_scoped_name': row['Enhanced_Scoped_Name'],
                'major_section_context': row['Major_Section_Context'] if pd.notna(row['Major_Section_Context']) else '',
                'section_context': row['Section_Context'] if pd.notna(row['Section_Context']) else ''
            }
        print(f"Loaded destination scoping: {len(dest_scoping)} fields")
    
    # Load destination Q1 values
    dest_file_path = "/Users/michaelkim/code/Bernstein/20240725_IPGP.US-IPG Photonics.xlsx"
    dest_wb = openpyxl.load_workbook(dest_file_path, data_only=True)
    dest_sheet = dest_wb['Reported']
    
    dest_q1_data = {}
    for row_idx in range(1, dest_sheet.max_row + 1):
        q1_value = dest_sheet.cell(row_idx, 70).value  # Column BR
        bs_value = dest_sheet.cell(row_idx, 71).value  # Column BS (current population)
        
        if q1_value is not None and row_idx in dest_scoping:
            dest_q1_data[row_idx] = {
                'original': q1_value,
                'rounded': round_for_comparison(q1_value),
                'current_bs_value': bs_value,
                'is_empty': bs_value is None or bs_value == ''
            }
    
    dest_wb.close()
    print(f"Loaded destination Q1 data: {len(dest_q1_data)} rows")
    
    # Count empty fields
    empty_count = sum(1 for data in dest_q1_data.values() if data['is_empty'])
    print(f"Empty fields needing population: {empty_count}")
    
    return dest_scoping, dest_q1_data


def perform_comprehensive_q1_verification(source_mappings: Dict[str, List[Dict]], 
                                        dest_scoping: Dict[int, Dict], 
                                        dest_q1_data: Dict[int, Dict]) -> List[Dict]:
    """Perform comprehensive Q1 verification across all source sheets."""
    
    print("\n=== PERFORMING COMPREHENSIVE Q1 VERIFICATION ===")
    
    # Create source Q1 lookup for fast matching
    source_q1_lookup = {}
    for sheet_name, fields in source_mappings.items():
        for field in fields:
            if field['q1_rounded'] is not None:
                q1_key = field['q1_rounded']
                if q1_key not in source_q1_lookup:
                    source_q1_lookup[q1_key] = []
                source_q1_lookup[q1_key].append(field)
    
    print(f"Built Q1 lookup with {len(source_q1_lookup)} unique Q1 values")
    
    matches = []
    used_source_combinations = set()  # Track (sheet, row) to avoid duplicates
    
    # Focus on empty destination fields
    empty_dest_fields = {row: data for row, data in dest_q1_data.items() 
                        if data['is_empty'] and row in dest_scoping}
    
    print(f"Processing {len(empty_dest_fields)} empty destination fields...")
    
    for dest_row, dest_q1_info in empty_dest_fields.items():
        dest_field_info = dest_scoping[dest_row]
        dest_q1_rounded = dest_q1_info['rounded']
        
        print(f"\nDEST Row {dest_row}: {dest_field_info['original_field_name']}")
        print(f"  Q1 value: {dest_q1_info['original']} (rounded: {dest_q1_rounded})")
        
        # Find source fields with matching Q1 value
        if dest_q1_rounded in source_q1_lookup:
            source_candidates = source_q1_lookup[dest_q1_rounded]
            
            best_match = None
            best_score = 0
            
            for source_field in source_candidates:
                source_key = (source_field['sheet_name'], source_field['row_number'])
                
                # Skip if already used (prefer unique mappings)
                if source_key in used_source_combinations:
                    continue
                
                # Calculate matching score
                score = 1.0  # Base Q1 match score
                
                # Boost score for field name similarity
                dest_name = dest_field_info['original_field_name'].lower()
                source_name = source_field['original_field_name'].lower()
                
                if dest_name == source_name:
                    score += 0.5
                elif any(word in source_name for word in dest_name.split() if len(word) > 3):
                    score += 0.3
                
                # Boost score for section context similarity
                dest_section = dest_field_info['section_context'].lower()
                source_section = source_field['section_context'].lower()
                
                if dest_section and source_section and dest_section == source_section:
                    score += 0.2
                
                # Prefer fields with Q2 data
                if source_field['q2_2024_value'] is not None:
                    score += 0.1
                
                if score > best_score:
                    best_score = score
                    best_match = source_field
            
            if best_match:
                source_key = (best_match['sheet_name'], best_match['row_number'])
                used_source_combinations.add(source_key)
                
                match = {
                    'Dest_Row_Number': dest_row,
                    'Dest_Field_Name': dest_field_info['original_field_name'],
                    'Dest_Enhanced_Scope': dest_field_info['enhanced_scoped_name'],
                    'Dest_Section_Context': dest_field_info['section_context'],
                    'Dest_Major_Section_Context': dest_field_info['major_section_context'],
                    
                    'Source_Sheet_Name': best_match['sheet_name'],
                    'Source_Row_Number': best_match['row_number'],
                    'Source_Field_Name': best_match['original_field_name'],
                    'Source_Enhanced_Scope': best_match['enhanced_scoped_name'],
                    'Source_Section_Context': best_match['section_context'],
                    
                    'Q1_Verification_Value': dest_q1_info['original'],
                    'Q1_Verification_Rounded': dest_q1_rounded,
                    'Source_Q1_Rounded': best_match['q1_rounded'],
                    'Source_Q2_Value': best_match['q2_2024_value'],
                    'Match_Method': 'Comprehensive_Q1_Verification',
                    'Match_Confidence': best_score,
                    'Match_Score': best_score
                }
                
                matches.append(match)
                
                print(f"  ✅ MATCHED to {best_match['sheet_name']} Row {best_match['row_number']}: {best_match['original_field_name']}")
                print(f"    Q1 verification: {dest_q1_rounded} = {best_match['q1_rounded']} ✓")
                print(f"    Q2 available: {best_match['q2_2024_value']}")
                print(f"    Match score: {best_score:.2f}")
            else:
                print(f"  ❌ Q1 match found but no unused source field available")
        else:
            print(f"  ❌ No Q1 match found for rounded value: {dest_q1_rounded}")
    
    print(f"\nComprehensive Q1 verification complete: {len(matches)} new matches found")
    return matches


def main():
    """Main entry point for comprehensive Q1 verification."""
    
    output_file = "/Users/michaelkim/code/Bernstein/comprehensive_q1_verified_mapping.csv"
    
    print("="*80)
    print("COMPREHENSIVE Q1 VERIFICATION - ALL FINANCIAL STATEMENTS")
    print("="*80)
    print("Finding additional matches for empty destination fields")
    print(f"Output file: {output_file}")
    
    try:
        # Load all source enhanced mappings
        source_mappings = load_all_source_enhanced_mappings()
        
        if not source_mappings:
            print("ERROR: No source mappings loaded")
            return
        
        # Load destination data
        dest_scoping, dest_q1_data = load_destination_data()
        
        # Perform comprehensive Q1 verification
        new_matches = perform_comprehensive_q1_verification(source_mappings, dest_scoping, dest_q1_data)
        
        # Save results
        if new_matches:
            with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = [
                    'Dest_Row_Number', 'Dest_Field_Name', 'Dest_Enhanced_Scope',
                    'Dest_Section_Context', 'Dest_Major_Section_Context',
                    'Source_Sheet_Name', 'Source_Row_Number', 'Source_Field_Name',
                    'Source_Enhanced_Scope', 'Source_Section_Context',
                    'Q1_Verification_Value', 'Q1_Verification_Rounded',
                    'Source_Q1_Rounded', 'Source_Q2_Value',
                    'Match_Method', 'Match_Confidence', 'Match_Score'
                ]
                
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(new_matches)
            
            print(f"\nComprehensive mapping saved to: {output_file}")
            print(f"New matches found: {len(new_matches)}")
            
            # Show breakdown by source sheet
            sheet_breakdown = {}
            for match in new_matches:
                sheet = match['Source_Sheet_Name']
                sheet_breakdown[sheet] = sheet_breakdown.get(sheet, 0) + 1
            
            print(f"\nNew matches by source sheet:")
            for sheet, count in sorted(sheet_breakdown.items()):
                print(f"  {sheet}: {count} matches")
            
            print(f"\n✅ SUCCESS: Found {len(new_matches)} additional matches!")
        else:
            print(f"\n❌ No additional matches found")
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
