#!/usr/bin/env python3
"""
Create All Enhanced Mappings

Creates enhanced scoped mappings for all source sheets:
- Income Statement
- Balance Sheet  
- Cash Flows

Author: AI Assistant
Date: October 2025
"""

import openpyxl
import pandas as pd
import csv
from pathlib import Path
from typing import Dict, List, Optional


def clean_field_name(field_name) -> str:
    """Clean and normalize field names for enhanced scoping."""
    if not field_name or pd.isna(field_name):
        return ""
    
    # Convert to string and strip
    cleaned = str(field_name).strip()
    
    # Remove special characters and replace with underscores
    cleaned = cleaned.replace('(', '').replace(')', '').replace(',', '').replace('&', 'And')
    cleaned = cleaned.replace(' ', '_').replace('-', '_').replace('/', '_')
    cleaned = cleaned.replace('__', '_').replace('___', '_')
    
    # Remove trailing underscores
    cleaned = cleaned.strip('_')
    
    return cleaned


def determine_section_context(row_idx: int, field_name: str, sheet_name: str) -> str:
    """Determine section context based on sheet and field characteristics."""
    
    if sheet_name == "Income Statement":
        if "revenue" in field_name.lower() or "sales" in field_name.lower():
            return "Revenue"
        elif "cost" in field_name.lower():
            return "Cost_Of_Sales"
        elif "gross" in field_name.lower():
            return "Gross_Profit"
        elif "operating" in field_name.lower():
            return "Operating"
        elif "interest" in field_name.lower():
            return "Interest"
        elif "tax" in field_name.lower():
            return "Tax"
        elif "income" in field_name.lower() or "earning" in field_name.lower():
            return "Net_Income"
        elif "share" in field_name.lower() or "per share" in field_name.lower():
            return "Earnings_Per_Share"
        else:
            return "Income_Statement_Other"
    
    elif sheet_name == "Balance Sheet":
        if "cash" in field_name.lower():
            return "Cash_And_Equivalents"
        elif "receivable" in field_name.lower():
            return "Receivables"
        elif "inventor" in field_name.lower():
            return "Inventory"
        elif "investment" in field_name.lower():
            return "Investments"
        elif "property" in field_name.lower() or "equipment" in field_name.lower():
            return "Property_Plant_Equipment"
        elif "goodwill" in field_name.lower():
            return "Goodwill"
        elif "intangible" in field_name.lower():
            return "Intangible_Assets"
        elif "asset" in field_name.lower():
            return "Assets"
        elif "payable" in field_name.lower():
            return "Payables"
        elif "debt" in field_name.lower():
            return "Debt"
        elif "liabilit" in field_name.lower():
            return "Liabilities"
        elif "equity" in field_name.lower() or "stock" in field_name.lower():
            return "Equity"
        else:
            return "Balance_Sheet_Other"
    
    elif sheet_name == "Cash Flows":
        if row_idx <= 20:
            return "Operating_Activities"
        elif row_idx <= 35:
            return "Investing_Activities"
        elif row_idx <= 50:
            return "Financing_Activities"
        else:
            return "Cash_Flow_Other"
    
    return "Unknown"


def create_enhanced_mapping_for_sheet(sheet_name: str, source_file: str) -> List[Dict]:
    """Create enhanced mapping for a specific sheet."""
    
    print(f"\n=== CREATING ENHANCED MAPPING FOR {sheet_name} ===")
    
    try:
        wb = openpyxl.load_workbook(source_file, data_only=True)
        
        if sheet_name not in wb.sheetnames:
            print(f"ERROR: Sheet '{sheet_name}' not found in {source_file}")
            return []
        
        sheet = wb[sheet_name]
        enhanced_fields = []
        
        print(f"Processing sheet: {sheet_name}")
        print(f"Max row: {sheet.max_row}")
        
        for row_idx in range(1, min(sheet.max_row + 1, 100)):  # Limit to first 100 rows
            # Column A = Field Name
            field_name_cell = sheet.cell(row_idx, 1)
            field_name = field_name_cell.value
            
            if not field_name or pd.isna(field_name) or str(field_name).strip() == "":
                continue
            
            field_name_str = str(field_name).strip()
            
            # Skip header rows or irrelevant rows
            if (field_name_str.startswith('=') or 
                field_name_str.lower() in ['', 'nan', 'none'] or
                len(field_name_str) < 2):
                continue
            
            # Get Q1 and Q2 values
            q1_value = sheet.cell(row_idx, 70).value  # Column BR
            q2_value = sheet.cell(row_idx, 93).value  # Column CO
            
            # Create section context
            section_context = determine_section_context(row_idx, field_name_str, sheet_name)
            
            # Create enhanced scoped name
            cleaned_field = clean_field_name(field_name_str)
            cleaned_section = clean_field_name(section_context)
            enhanced_scoped_name = f"{sheet_name.replace(' ', '_')}.{cleaned_section}.{cleaned_field}"
            
            enhanced_field = {
                'Row_Number': row_idx,
                'Original_Field_Name': field_name_str,
                'Enhanced_Scoped_Name': enhanced_scoped_name,
                'Section_Context': section_context,
                'Major_Section_Context': sheet_name,
                'Q1_2024_Value': q1_value if q1_value is not None else '',
                'Q2_2024_Value': q2_value if q2_value is not None else '',
                'Has_Q1_Data': q1_value is not None,
                'Has_Q2_Data': q2_value is not None
            }
            
            enhanced_fields.append(enhanced_field)
            
            if len(enhanced_fields) <= 10:  # Show first 10
                print(f"  Row {row_idx}: {field_name_str} → {enhanced_scoped_name}")
        
        wb.close()
        print(f"Created enhanced mapping: {len(enhanced_fields)} fields")
        return enhanced_fields
        
    except Exception as e:
        print(f"ERROR processing {sheet_name}: {e}")
        return []


def main():
    """Create enhanced mappings for all financial statement sheets."""
    
    source_file = "/Users/michaelkim/code/Bernstein/IPGP-Financial-Data-Workbook-2024-Q2.xlsx"
    
    sheets_to_process = [
        "Income Statement",
        "Balance Sheet", 
        "Cash Flows"
    ]
    
    print("="*80)
    print("CREATING ALL ENHANCED MAPPINGS")
    print("="*80)
    print(f"Source file: {source_file}")
    
    for sheet_name in sheets_to_process:
        enhanced_fields = create_enhanced_mapping_for_sheet(sheet_name, source_file)
        
        if enhanced_fields:
            # Save to CSV
            output_file = f"/Users/michaelkim/code/Bernstein/{sheet_name.lower().replace(' ', '_')}_enhanced_mapping.csv"
            
            with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = [
                    'Row_Number', 'Original_Field_Name', 'Enhanced_Scoped_Name',
                    'Section_Context', 'Major_Section_Context', 
                    'Q1_2024_Value', 'Q2_2024_Value', 'Has_Q1_Data', 'Has_Q2_Data'
                ]
                
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(enhanced_fields)
            
            print(f"✅ Saved: {output_file} ({len(enhanced_fields)} fields)")
        else:
            print(f"❌ No fields created for {sheet_name}")
    
    print(f"\n" + "="*80)
    print("ENHANCED MAPPING CREATION COMPLETE")
    print("="*80)


if __name__ == "__main__":
    main()
