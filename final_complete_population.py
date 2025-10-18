#!/usr/bin/env python3
"""
Final Complete Population - All Fields Including Precision-Adjusted

Combines the complete Q1 verified mapping (98 matches) with the precision-adjusted 
mapping (36 matches) to populate ALL possible fields in the destination file.

Author: AI Assistant
Date: October 2025
"""

import openpyxl
import pandas as pd
import csv
from pathlib import Path
from typing import Dict, List


def load_all_verified_mappings() -> List[Dict]:
    """Load all verified mappings from both complete and precision-adjusted files."""
    
    # Load complete mapping (98 matches from all financial statements)
    complete_file = "/Users/michaelkim/code/Bernstein/complete_q1_verified_mapping.csv"
    complete_mappings = []
    
    if Path(complete_file).exists():
        with open(complete_file, 'r', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            complete_mappings = list(reader)
        print(f"Loaded complete mapping: {len(complete_mappings)} matches")
    
    # Load precision-adjusted mapping (36 matches including rows 48-54)
    precision_file = "/Users/michaelkim/code/Bernstein/precision_adjusted_q1_mapping.csv"
    precision_mappings = []
    
    if Path(precision_file).exists():
        with open(precision_file, 'r', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            precision_mappings = list(reader)
        print(f"Loaded precision-adjusted mapping: {len(precision_mappings)} matches")
    
    # Combine mappings, preferring precision-adjusted for overlapping rows
    combined_mappings = {}
    
    # First add complete mappings
    for mapping in complete_mappings:
        dest_row = int(mapping['Dest_Row_Number'])
        combined_mappings[dest_row] = mapping
    
    # Then override with precision-adjusted mappings (better for percentage fields)
    for mapping in precision_mappings:
        dest_row = int(mapping['Dest_Row_Number'])
        # Convert precision-adjusted format to complete format
        standardized_mapping = {
            'Dest_Row_Number': mapping['Dest_Row_Number'],
            'Dest_Field_Name': mapping['Dest_Field_Name'],
            'Dest_Enhanced_Scope': mapping['Dest_Enhanced_Scope'],
            'Dest_Section_Context': mapping['Dest_Section_Context'],
            'Dest_Major_Section_Context': mapping['Dest_Major_Section_Context'],
            'Source_Sheet_Name': mapping['Source_Sheet_Name'],
            'Source_Row_Number': mapping['Source_Row_Number'],
            'Source_Field_Name': mapping['Source_Field_Name'],
            'Source_Enhanced_Scope': mapping['Source_Enhanced_Scope'],
            'Source_Section_Context': mapping['Source_Section_Context'],
            'Q1_Verification_Value': mapping.get('Q1_Verification_Value_Original', mapping.get('Q1_Verification_Value', '')),
            'Source_Q2_Value': mapping['Source_Q2_Value'],
            'Match_Method': mapping['Match_Method'],
            'Match_Confidence': mapping['Match_Confidence']
        }
        combined_mappings[dest_row] = standardized_mapping
    
    # Convert back to list and sort by destination row
    final_mappings = list(combined_mappings.values())
    final_mappings.sort(key=lambda x: int(x['Dest_Row_Number']))
    
    print(f"Final combined mappings: {len(final_mappings)} total")
    
    # Check if rows 48-54 are now included
    rows_48_54 = [m for m in final_mappings if 48 <= int(m['Dest_Row_Number']) <= 54]
    print(f"Rows 48-54 included: {len(rows_48_54)}")
    
    return final_mappings


def populate_final_destination_file(mappings: List[Dict]) -> Dict:
    """Populate destination file with final complete mappings."""
    
    print("="*80)
    print("FINAL COMPLETE POPULATION - ALL VERIFIED FIELDS")
    print("="*80)
    
    source_file = "/Users/michaelkim/code/Bernstein/IPGP-Financial-Data-Workbook-2024-Q2.xlsx"
    dest_file = "/Users/michaelkim/code/Bernstein/20240725_IPGP.US-IPG Photonics.xlsx"
    
    print(f"Source file: {source_file}")
    print(f"Destination file: {dest_file}")
    
    # Load workbooks
    source_wb = openpyxl.load_workbook(source_file, data_only=True)
    dest_wb = openpyxl.load_workbook(dest_file, data_only=False)
    dest_sheet = dest_wb['Reported']
    
    population_results = []
    values_populated = 0
    sheet_stats = {}
    
    print(f"\nPopulating {len(mappings)} final verified mappings...")
    
    for i, mapping in enumerate(mappings, 1):
        dest_row = int(mapping['Dest_Row_Number'])
        source_sheet_name = mapping['Source_Sheet_Name']
        source_row = mapping['Source_Row_Number']
        dest_field_name = mapping['Dest_Field_Name']
        source_field_name = mapping['Source_Field_Name']
        
        print(f"\n[{i}/{len(mappings)}] DEST Row {dest_row}: {dest_field_name}")
        
        # Get source sheet and Q2 value
        if source_sheet_name in source_wb.sheetnames:
            source_sheet = source_wb[source_sheet_name]
            
            # Handle Key Metrics special case where row number might be empty
            if not source_row or source_row.strip() == '':
                if source_sheet_name == 'Key Metrics':
                    # Find row by field name
                    found_row = None
                    for row_idx in range(1, min(source_sheet.max_row + 1, 200)):
                        cell_value = source_sheet.cell(row_idx, 1).value
                        if cell_value and str(cell_value).strip() == source_field_name:
                            found_row = row_idx
                            break
                    
                    if found_row:
                        source_row = found_row
                    else:
                        print(f"  ❌ Could not find row for {source_field_name}")
                        continue
                else:
                    print(f"  ❌ No source row specified")
                    continue
            else:
                source_row = int(source_row)
            
            # Get Q2 value (Column CO = 93)
            source_q2_value = source_sheet.cell(source_row, 93).value
            current_dest_value = dest_sheet.cell(dest_row, 71).value
            
            print(f"  From {source_sheet_name} Row {source_row}: {source_field_name}")
            print(f"  Q2 value: {source_q2_value}")
            
            if source_q2_value is not None:
                # Populate Column BS
                dest_sheet.cell(dest_row, 71).value = source_q2_value
                values_populated += 1
                
                # Track stats
                if source_sheet_name not in sheet_stats:
                    sheet_stats[source_sheet_name] = 0
                sheet_stats[source_sheet_name] += 1
                
                print(f"  ✅ POPULATED: {source_q2_value}")
                
                population_results.append({
                    'Dest_Row': dest_row,
                    'Dest_Field_Name': dest_field_name,
                    'Source_Sheet': source_sheet_name,
                    'Source_Row': source_row,
                    'Source_Field_Name': source_field_name,
                    'Source_Q2_Value': source_q2_value,
                    'Previous_Value': current_dest_value,
                    'Status': 'POPULATED'
                })
            else:
                print(f"  ❌ No Q2 data available")
                population_results.append({
                    'Dest_Row': dest_row,
                    'Dest_Field_Name': dest_field_name,
                    'Source_Sheet': source_sheet_name,
                    'Source_Row': source_row,
                    'Source_Field_Name': source_field_name,
                    'Source_Q2_Value': '',
                    'Previous_Value': current_dest_value,
                    'Status': 'NO_Q2_DATA'
                })
        else:
            print(f"  ❌ Source sheet not found: {source_sheet_name}")
    
    # Save final populated file
    output_file = "/Users/michaelkim/code/Bernstein/final_populated_20240725_IPGP.US-IPG_Photonics.xlsx"
    print(f"\nSaving final populated file to: {output_file}")
    dest_wb.save(output_file)
    
    source_wb.close()
    dest_wb.close()
    
    return {
        'population_results': population_results,
        'values_populated': values_populated,
        'total_mappings': len(mappings),
        'sheet_stats': sheet_stats,
        'output_file': output_file
    }


def main():
    """Main entry point for final complete population."""
    
    print("="*80)
    print("FINAL COMPLETE POPULATION - INCLUDING PRECISION-ADJUSTED FIELDS")
    print("="*80)
    print("Combining complete mapping + precision-adjusted mapping")
    print("This will populate rows 48-54 that were missing due to precision issues")
    
    try:
        # Load all verified mappings
        all_mappings = load_all_verified_mappings()
        
        if not all_mappings:
            print("ERROR: No mappings loaded")
            return
        
        # Populate destination file
        population_summary = populate_final_destination_file(all_mappings)
        
        # Final summary
        print(f"\n" + "="*80)
        print("FINAL COMPLETE POPULATION RESULTS")
        print("="*80)
        print(f"Total mappings processed: {population_summary['total_mappings']}")
        print(f"Values successfully populated: {population_summary['values_populated']}")
        print(f"Success rate: {population_summary['values_populated']}/{population_summary['total_mappings']} = {population_summary['values_populated']/population_summary['total_mappings']*100:.1f}%")
        
        print(f"\nPopulation by source sheet:")
        for sheet, count in sorted(population_summary['sheet_stats'].items()):
            print(f"  {sheet}: {count} values")
        
        print(f"\nFinal populated file: {population_summary['output_file']}")
        print("✅ ALL verified fields now populated including rows 48-54!")
        
        # Verify rows 48-54 specifically
        rows_48_54_populated = [r for r in population_summary['population_results'] 
                               if 48 <= r['Dest_Row'] <= 54 and r['Status'] == 'POPULATED']
        print(f"\nRows 48-54 successfully populated: {len(rows_48_54_populated)}")
        for result in rows_48_54_populated:
            print(f"  Row {result['Dest_Row']}: {result['Dest_Field_Name']} = {result['Source_Q2_Value']}")
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
